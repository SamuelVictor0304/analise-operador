from pathlib import Path
from html import escape
from io import BytesIO
import math
import os
import re
import unicodedata

import altair as alt
import numpy as np
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components


BASE_DIR = Path(__file__).parent
EVENTOS_FILE = BASE_DIR / "Cobmais-Eventos-908-2026050417.xlsx"
RESULTADOS_FILE = BASE_DIR / "NOVA BASE RESULTADOS 2026.xlsm"
COLABORADORES_FILE = BASE_DIR / "Base de colaboradores.xlsx"
EXCLUDED_OPERATORS = {"samuel.levi"}
EXCLUDED_OPERATOR_PREFIXES = ()
DEFAULT_OPERATOR_GOAL = 200000
POST_REPOSSESSED_GOAL = 900000
SPECIAL_OPERATOR_GOALS = {"victor.lima": POST_REPOSSESSED_GOAL}
IGNORED_META_OPERATORS = {"luiz.mauro", "cecilia.bonfim", "edmilson.silva"}
ALWAYS_INCLUDED_NEGOTIATORS = {"gabriela.rodrigues1"}
DEFAULT_RESULT_MONTH = "AGOSTO"
GOAL_FALLBACK_SOURCE_MONTH = "JULHO"
GOAL_FALLBACK_TARGET_MONTH = "AGOSTO"
OPERATIONAL_BUSINESS_DAY_TOTALS = {
    (2026, 7): 23,
    (2026, 8): 21,
}
POSTGRES_DEFAULTS = {
    "host": "",
    "port": 5432,
    "database": "",
    "user": "",
    "password": "",
    "schema": "workplan",
    "table": "casos_workplan",
}


def stretch_altair_chart(chart):
    try:
        st.altair_chart(chart, width="stretch")
    except TypeError:
        st.altair_chart(chart, use_container_width=True)


def stretch_dataframe(df, **kwargs):
    try:
        st.dataframe(df, width="stretch", **kwargs)
    except TypeError:
        st.dataframe(df, use_container_width=True, **kwargs)


def html_block(html, height=None):
    if hasattr(st, "html"):
        st.html(html)
    else:
        components.html(html, height=height)


def latest_file(pattern):
    files = [p for p in BASE_DIR.glob(pattern) if not p.name.startswith("~$")]
    if not files:
        raise FileNotFoundError(f"Nenhum arquivo encontrado para o padrão: {pattern}")
    return max(files, key=lambda p: p.stat().st_mtime)


def file_version(path):
    path = Path(path)
    stat = path.stat()
    return (path.name, stat.st_size, stat.st_mtime_ns)


def data_file_versions():
    return (
        file_version(latest_file("Cobmais-Eventos-908-*.xlsx")),
        file_version(latest_file("Pesquisa-Cliente-908-*.xlsx")),
        file_version(RESULTADOS_FILE),
    )


st.set_page_config(
    page_title="Performance Operacional",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
    <style>
    div[data-testid="stMetric"] {
        overflow: visible;
    }
    .metric-card {
        min-height: 96px;
        border: 1px solid rgba(148, 163, 184, 0.24);
        border-radius: 8px;
        padding: 12px 14px;
        background: rgba(15, 23, 42, 0.18);
    }
    .metric-card__label {
        margin-bottom: 8px;
        color: rgba(255, 255, 255, 0.82);
        font-size: 0.82rem;
        font-weight: 700;
        line-height: 1.2;
    }
    .metric-card__value {
        color: #ffffff;
        font-size: 1.48rem;
        font-weight: 650;
        line-height: 1.18;
        white-space: normal;
        overflow-wrap: anywhere;
        word-break: break-word;
    }
    .metric-card--compact .metric-card__value {
        font-size: 1.28rem;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


CORP_PALETTE = ["#213547", "#2f6f73", "#b7791f", "#6b7280", "#8a3ffc", "#bf616a"]
MONTH_NAMES_PT = {
    1: "JANEIRO",
    2: "FEVEREIRO",
    3: "MARÇO",
    4: "ABRIL",
    5: "MAIO",
    6: "JUNHO",
    7: "JULHO",
    8: "AGOSTO",
    9: "SETEMBRO",
    10: "OUTUBRO",
    11: "NOVEMBRO",
    12: "DEZEMBRO",
}

FIELD_HELP = {
    "OPERADOR": ("Operador", "Negociador responsável pelo evento ou acordo."),
    "FAIXA_ATRASO": ("Faixa de atraso", "Agrupamento do atraso/DPD do contrato em faixas gerenciais."),
    "SEGMENTO_DPD": ("Segmento DPD", "Classificação do contrato pela coluna Y/DPD Formula: POTLOSS 1-720, SALVAGE 721-1440 e SALVAGE + acima de 1440."),
    "prioridade_workplan": ("Prioridade", "Classificação do contrato no Workplan pelo score de recuperação."),
    "score_recuperacao": ("Score recuperação", "Score de priorização futura combinando segmento, valor, histórico de CPC e tempo sem contato. Contratos com pagamento ou acordo em aberto são excluídos da recomendação."),
    "probabilidade_recuperacao": ("Chance recuperação", "Probabilidade estimada de recuperação com base no histórico de contratos de perfil semelhante."),
    "valor_potencial": ("Valor potencial", "Valor financeiro em aberto do contrato usado para dimensionar a oportunidade."),
    "valor_esperado_recuperacao": ("Valor esperado", "Valor potencial ponderado pela chance e pela recuperação média histórica do perfil."),
    "base_probabilidade": ("Base probabilidade", "Nível de perfil histórico usado para estimar a chance de recuperação."),
    "contratos_hist": ("Contratos hist.", "Quantidade de contratos históricos usados no perfil de comparação."),
    "recuperacao_media_perfil": ("% recuperado perfil", "Valor pago dividido pelo valor negociado no perfil histórico semelhante."),
    "risco_quebra_perfil": ("Risco quebra perfil", "Acordos não pagos divididos por acordos históricos do perfil semelhante."),
    "motivo_abordagem": ("Motivo abordagem", "Motivo prático sugerido para abordar ou priorizar o cliente."),
    "motivo_priorizacao": ("Motivo priorização", "Principais fatores que aumentaram a prioridade do contrato."),
    "inadimplencia_precoce_tipo": ("FPD/EPD", "Classificacao por primeira parcela nao paga: FPD na 1a parcela, EPD na 5a parcela e demais casos separados."),
    "no_first_ins_unpaid": ("Primeira parcela nao paga", "Numero da primeira parcela em atraso/nao paga usado para identificar FPD e EPD."),
    "cpf_cnpj": ("CPF/CNPJ", "Documento do cliente conforme Workplan."),
    "PRODUTO": ("Produto", "Produto/carteira localizado no histórico de contratos."),
    "total_amount_due": ("Valor em aberto", "Valor total em aberto do contrato no Workplan."),
    "dpd": ("DPD", "Dias de atraso do contrato no Workplan."),
    "dias_sem_contato": ("Dias sem contato", "Dias desde o último contato registrado no Workplan ou último acionamento no histórico."),
    "acionamentos_hist": ("Acionamentos históricos", "Quantidade de acionamentos históricos localizados para o contrato."),
    "cpcs_hist": ("CPCs históricos", "Quantidade de CPCs históricos localizados para o contrato."),
    "acordos_hist": ("Acordos históricos", "Quantidade de acordos históricos localizados para o contrato."),
    "pagamentos_hist": ("Pagamentos históricos", "Quantidade de pagamentos históricos localizados para o contrato."),
    "carteira_elegivel": ("Carteira elegível", "Soma do valor em aberto dos contratos elegíveis no grupo."),
    "contratos_elegiveis": ("Contratos elegíveis", "Quantidade de contratos elegíveis no grupo."),
    "recuperacao_esperada": ("Recuperação esperada", "Carteira elegível ponderada pelas taxas históricas de contato, CPC, acordo, pagamento e percentual recuperado."),
    "recuperacao_esperada_pct_carteira": ("% esperado da carteira", "Recuperação esperada dividida pela carteira elegível."),
    "taxa_contato": ("Taxa acionamento -> contato", "Contatos com cliente divididos pelos acionamentos históricos no grupo."),
    "taxa_cpc": ("Taxa contato -> CPC", "CPCs históricos divididos pelos contatos com cliente no grupo."),
    "taxa_acordo": ("Taxa CPC -> acordo", "Acordos históricos divididos pelos CPCs históricos no grupo."),
    "taxa_pagamento": ("Taxa acordo -> pagamento", "Pagamentos históricos divididos pelos acordos históricos no grupo."),
    "percentual_medio_recuperado": ("% médio recuperado", "Valor pago histórico dividido pelo valor negociado histórico no grupo."),
    "base_taxas": ("Base das taxas", "Indica se as taxas foram calculadas pelo grupo ou pela média geral por baixa amostra."),
    "flag_cobravel": ("Cobravel", "Indica se o contrato está marcado como cobrável no Workplan."),
    "status_cpc": ("Status CPC", "Status de CPC disponível no Workplan."),
    "REGIÃO": ("Região", "Região cadastrada na base de resultados."),
    "clientes": ("Clientes", "Quantidade distinta de contratos/clientes no agrupamento."),
    "clientes_trabalhados": ("Clientes trabalhados", "Quantidade distinta de contratos acionados pelo operador."),
    "acionamentos": ("Acionamentos", "Total de eventos válidos, sem AUTO/importação por padrão."),
    "contatos_efetivos": ("Contatos efetivos", "Mesmo critério de CPC: eventos iniciados por 02, 03, 04 ou 05."),
    "contatos_cliente": ("Contatos cliente", "Eventos iniciados por 02 ou 03, contato direto com cliente."),
    "cpcs": ("CPCs", "Eventos produtivos iniciados por 02, 03, 04 ou 05."),
    "cpcs_unicos": ("CPCs únicos", "Clientes distintos por CPF/CNPJ com pelo menos um CPC por operador. Remove acionamentos repetidos do mesmo cliente."),
    "clientes_cpc": ("Clientes com CPC", "Clientes distintos por CPF/CNPJ que tiveram ao menos um CPC."),
    "contratos_cpc": ("Contratos com CPC", "Contratos distintos que tiveram ao menos um CPC."),
    "acordos": ("Acordos", "Quantidade de acordos localizados na base de resultados."),
    "pagamentos": ("Pagamentos", "Acordos com status PAGOU ou data de pagamento preenchida."),
    "acordos_sem_pagamento": ("Acordos sem pagamento", "Acordos sem status pago e sem data de pagamento."),
    "acordos_em_aberto": ("Acordos em aberto", "Acordos com status EM ABERTO na base de resultados."),
    "acordos_nao_pagou": ("Acordos não pagos", "Acordos com status NÃO PAGOU na base de resultados."),
    "pct_quebra": ("% quebras", "Impacto financeiro das quebras na meta: valor dos acordos não pagos dividido pela meta aplicável."),
    "tx_contato": ("Taxa de CPC", "CPCs divididos pelo total de acionamentos."),
    "tx_acordo": ("Taxa CPC -> acordo", "Acordos divididos pelo total de CPCs do operador."),
    "tx_acordo_cliente_cpc": ("Taxa contrato CPC -> acordo", "Acordos divididos pelos contratos distintos com CPC."),
    "tx_pagamento": ("Taxa acordo -> pagamento", "Pagamentos divididos por acordos."),
    "efetividade_pagamento": ("Efetividade pagamento", "Pagamentos divididos por pagamentos mais acordos quebrados/não pagos."),
    "tx_pagamento_cpc": ("Taxa CPC -> pagamento", "Pagamentos divididos pelo total de CPCs do operador."),
    "tx_sem_pagamento": ("Taxa sem pagamento", "Acordos sem pagamento divididos pelo total de acordos."),
    "tx_cpc_acordo": ("Taxa CPC -> acordo", "Contratos com CPC que geraram acordo, divididos pelos contratos com CPC."),
    "tx_cpc_pagamento": ("Taxa CPC -> pagamento", "Contratos com CPC que geraram pagamento, divididos pelos contratos com CPC."),
    "tx_cpc_unico_acordo": ("Taxa CPC único -> acordo", "Clientes únicos por CPF/CNPJ com CPC que geraram acordo, divididos pelos clientes únicos com CPC."),
    "tx_cpc_unico_pagamento": ("Taxa CPC único -> pagamento", "Clientes únicos por CPF/CNPJ com CPC que geraram pagamento, divididos pelos clientes únicos com CPC."),
    "tx_acordo_pagamento": ("Taxa acordo -> pagamento", "Pagamentos divididos pelos acordos originados em contratos com CPC."),
    "tx_acordo_sem_pagamento": ("Taxa acordo sem pagamento", "Acordos sem pagamento divididos pelos acordos originados em contratos com CPC."),
    "valor_negociado": ("Valor negociado", "Soma da coluna VALOR DO BANCO - META da base de resultados."),
    "valor_pago": ("Valor recebido", "Valor negociado dos acordos pagos; acordos não pagos entram como R$ 0,00."),
    "valor_em_aberto": ("Valor em aberto", "Valor negociado dos acordos com status EM ABERTO na base de resultados."),
    "valor_nao_pagou": ("Valor não pago", "Valor negociado dos acordos com status NÃO PAGOU."),
    "valor_quebra": ("Valor quebras", "Valor negociado dos acordos não pagos."),
    "ticket_medio": ("Ticket médio", "Valor negociado médio dos acordos."),
    "recuperacao": ("% recuperação", "Valor recebido dividido pelo valor negociado."),
    "score": ("Score", "Índice composto que pondera contato, acordo, pagamento, valor recebido e volume."),
    "meta_individual": ("Meta individual", "Meta mensal do negociador: R$ 200 mil; Victor Lima usa R$ 900 mil de pós retomado."),
    "atingimento_meta_individual": ("% meta individual", "Valor recebido dividido pela meta individual do negociador."),
    "pct_aberto_meta_individual": ("% aberto/meta individual", "Valor em aberto dividido pela meta individual do negociador."),
    "saldo_meta_individual": ("Saldo meta individual", "Valor recebido menos meta individual. Negativo indica falta para bater meta."),
    "meta_geral_escritorio": ("Meta geral escritório", "Meta geral do escritório para o mês, lida na aba METAS."),
    "participacao_meta_geral": ("% meta geral", "Quanto o operador contribuiu para a meta geral do escritório."),
    "quartil_meta_individual": ("Quartil meta individual", "Quartil do atingimento da meta individual. Q4 é o melhor grupo."),
    "quartil_meta_geral": ("Quartil meta geral", "Quartil da participação na meta geral do escritório. Q4 é o melhor grupo."),
    "diagnostico_meta": ("Diagnóstico meta", "Leitura gerencial combinando atingimento individual e contribuição na meta geral."),
    "quartil_cpc_acordo": ("Quartil CPC → acordo", "Quartil da taxa de conversão CPC único → acordo. Q4 é o melhor grupo. Mínimo de 3 CPCs únicos para entrar no ranking."),
    "quartil_cpc_pagamento": ("Quartil CPC → pagamento", "Quartil da taxa de conversão CPC único → pagamento. Q4 é o melhor grupo. Mínimo de 3 CPCs únicos para entrar no ranking."),
    "quartil_cpc_volume": ("Quartil volume CPC", "Quartil da quantidade de CPCs únicos gerados pelo operador. Q4 é o melhor grupo. Mínimo de 1 CPC único para entrar no ranking."),
    "diagnostico_cpc": ("Diagnóstico CPC", "Classificação da qualidade de conversão do CPC em pagamento: Alta conversão (≥20%), Boa conversão (≥10%), Atenção (≥5%) e Crítico (<5%). Sem base quando menos de 3 CPCs únicos."),
    "nome_colaborador": ("Nome colaborador", "Nome completo do colaborador conforme Base de colaboradores."),
    "base_colaborador": ("Base colaborador", "Aba/carteira da Base de colaboradores onde o login foi encontrado."),
    "cargo_colaborador": ("Cargo", "Cargo do colaborador conforme Base de colaboradores."),
    "negociador_cadastrado": ("Cadastro colaborador", "Indica se o operador está cadastrado como negociador na Base de colaboradores."),
}


def normalize_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def format_operator_label(operador):
    text = normalize_text(operador)
    if not text:
        return "-"
    if "." in text and text == text.lower():
        return " ".join(part.capitalize() for part in text.split("."))
    return text


def normalize_status(value):
    text = normalize_text(value).upper()
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def normalize_month_key(value):
    return normalize_status(value)


def carry_over_month_goal(goals, source_month=GOAL_FALLBACK_SOURCE_MONTH, target_month=GOAL_FALLBACK_TARGET_MONTH):
    source_key = normalize_month_key(source_month)
    target_key = normalize_month_key(target_month)
    if goals.get(target_key, 0) <= 0 and goals.get(source_key, 0) > 0:
        goals[target_key] = goals[source_key]
    return goals


def normalize_contract(value):
    if pd.isna(value):
        return np.nan
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    return digits or np.nan


def normalize_document(value):
    return normalize_contract(value)


def normalize_operator(value):
    text = normalize_text(value).lower()
    return text if text else np.nan


def is_excluded_operator(value):
    text = normalize_operator(value)
    if pd.isna(text):
        return False
    return text in EXCLUDED_OPERATORS or any(text.startswith(prefix) for prefix in EXCLUDED_OPERATOR_PREFIXES)


def atraso_faixa(days):
    if pd.isna(days):
        return "Sem atraso"
    days = float(days)
    if days <= 30:
        return "000-030"
    if days <= 60:
        return "031-060"
    if days <= 90:
        return "061-090"
    if days <= 120:
        return "091-120"
    if days <= 180:
        return "121-180"
    if days <= 360:
        return "181-360"
    return "361+"


def segmento_dpd(days):
    if pd.isna(days):
        return "Sem DPD"
    text = normalize_text(days).upper()
    if text:
        compact = re.sub(r"[^A-Z0-9+]", "", text)
        if compact == "POTLOSS":
            return "POTLOSS"
        if compact == "SALVAGE":
            return "SALVAGE"
        if compact in {"SALVAGE+", "SALVAGEPLUS"}:
            return "SALVAGE +"
    numeric_days = pd.to_numeric(days, errors="coerce")
    if pd.isna(numeric_days):
        return "Sem DPD"
    if numeric_days < 1:
        return "Sem DPD"
    if numeric_days <= 720:
        return "POTLOSS"
    if numeric_days <= 1440:
        return "SALVAGE"
    return "SALVAGE +"


def money_fmt(value):
    value = 0 if pd.isna(value) else float(value)
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def pct_fmt(value):
    value = 0 if pd.isna(value) or np.isinf(value) else float(value)
    return f"{value:.1%}".replace(".", ",")


def num_fmt(value):
    value = 0 if pd.isna(value) else float(value)
    return f"{value:,.0f}".replace(",", ".")


def safe_div(num, den):
    return np.where(den == 0, 0, num / den)


def scalar_safe_div(num, den):
    return 0 if pd.isna(den) or den == 0 else num / den


def normalized_upper(value):
    return normalize_text(value).upper()


def is_pos_retomado(value):
    text = normalized_upper(value)
    compact = re.sub(r"[^A-Z0-9]", "", text)
    return compact in {"POSRETOMADO", "POSTRETOMADO", "POSRETOMADA", "POSTRETOMADA"}


def parse_installment_number(value):
    if pd.isna(value):
        return np.nan
    text = normalize_text(value)
    if not text:
        return np.nan
    match = re.search(r"\d+", text)
    return float(match.group(0)) if match else np.nan


def early_default_type(value):
    installment = pd.to_numeric(value, errors="coerce")
    if pd.isna(installment) or installment <= 0:
        return "Demais"
    installment = int(installment)
    if installment == 1:
        return "FPD"
    if installment == 5:
        return "EPD"
    return "Demais"


def streamlit_secret_section(name):
    try:
        return st.secrets.get(name, {})
    except Exception:
        return {}


def streamlit_secret_value(name, default=""):
    try:
        return st.secrets.get(name, default)
    except Exception:
        return default


def config_value(name, secrets, default=""):
    return os.getenv(name) or streamlit_secret_value(name) or secrets.get(name.lower(), default)


def postgres_config():
    secrets = streamlit_secret_section("postgres")
    supabase_ref = config_value("SUPABASE_PROJECT_REF", secrets)
    supabase_password = config_value("SUPABASE_DB_PASSWORD", secrets)
    supabase_user = config_value("SUPABASE_DB_USER", secrets, "postgres")
    supabase_port = int(config_value("SUPABASE_DB_PORT", secrets, "6543"))
    supabase_host = config_value("SUPABASE_DB_HOST", secrets, "aws-1-sa-east-1.pooler.supabase.com")
    database_url = (
        ""
        if supabase_ref and supabase_password
        else os.getenv("SUPABASE_DB_URL")
        or os.getenv("DATABASE_URL")
        or streamlit_secret_value("SUPABASE_DB_URL")
        or streamlit_secret_value("supabase_db_url")
        or streamlit_secret_value("DATABASE_URL")
        or streamlit_secret_value("database_url")
        or secrets.get("database_url", secrets.get("url", ""))
    )
    has_connection_parts = any(
        [
            os.getenv("PGHOST"),
            os.getenv("PGDATABASE"),
            os.getenv("PGUSER"),
            os.getenv("PGPASSWORD"),
            secrets.get("host"),
            secrets.get("database"),
            secrets.get("user"),
            secrets.get("password"),
            supabase_ref,
            supabase_password,
        ]
    )
    if supabase_ref and supabase_password:
        host = supabase_host
        port = supabase_port
        database = "postgres"
        user = f"{supabase_user}.{supabase_ref}"
        password = supabase_password
    else:
        host = os.getenv("PGHOST", secrets.get("host", POSTGRES_DEFAULTS["host"]))
        port = int(os.getenv("PGPORT", secrets.get("port", POSTGRES_DEFAULTS["port"])))
        database = os.getenv("PGDATABASE", secrets.get("database", POSTGRES_DEFAULTS["database"]))
        user = os.getenv("PGUSER", secrets.get("user", POSTGRES_DEFAULTS["user"]))
        password = os.getenv("PGPASSWORD", secrets.get("password", POSTGRES_DEFAULTS["password"]))
    return {
        "database_url": database_url,
        "configured": bool(database_url or has_connection_parts),
        "host": host,
        "port": port,
        "database": database,
        "user": user,
        "password": password,
        "schema": os.getenv("PGSCHEMA", secrets.get("schema", POSTGRES_DEFAULTS["schema"])),
        "table": os.getenv("PGTABLE", secrets.get("table", POSTGRES_DEFAULTS["table"])),
    }


def quartile_label(series, higher_is_better=True, min_series=None, min_value=1):
    values = pd.to_numeric(series, errors="coerce")
    valid = values.notna()
    if min_series is not None:
        valid = valid & (pd.to_numeric(min_series, errors="coerce").fillna(0) >= min_value)

    labels = pd.Series("Sem base", index=series.index, dtype="object")
    if valid.sum() == 0:
        return labels

    ranked = values[valid].rank(method="average", pct=True)
    if not higher_is_better:
        ranked = 1 - ranked + (1 / valid.sum())

    labels.loc[ranked[ranked <= 0.25].index] = "Q1 - crítico"
    labels.loc[ranked[(ranked > 0.25) & (ranked <= 0.50)].index] = "Q2 - atenção"
    labels.loc[ranked[(ranked > 0.50) & (ranked <= 0.75)].index] = "Q3 - bom"
    labels.loc[ranked[ranked > 0.75].index] = "Q4 - destaque"
    return labels


def selected_months(resultados):
    return result_months_frame(resultados)["MES_RESULTADO"].tolist()


def result_months_frame(resultados):
    meses = (
        resultados[["MES_RESULTADO", "MES_NUM"]]
        .replace("", np.nan)
        .dropna(subset=["MES_RESULTADO"])
        .copy()
    )
    if meses.empty:
        return meses
    meses["MES_NUM"] = pd.to_numeric(meses["MES_NUM"], errors="coerce")
    meses["MES_KEY"] = meses["MES_RESULTADO"].map(normalize_month_key)
    meses = (
        meses.sort_values(["MES_NUM", "MES_RESULTADO"], na_position="last")
        .drop_duplicates(subset=["MES_KEY"], keep="first")
        .sort_values(["MES_NUM", "MES_RESULTADO"], na_position="last")
    )
    return meses


def selected_months_count(resultados):
    return max(len(selected_months(resultados)), 1)


def office_goal_for_resultados(resultados):
    return office_goal_for_months(selected_months(resultados))


def office_goal_for_months(months):
    metas_gerais = load_office_goals(file_version(RESULTADOS_FILE))
    month_keys = list(dict.fromkeys(normalize_month_key(mes) for mes in months if normalize_month_key(mes)))
    selected_goals = [metas_gerais.get(mes, 0) for mes in month_keys]
    positive_goals = [goal for goal in selected_goals if goal > 0]
    if positive_goals:
        return sum(positive_goals)

    available_goals = [goal for goal in metas_gerais.values() if goal > 0]
    if not available_goals:
        return 0
    return available_goals[-1]


def operator_goal_series(operadores, meses_count):
    operadores = operadores.fillna("")
    base_goal = operadores.map(SPECIAL_OPERATOR_GOALS).fillna(DEFAULT_OPERATOR_GOAL)
    base_goal = base_goal.where(~operadores.isin(IGNORED_META_OPERATORS), 0)
    return pd.Series(base_goal * meses_count, index=operadores.index)


@st.cache_data(show_spinner=False)
def load_collaborators():
    if not COLABORADORES_FILE.exists():
        return pd.DataFrame(columns=["OPERADOR", "nome_colaborador", "base_colaborador", "cargo_colaborador"])

    sheets = pd.read_excel(COLABORADORES_FILE, sheet_name=None)
    frames = []
    for sheet_name, df in sheets.items():
        df.columns = [normalize_text(c).upper() for c in df.columns]
        if "LOGIN COBMAIS" not in df.columns:
            continue
        if "ATIVO" in df.columns:
            df = df[df["ATIVO"].map(normalize_text).str.upper().eq("SIM")].copy()
        base = pd.DataFrame(
            {
                "OPERADOR": df["LOGIN COBMAIS"].map(normalize_operator),
                "nome_colaborador": df.get("NOME COLABORADOR", pd.Series(index=df.index, dtype=object)).map(normalize_text),
                "base_colaborador": sheet_name,
                "cargo_colaborador": df.get("CARGO", pd.Series(index=df.index, dtype=object)).map(normalize_text).str.upper(),
            }
        )
        base = base[base["OPERADOR"].notna()]
        base = base[~base["OPERADOR"].isin(["escritório", "escritorio"])]
        base = base[~base["OPERADOR"].map(is_excluded_operator)]
        frames.append(base)

    if not frames:
        return pd.DataFrame(columns=["OPERADOR", "nome_colaborador", "base_colaborador", "cargo_colaborador"])

    colaboradores = pd.concat(frames, ignore_index=True).drop_duplicates(["OPERADOR", "base_colaborador"])
    judicial = colaboradores[colaboradores["base_colaborador"].eq("JUDICIAL")].copy()
    if judicial.empty:
        judicial = colaboradores[colaboradores["cargo_colaborador"].eq("NEGOCIADOR")].copy()
    return judicial.drop_duplicates("OPERADOR")


@st.cache_data(show_spinner=False)
def load_office_goals(data_version):
    metas = pd.read_excel(RESULTADOS_FILE, sheet_name="METAS", header=None)
    header_idx = metas.index[metas.iloc[:, 0].astype(str).str.strip().str.upper().eq("NEGOCIADOR")]
    if len(header_idx) == 0:
        return {}

    header_row = header_idx[0]
    headers = metas.iloc[header_row].map(normalize_month_key).tolist()
    totals = metas.iloc[header_row + 1:].copy()
    total_rows = totals[totals.iloc[:, 0].astype(str).str.strip().str.upper().eq("TOTAL")]
    if total_rows.empty:
        return {}

    total_row = total_rows.iloc[0]
    goals = {}
    month_abbr = {
        "JANEIRO": "JAN",
        "FEVEREIRO": "FEV",
        "MARÇO": "MAR",
        "MARCO": "MAR",
        "ABRIL": "ABR",
        "MAIO": "MAI",
        "JUNHO": "JUN",
        "JULHO": "JUL",
        "AGOSTO": "AGO",
        "SETEMBRO": "SET",
        "OUTUBRO": "OUT",
        "NOVEMBRO": "NOV",
        "DEZEMBRO": "DEZ",
    }
    for month, abbr in month_abbr.items():
        target_col = None
        for i, header in enumerate(headers):
            if header == f"META {abbr}":
                target_col = i
                break
        if target_col is not None:
            goals[normalize_month_key(month)] = pd.to_numeric(total_row.iloc[target_col], errors="coerce")
    goals = {k: float(v) for k, v in goals.items() if pd.notna(v)}
    # Meta de agosto ainda nao cadastrada na aba METAS: reaproveita a meta de julho por enquanto.
    return carry_over_month_goal(goals)


@st.cache_data(show_spinner=False)
def load_office_received(data_version):
    metas = pd.read_excel(RESULTADOS_FILE, sheet_name="METAS", header=None)
    header_idx = metas.index[metas.iloc[:, 0].astype(str).str.strip().str.upper().eq("NEGOCIADOR")]
    if len(header_idx) == 0:
        return {}

    header_row = header_idx[0]
    headers = metas.iloc[header_row].map(normalize_text).str.upper().tolist()
    totals = metas.iloc[header_row + 1:].copy()
    total_rows = totals[totals.iloc[:, 0].astype(str).str.strip().str.upper().eq("TOTAL")]
    if total_rows.empty:
        return {}

    total_row = total_rows.iloc[0]
    received = {}
    months = [
        "JANEIRO",
        "FEVEREIRO",
        "MARÇO",
        "ABRIL",
        "MAIO",
        "JUNHO",
        "JULHO",
        "AGOSTO",
        "SETEMBRO",
        "OUTUBRO",
        "NOVEMBRO",
        "DEZEMBRO",
    ]
    for month in months:
        target_col = None
        for i, header in enumerate(headers):
            if header == month:
                target_col = i
                break
        if target_col is not None:
            received[month] = pd.to_numeric(total_row.iloc[target_col], errors="coerce")
    return {k: float(v) for k, v in received.items() if pd.notna(v)}


@st.cache_data(show_spinner=False)
def load_region_goals(data_version):
    metas = pd.read_excel(RESULTADOS_FILE, sheet_name="METAS", header=None)
    header_idx = metas.index[metas.iloc[:, 0].astype(str).str.strip().str.upper().eq("NEGOCIADOR")]
    if len(header_idx) == 0:
        return pd.DataFrame(columns=["REGIÃO", "MES_RESULTADO", "valor_pago", "meta_regiao"])

    header_row = header_idx[0]
    headers = metas.iloc[header_row].map(normalize_month_key).tolist()
    month_abbr = {
        "JANEIRO": "JAN",
        "FEVEREIRO": "FEV",
        "MARÇO": "MAR",
        "MARCO": "MAR",
        "ABRIL": "ABR",
        "MAIO": "MAI",
        "JUNHO": "JUN",
        "JULHO": "JUL",
        "AGOSTO": "AGO",
        "SETEMBRO": "SET",
        "OUTUBRO": "OUT",
        "NOVEMBRO": "NOV",
        "DEZEMBRO": "DEZ",
    }

    rows = []
    for row_idx in range(header_row + 1, len(metas)):
        row = metas.iloc[row_idx]
        regiao = normalize_text(row.iloc[0])
        regiao_key = normalize_month_key(regiao)
        if not regiao_key:
            break
        if regiao_key == "TOTAL":
            continue

        for month, abbr in month_abbr.items():
            valor_col = next((i for i, header in enumerate(headers) if header == month), None)
            meta_col = next((i for i, header in enumerate(headers) if header == f"META {abbr}"), None)
            valor = pd.to_numeric(row.iloc[valor_col], errors="coerce") if valor_col is not None else 0
            meta = pd.to_numeric(row.iloc[meta_col], errors="coerce") if meta_col is not None else 0
            rows.append(
                {
                    "REGIÃO": regiao,
                    "REGIAO_KEY": regiao_key,
                    "MES_RESULTADO": month,
                    "valor_pago": 0 if pd.isna(valor) else float(valor),
                    "meta_regiao": 0 if pd.isna(meta) else float(meta),
                }
            )

    df = pd.DataFrame(rows)
    if not df.empty:
        # Meta regional de agosto ainda nao cadastrada na aba METAS: reaproveita a meta de julho por enquanto.
        fonte = df.loc[df["MES_RESULTADO"] == GOAL_FALLBACK_SOURCE_MONTH].set_index("REGIAO_KEY")["meta_regiao"]
        alvo_mask = (df["MES_RESULTADO"] == GOAL_FALLBACK_TARGET_MONTH) & (df["meta_regiao"] <= 0)
        df.loc[alvo_mask, "meta_regiao"] = df.loc[alvo_mask, "REGIAO_KEY"].map(fonte).fillna(0)
    return df


def build_region_goal_map(months, resultados=None):
    region_goals = load_region_goals(file_version(RESULTADOS_FILE))
    if region_goals.empty:
        return region_goals

    month_keys = [normalize_month_key(month) for month in months]
    selected = region_goals[region_goals["MES_RESULTADO"].isin(month_keys)].copy()
    selected = selected[selected["meta_regiao"] > 0]
    if selected.empty:
        positive = region_goals[region_goals["meta_regiao"] > 0].copy()
        if positive.empty:
            return positive
        latest_month = positive["MES_RESULTADO"].drop_duplicates().iloc[-1]
        selected = positive[positive["MES_RESULTADO"].eq(latest_month)].copy()
        month_keys = [latest_month]

    df = (
        selected.groupby(["REGIÃO", "REGIAO_KEY"], dropna=False)
        .agg(meta_regiao=("meta_regiao", "sum"))
        .reset_index()
    )

    # valor_pago real vem do dataframe de resultados filtrado pelos meses; a aba
    # METAS so traz a meta cadastrada, entao usamos os pagamentos efetivos.
    df["valor_pago"] = 0.0
    if resultados is not None and not resultados.empty:
        pagos = resultados.copy()
        if "MES_RESULTADO" in pagos.columns and month_keys:
            pagos = pagos[pagos["MES_RESULTADO"].isin(month_keys)]
        if not pagos.empty and "REGIÃO" in pagos.columns and "VALOR_PAGO" in pagos.columns:
            pagos["REGIAO_KEY"] = pagos["REGIÃO"].map(normalize_month_key)
            soma = pagos.groupby("REGIAO_KEY", dropna=False)["VALOR_PAGO"].sum()
            df["valor_pago"] = df["REGIAO_KEY"].map(soma).fillna(0.0)

    df["pct_meta_regiao"] = safe_div(df["valor_pago"], df["meta_regiao"])
    return df.sort_values("pct_meta_regiao", ascending=False)


@st.cache_data(show_spinner=False)
def _load_brazil_regions_svg():
    path = Path(__file__).parent / "assets" / "brazil_regions.svg"
    try:
        return path.read_text(encoding="utf-8")
    except OSError:
        return ""


def region_meta_map(region_df, title):
    if region_df.empty:
        st.info("Sem metas regionais cadastradas na aba METAS para montar o mapa.")
        return

    # IBGE region IDs no SVG: 1=Norte, 2=Nordeste, 3=Sudeste, 4=Sul, 5=Centro-Oeste
    active_keys = {"CENTRO-OESTE", "SUDESTE", "SUL"}
    region_id = {"NORTE": "1", "NORDESTE": "2", "SUDESTE": "3", "SUL": "4", "CENTRO-OESTE": "5"}

    # Coordenadas (lon, lat) aproximadas do centroide de cada regiao
    centroids_lonlat = {
        "CENTRO-OESTE": (-53.0, -14.0),
        "SUDESTE": (-43.5, -20.0),
        "SUL": (-52.0, -27.5),
    }
    # viewBox do SVG IBGE: "-73.9833 -5.2718 39.1806 39.0157" (Y = -lat)
    vb_x, vb_y, vb_w, vb_h = -73.9833, -5.2718, 39.1806, 39.0157

    def to_pct(lon, lat):
        x_pct = (lon - vb_x) / vb_w * 100
        y_pct = (-lat - vb_y) / vb_h * 100
        return f"{x_pct:.2f}%", f"{y_pct:.2f}%"

    svg_raw = _load_brazil_regions_svg()
    if not svg_raw:
        st.warning("SVG do mapa do Brasil nao encontrado em assets/brazil_regions.svg.")
        return

    # Aplica cores: regioes ativas em verde-azulado, demais em cinza neutro
    color_active = "#2f6f73"
    color_inactive = "#5b6b73"
    stroke = "#102733"
    svg = svg_raw
    for key, rid in region_id.items():
        fill = color_active if key in active_keys else color_inactive
        svg = svg.replace(f'<path id="{rid}"', f'<path id="{rid}" fill="{fill}" stroke="{stroke}" stroke-width="600"')
    # garante que o svg externo preencha o container e seja responsivo
    svg = svg.replace("<svg ", '<svg style="width:100%;height:100%;display:block;" ', 1)
    svg = re.sub(r'width="\d+"\s*height="\d+"\s*', "", svg, count=1)

    cards = []
    for _, row in region_df.iterrows():
        key = row["REGIAO_KEY"]
        if key not in centroids_lonlat:
            continue
        lon, lat = centroids_lonlat[key]
        left, top = to_pct(lon, lat)
        label = escape(normalize_text(row["REGIÃO"]))
        pct = escape(pct_fmt(row["pct_meta_regiao"]))
        meta = escape(money_fmt(row["meta_regiao"]))
        received = escape(money_fmt(row["valor_pago"]))
        cards.append(
            f"""
            <div class="region-map__pin" style="left:{left};top:{top};"></div>
            <div class="region-map__card" style="left:{left};top:calc({top} + 14px);">
                <div class="region-map__label">{label}</div>
                <div class="region-map__pct">{pct}</div>
                <div class="region-map__sub">{received} / {meta}</div>
            </div>
            """
        )

    html_block(
        f"""
        <style>
        .region-map {{
            position: relative;
            padding: 6px 0 8px 0;
            background: transparent;
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
        }}
        .region-map__title {{
            color: inherit;
            font-weight: 800;
            font-size: 1rem;
            margin: 0 0 8px 2px;
        }}
        .region-map__stage {{
            position: relative;
            width: 100%;
            max-width: 560px;
            aspect-ratio: 1 / 1;
            margin: 0 auto;
        }}
        .region-map__stage svg {{
            width: 100%;
            height: 100%;
            display: block;
            filter: drop-shadow(0 4px 6px rgba(0,0,0,.18));
        }}
        .region-map__pin {{
            position: absolute;
            z-index: 3;
            width: 14px;
            height: 14px;
            border-radius: 50%;
            border: 3px solid #0e2230;
            background: #ffffff;
            transform: translate(-50%, -50%);
            box-shadow: 0 2px 4px rgba(0,0,0,.35);
        }}
        .region-map__pin::after {{
            content: "";
            position: absolute;
            left: 50%; top: 100%;
            width: 0; height: 0;
            border-left: 6px solid transparent;
            border-right: 6px solid transparent;
            border-top: 9px solid #0e2230;
            transform: translateX(-50%);
        }}
        .region-map__card {{
            position: absolute;
            z-index: 4;
            min-width: 118px;
            transform: translate(-50%, 0);
            text-align: center;
            filter: drop-shadow(4px 4px 0 rgba(0,0,0,.18));
        }}
        .region-map__label {{
            display: inline-block;
            padding: 4px 10px;
            background: #eef3f5;
            color: #36434d;
            font-size: .76rem;
            font-weight: 800;
            border-radius: 2px;
        }}
        .region-map__pct {{
            margin: 2px auto 0;
            padding: 6px 14px;
            width: max-content;
            min-width: 86px;
            background: rgba(14, 34, 48, .95);
            color: #ffffff;
            font-size: 1.35rem;
            font-weight: 900;
            line-height: 1;
            border-radius: 2px;
        }}
        .region-map__sub {{
            margin: 1px auto 0;
            padding: 4px 8px;
            width: max-content;
            max-width: 180px;
            background: rgba(14, 34, 48, .82);
            color: rgba(255,255,255,.85);
            font-size: .66rem;
            white-space: nowrap;
            border-radius: 2px;
        }}
        </style>
        <div class="region-map">
            <div class="region-map__title">{escape(title)}</div>
            <div class="region-map__stage">
                {svg}
                {''.join(cards)}
            </div>
        </div>
        """,
        height=620,
    )


def build_meta_analysis(operador_df, resultados, operadores_scope=None):
    colaboradores = load_collaborators()
    missing_negotiators = ALWAYS_INCLUDED_NEGOTIATORS - set(colaboradores["OPERADOR"].dropna())
    if missing_negotiators:
        colaboradores = pd.concat(
            [
                colaboradores,
                pd.DataFrame(
                    {
                        "OPERADOR": sorted(missing_negotiators),
                        "nome_colaborador": sorted(missing_negotiators),
                        "base_colaborador": "MANUAL",
                        "cargo_colaborador": "NEGOCIADOR",
                    }
                ),
            ],
            ignore_index=True,
        )
    if operadores_scope:
        colaboradores = colaboradores[colaboradores["OPERADOR"].isin(operadores_scope)].copy()
    meses = selected_months(resultados)
    meses_count = max(len(meses), 1)
    meta_geral = office_goal_for_months(meses)

    operadores_base = colaboradores[["OPERADOR"]].drop_duplicates()
    operadores_base = operadores_base[~operadores_base["OPERADOR"].isin(IGNORED_META_OPERATORS)].copy()
    df = operadores_base.merge(operador_df, on="OPERADOR", how="left")
    df = df.merge(colaboradores, on="OPERADOR", how="left")
    metric_cols = [
        "acionamentos",
        "clientes_trabalhados",
        "contatos_efetivos",
        "contatos_cliente",
        "cpcs",
        "clientes_cpc",
        "acordos",
        "pagamentos",
        "acordos_sem_pagamento",
        "acordos_em_aberto",
        "acordos_nao_pagou",
        "pct_quebra",
        "valor_negociado",
        "valor_pago",
        "valor_em_aberto",
        "valor_nao_pagou",
        "valor_quebra",
        "ticket_medio",
        "tx_contato",
        "tx_acordo",
        "tx_acordo_cliente_cpc",
        "tx_pagamento",
        "efetividade_pagamento",
        "tx_pagamento_cpc",
        "tx_sem_pagamento",
        "recuperacao",
        "score",
    ]
    for col in metric_cols:
        if col in df:
            df[col] = df[col].fillna(0)
    df["negociador_cadastrado"] = np.where(df["nome_colaborador"].notna(), "Sim", "Não")
    df["nome_colaborador"] = df["nome_colaborador"].fillna("Não localizado na base")
    df["base_colaborador"] = df["base_colaborador"].fillna("Fora da base")
    df["cargo_colaborador"] = df["cargo_colaborador"].fillna("Não localizado")

    df["meta_individual"] = operator_goal_series(df["OPERADOR"], meses_count)
    df["meta_geral_escritorio"] = meta_geral
    df["pct_quebra"] = safe_div(df["valor_quebra"], df["meta_individual"])
    df["atingimento_meta_individual"] = safe_div(df["valor_pago"], df["meta_individual"])
    df["pct_aberto_meta_individual"] = safe_div(df["valor_em_aberto"], df["meta_individual"])
    df["saldo_meta_individual"] = df["valor_pago"] - df["meta_individual"]
    df["participacao_meta_geral"] = safe_div(df["valor_pago"], df["meta_geral_escritorio"])
    df["quartil_meta_individual"] = quartile_label(df["atingimento_meta_individual"], min_series=df["valor_pago"], min_value=0.01)
    df["quartil_meta_geral"] = quartile_label(df["participacao_meta_geral"], min_series=df["valor_pago"], min_value=0.01)
    df["diagnostico_meta"] = np.select(
        [
            df["valor_pago"] <= 0,
            df["atingimento_meta_individual"] >= 1,
            df["atingimento_meta_individual"] >= 0.75,
            df["atingimento_meta_individual"] >= 0.50,
        ],
        ["Sem recebimento", "Meta batida", "Próximo da meta", "Atenção"],
        default="Crítico",
    )
    return df.sort_values(["atingimento_meta_individual", "valor_pago"], ascending=False), meses, meta_geral


def meta_operator_groups(df):
    metrics = [
        ("Atingimento da meta individual", "quartil_meta_individual"),
        ("Participação na meta geral", "quartil_meta_geral"),
    ]
    groups = ["Q4 - destaque", "Q3 - bom", "Q2 - atenção", "Q1 - crítico", "Sem base"]
    rows = []
    for metric_label, metric_col in metrics:
        for group in groups:
            operadores = sorted(df.loc[df[metric_col].eq(group), "OPERADOR"].dropna().astype(str).tolist())
            rows.append(
                {
                    "Métrica": metric_label,
                    "Grupo": group,
                    "Qtd. operadores": len(operadores),
                    "Operadores": ", ".join(operadores) if operadores else "-",
                }
            )
    return pd.DataFrame(rows)


def cpc_operator_groups(df):
    metrics = [
        ("Volume de CPCs únicos", "quartil_cpc_volume"),
        ("Conversão CPC único → acordo", "quartil_cpc_acordo"),
        ("Conversão CPC único → pagamento", "quartil_cpc_pagamento"),
    ]
    groups = ["Q4 - destaque", "Q3 - bom", "Q2 - atenção", "Q1 - crítico", "Sem base"]
    rows = []
    for metric_label, metric_col in metrics:
        for group in groups:
            operadores = sorted(df.loc[df[metric_col].eq(group), "OPERADOR"].dropna().astype(str).tolist())
            rows.append(
                {
                    "Métrica": metric_label,
                    "Grupo": group,
                    "Qtd. operadores": len(operadores),
                    "Operadores": ", ".join(operadores) if operadores else "-",
                }
            )
    return pd.DataFrame(rows)


def cpc_quartil_descritivo(df):
    ordem = ["Q4 - destaque", "Q3 - bom", "Q2 - atenção", "Q1 - crítico", "Sem base"]
    rows = []
    for grupo in ordem:
        sub = df[df["quartil_cpc_pagamento"].eq(grupo)]
        if sub.empty:
            continue
        com_base = sub[sub["cpcs_unicos"] >= 3]
        rows.append({
            "Quartil": grupo,
            "Qtd. operadores": len(sub),
            "CPCs únicos (média)": num_fmt(com_base["cpcs_unicos"].mean()) if not com_base.empty else "-",
            "CPC → acordo (média)": pct_fmt(com_base["tx_cpc_unico_acordo"].mean()) if not com_base.empty else "-",
            "CPC → pgto (média)": pct_fmt(com_base["tx_cpc_unico_pagamento"].mean()) if not com_base.empty else "-",
            "Efetividade pgto (média)": pct_fmt(com_base["efetividade_pagamento"].mean()) if not com_base.empty else "-",
            "Acordos (total)": num_fmt(sub["acordos"].sum()),
            "Pagamentos (total)": num_fmt(sub["pagamentos"].sum()),
            "Acordos s/ pgto (total)": num_fmt(sub["acordos_sem_pagamento"].sum()),
            "Valor recebido (total)": money_fmt(sub["valor_pago"].sum()),
        })
    return pd.DataFrame(rows)


def cpc_volume_quartil_descritivo(df):
    ordem = ["Q4 - destaque", "Q3 - bom", "Q2 - atenção", "Q1 - crítico", "Sem base"]
    rows = []
    for grupo in ordem:
        sub = df[df["quartil_cpc_volume"].eq(grupo)]
        if sub.empty:
            continue
        com_base = sub[sub["cpcs_unicos"] >= 1]
        rows.append({
            "Quartil": grupo,
            "Qtd. operadores": len(sub),
            "CPCs únicos (média)": num_fmt(com_base["cpcs_unicos"].mean()) if not com_base.empty else "-",
            "CPCs únicos (mín.)": num_fmt(com_base["cpcs_unicos"].min()) if not com_base.empty else "-",
            "CPCs únicos (máx.)": num_fmt(com_base["cpcs_unicos"].max()) if not com_base.empty else "-",
            "CPC → acordo (média)": pct_fmt(com_base["tx_cpc_unico_acordo"].mean()) if not com_base.empty else "-",
            "CPC → pgto (média)": pct_fmt(com_base["tx_cpc_unico_pagamento"].mean()) if not com_base.empty else "-",
            "Acordos (total)": num_fmt(sub["acordos"].sum()),
            "Pagamentos (total)": num_fmt(sub["pagamentos"].sum()),
            "Valor recebido (total)": money_fmt(sub["valor_pago"].sum()),
        })
    return pd.DataFrame(rows)


@st.cache_data(show_spinner=False)
def load_workplan():
    try:
        import psycopg2
    except ImportError:
        return pd.DataFrame(), "Driver psycopg2-binary não instalado."

    cfg = postgres_config()
    if not cfg["configured"]:
        return (
            pd.DataFrame(),
            "Workplan nao configurado. Configure SUPABASE_DB_URL/DATABASE_URL ou PGHOST/PGDATABASE/PGUSER/PGPASSWORD nas variaveis de ambiente ou em st.secrets.",
        )

    query = f"""
        SELECT
            agreement_no,
            cust_name,
            cpf_cnpj,
            dpd,
            total_amount_due,
            last_contact_date,
            allocation_date,
            last_marking_date,
            city,
            state,
            last_marking_value,
            pct_of_margin_money,
            no_first_ins_unpaid,
            status,
            flag_cobravel,
            status_base,
            flag_cpc,
            status_cpc,
            probabilidade,
            faixa_atraso,
            regiao,
            uf
        FROM "{cfg['schema']}"."{cfg['table']}"
    """
    try:
        if cfg["database_url"]:
            conn = psycopg2.connect(cfg["database_url"])
        else:
            conn = psycopg2.connect(
                host=cfg["host"],
                port=cfg["port"],
                dbname=cfg["database"],
                user=cfg["user"],
                password=cfg["password"],
            )
        with conn:
            df = pd.read_sql_query(query, conn)
    except Exception as exc:
        host = str(cfg.get("host", "")).lower()
        message = str(exc)
        if host in {"localhost", "127.0.0.1", "::1"} and "connection refused" in message.lower():
            return (
                pd.DataFrame(),
                "PostgreSQL local nao esta acessivel a partir deste ambiente. Em deploy, localhost aponta para o servidor do Streamlit; configure SUPABASE_DB_URL ou DATABASE_URL com a connection string do Supabase.",
            )
        return pd.DataFrame(), f"Não foi possível carregar o Workplan: {exc}"
    finally:
        if "conn" in locals():
            conn.close()

    df.columns = [normalize_text(c).lower() for c in df.columns]
    df["CONTRATO_KEY"] = df["agreement_no"].map(normalize_contract)
    df["cpf_cnpj"] = df["cpf_cnpj"].map(normalize_text)
    df["dpd"] = pd.to_numeric(df["dpd"], errors="coerce")
    df["total_amount_due"] = pd.to_numeric(df["total_amount_due"], errors="coerce").fillna(0)
    df["pct_of_margin_money"] = pd.to_numeric(df["pct_of_margin_money"], errors="coerce")
    df["no_first_ins_unpaid"] = pd.to_numeric(df["no_first_ins_unpaid"], errors="coerce")
    df["inadimplencia_precoce_tipo"] = df["no_first_ins_unpaid"].map(early_default_type)
    for col in ["last_contact_date", "allocation_date", "last_marking_date"]:
        df[col] = pd.to_datetime(df[col], errors="coerce")
    df["SEGMENTO_DPD"] = df["dpd"].map(segmento_dpd)
    df["FAIXA_ATRASO"] = df["faixa_atraso"].map(normalize_text)
    df["FAIXA_ATRASO"] = df["FAIXA_ATRASO"].where(df["FAIXA_ATRASO"].ne(""), df["dpd"].map(atraso_faixa))
    df["REGIÃO"] = df["regiao"].map(normalize_text)
    df["flag_cobravel"] = df["flag_cobravel"].map(normalize_text).str.upper()
    df["flag_cpc"] = df["flag_cpc"].map(normalize_text).str.upper()
    df["status_cpc"] = df["status_cpc"].map(normalize_text)
    df["status_base"] = df["status_base"].map(normalize_text)
    return df[df["CONTRATO_KEY"].notna()].copy(), None


def build_workplan_analysis(workplan, eventos_hist, resultados_hist):
    if workplan.empty:
        return workplan

    eventos_contrato = eventos_hist.groupby("CONTRATO_KEY", dropna=True).agg(
        acionamentos_hist=("EVENTO_TXT", "size"),
        cpcs_hist=("IS_CPC", "sum"),
        ultimo_acionamento=("DATA", "max"),
    ).reset_index()
    eventos_perfil_cols = ["CONTRATO_KEY"] + [col for col in ["PRODUTO"] if col in eventos_hist.columns]
    eventos_perfil = eventos_hist[eventos_perfil_cols].dropna(subset=["CONTRATO_KEY"]).drop_duplicates("CONTRATO_KEY")
    resultados_early = resultados_hist.copy()
    if "PARCELA_NUM" in resultados_early.columns:
        resultados_early["PARCELA_NAO_PAGA_NUM"] = resultados_early["PARCELA_NUM"].where(resultados_early["IS_NAO_PAGOU"])
    else:
        resultados_early["PARCELA_NAO_PAGA_NUM"] = np.nan
    resultados_contrato = resultados_early.groupby("CONTRATO_KEY", dropna=True).agg(
        acordos_hist=("CONTRATO_KEY", "count"),
        pagamentos_hist=("IS_PAGO", "sum"),
        acordos_em_aberto_hist=("IS_EM_ABERTO", "sum"),
        acordos_nao_pagou_hist=("IS_NAO_PAGOU", "sum"),
        valor_negociado_hist=("VALOR_NEGOCIADO", "sum"),
        valor_pago_hist=("VALOR_PAGO", "sum"),
        primeira_parcela_nao_paga_hist=("PARCELA_NAO_PAGA_NUM", "min"),
    ).reset_index()

    df = workplan.merge(eventos_contrato, on="CONTRATO_KEY", how="left")
    df = df.merge(eventos_perfil, on="CONTRATO_KEY", how="left")
    df = df.merge(resultados_contrato, on="CONTRATO_KEY", how="left")
    fill_zero = [
        "acionamentos_hist",
        "cpcs_hist",
        "acordos_hist",
        "pagamentos_hist",
        "acordos_em_aberto_hist",
        "acordos_nao_pagou_hist",
        "valor_negociado_hist",
        "valor_pago_hist",
        "primeira_parcela_nao_paga_hist",
    ]
    for col in fill_zero:
        df[col] = df[col].fillna(0)
    historico_precoce = df["primeira_parcela_nao_paga_hist"].map(early_default_type)
    df["inadimplencia_precoce_tipo"] = df["inadimplencia_precoce_tipo"].where(
        df["inadimplencia_precoce_tipo"].ne("Demais"),
        historico_precoce,
    )

    ultimo_contato = df[["last_contact_date", "ultimo_acionamento"]].max(axis=1)
    hoje = pd.Timestamp.today().normalize()
    df["dias_sem_contato"] = (hoje - ultimo_contato).dt.days
    df["dias_sem_contato"] = df["dias_sem_contato"].fillna(999).clip(lower=0)

    profile_rates, global_profile_rates = build_recovery_profile_rates(eventos_hist, resultados_hist)
    df = apply_recovery_profile_rates(df, profile_rates, global_profile_rates)

    df["probabilidade_recuperacao"] = (
        df["prob_pagamento_perfil"].fillna(0) * 0.70
        + df["prob_acordo_perfil"].fillna(0) * 0.20
        + df["recuperacao_media_perfil"].fillna(0) * 0.10
    ).clip(lower=0, upper=1)
    df["probabilidade_recuperacao"] = np.where(
        (df["cpcs_hist"] > 0) | df["flag_cpc"].eq("SIM"),
        np.minimum(df["probabilidade_recuperacao"] * 1.12, 1),
        df["probabilidade_recuperacao"],
    )
    df["probabilidade_recuperacao"] = np.where(
        df["dias_sem_contato"] >= 60,
        df["probabilidade_recuperacao"] * 0.95,
        df["probabilidade_recuperacao"],
    )
    df["probabilidade_recuperacao"] = np.where(
        df["flag_cobravel"].eq("SIM"),
        df["probabilidade_recuperacao"],
        df["probabilidade_recuperacao"] * 0.35,
    )
    df["ajuste_fpd_epd"] = np.select(
        [
            df["inadimplencia_precoce_tipo"].eq("FPD"),
            df["inadimplencia_precoce_tipo"].eq("EPD"),
        ],
        [0.45, 0.75],
        default=1.0,
    )
    df["probabilidade_recuperacao"] = (df["probabilidade_recuperacao"] * df["ajuste_fpd_epd"]).clip(lower=0, upper=1)
    df["probabilidade_recuperacao"] = np.where(
        df["inadimplencia_precoce_tipo"].eq("FPD"),
        np.minimum(df["probabilidade_recuperacao"], 0.25),
        df["probabilidade_recuperacao"],
    )
    df["probabilidade_recuperacao"] = np.where(
        df["inadimplencia_precoce_tipo"].eq("EPD"),
        np.minimum(df["probabilidade_recuperacao"], 0.50),
        df["probabilidade_recuperacao"],
    )

    df["valor_potencial"] = df["total_amount_due"].fillna(0)
    df["valor_esperado_recuperacao"] = (
        df["valor_potencial"]
        * df["probabilidade_recuperacao"]
        * df["recuperacao_media_perfil"].fillna(global_profile_rates.get("recuperacao_media_perfil", 0))
    )
    valor_rank = df["valor_potencial"].rank(pct=True).fillna(0)
    esperado_rank = df["valor_esperado_recuperacao"].rank(pct=True).fillna(0)
    prob_rank = df["probabilidade_recuperacao"].rank(pct=True).fillna(0)
    df["score_recuperacao"] = (prob_rank * 0.45 + esperado_rank * 0.40 + valor_rank * 0.15).clip(0, 1)
    df["prioridade_workplan"] = np.select(
        [
            df["score_recuperacao"] >= 0.70,
            df["score_recuperacao"] >= 0.45,
        ],
        ["Alta", "Média"],
        default="Baixa",
    )

    valor_alto_limite = df["valor_potencial"].quantile(0.75) if not df.empty else 0
    probabilidade_alta_limite = df["probabilidade_recuperacao"].quantile(0.75) if not df.empty else 0
    df["valor_alto_limite"] = valor_alto_limite
    df["probabilidade_alta_limite"] = probabilidade_alta_limite
    df["motivo_abordagem"] = df.apply(abordagem_workplan, axis=1)
    df["motivo_priorizacao"] = df["motivo_abordagem"]
    df = df.drop(columns=["valor_alto_limite", "probabilidade_alta_limite"], errors="ignore")
    return df.sort_values(["valor_esperado_recuperacao", "score_recuperacao", "total_amount_due"], ascending=False)

    motivos = []
    for _, row in df.iterrows():
        parts = []
        if row["SEGMENTO_DPD"] in {"POTLOSS", "SALVAGE"}:
            parts.append(row["SEGMENTO_DPD"])
        if row["total_amount_due"] >= df["total_amount_due"].quantile(0.75):
            parts.append("alto valor")
        if row["cpcs_hist"] > 0 or row["flag_cpc"] == "SIM":
            parts.append("histórico de CPC")
        if row["dias_sem_contato"] >= 15:
            parts.append("sem contato recente")
        motivos.append(", ".join(parts) if parts else "baixo sinal histórico")
    df["motivo_priorizacao"] = motivos
    return df.sort_values(["score_recuperacao", "total_amount_due"], ascending=False)


@st.cache_data(show_spinner=False)
def load_data(data_version):
    eventos = pd.read_excel(EVENTOS_FILE, sheet_name="Eventos", usecols=[0, 1, 4, 7, 8, 11])
    clientes_file = latest_file("Pesquisa-Cliente-908-*.xlsx")
    contratos = pd.read_excel(clientes_file, sheet_name="Contratos", usecols=[0, 5, 6, 7, 8, 9, 14, 18])
    resultados = pd.read_excel(
        RESULTADOS_FILE,
        sheet_name="BASE",
        usecols=[0, 1, 3, 5, 6, 7, 8, 9, 11, 12, 15, 16, 19, 20, 22, 24, 25],
    )

    eventos.columns = [normalize_text(c).upper() for c in eventos.columns]
    contratos.columns = [normalize_text(c).upper() for c in contratos.columns]
    resultados.columns = [normalize_text(c).upper() for c in resultados.columns]

    eventos = eventos.dropna(how="all").copy()
    contratos = contratos.dropna(how="all").copy()
    resultados = resultados.dropna(subset=["Nº CONTRATO", "ACORDO POR"], how="all").copy()

    eventos["CONTRATO_KEY"] = eventos["CONTRATO"].map(normalize_contract)
    eventos["CPF_CNPJ_KEY"] = eventos.get("CPF/CNPJ", pd.Series(index=eventos.index, dtype="object")).map(normalize_document)
    eventos["OPERADOR"] = eventos["OPERADOR"].map(normalize_operator)
    eventos = eventos[~eventos["OPERADOR"].map(is_excluded_operator)].copy()
    eventos["DATA"] = pd.to_datetime(eventos["DATA"], dayfirst=True, errors="coerce")
    eventos["EVENTO_TXT"] = eventos["EVENTO"].map(normalize_text)
    eventos["EVENTO_UPPER"] = eventos["EVENTO_TXT"].str.upper()
    eventos["TIPO DE ACIONAMENTO"] = eventos["TIPO DE ACIONAMENTO"].map(normalize_text)

    contratos["CONTRATO_KEY"] = contratos["CONTRATO"].map(normalize_contract)
    contratos["CPF_CNPJ_KEY"] = contratos.get("CPF/CNPJ", pd.Series(index=contratos.index, dtype="object")).map(normalize_document)
    contratos["ATRASO"] = pd.to_numeric(contratos["ATRASO"], errors="coerce")
    contratos["TOTAL ABERTO"] = pd.to_numeric(contratos["TOTAL ABERTO"], errors="coerce")
    contratos["FAIXA_ATRASO"] = contratos["ATRASO"].map(atraso_faixa)
    contratos["SEGMENTO_DPD"] = contratos["ATRASO"].map(segmento_dpd)

    contrato_cols = [
        "CONTRATO_KEY",
        "CPF_CNPJ_KEY",
        "PRODUTO",
        "REGIAO",
        "FILIAL",
        "ESTAGIO",
        "ATRASO",
        "TOTAL ABERTO",
        "FAIXA_ATRASO",
        "SEGMENTO_DPD",
    ]
    contrato_cols = [c for c in contrato_cols if c in contratos.columns]
    contratos_lookup = contratos[contrato_cols].drop_duplicates("CONTRATO_KEY")
    eventos = eventos.merge(contratos_lookup, on="CONTRATO_KEY", how="left", suffixes=("", "_CONTRATO"))
    if "CPF_CNPJ_KEY_CONTRATO" in eventos.columns:
        eventos["CPF_CNPJ_KEY"] = eventos["CPF_CNPJ_KEY"].fillna(eventos["CPF_CNPJ_KEY_CONTRATO"])
        eventos = eventos.drop(columns=["CPF_CNPJ_KEY_CONTRATO"])
    eventos["CPC_CLIENT_KEY"] = eventos["CPF_CNPJ_KEY"].fillna(eventos["CONTRATO_KEY"])

    eventos["IS_AUTO"] = eventos["OPERADOR"].eq("auto")
    eventos["IS_IMPORTACAO"] = eventos["EVENTO_UPPER"].str.contains("IMPORTACAO|IMPORTAÇÃO", na=False)
    eventos["IS_ACIONAMENTO"] = ~(eventos["IS_AUTO"] | eventos["IS_IMPORTACAO"])
    eventos["IS_CPC"] = eventos["EVENTO_UPPER"].str.match(r"^\s*(02|03|04|05)\b", na=False)
    eventos["IS_CONTATO_EFETIVO"] = eventos["IS_CPC"]
    eventos["IS_CONTATO_CLIENTE"] = eventos["EVENTO_UPPER"].str.match(r"^\s*(02|03)\b", na=False)
    eventos["MES"] = eventos["DATA"].dt.to_period("M").astype(str)

    resultados["CONTRATO_KEY"] = resultados["Nº CONTRATO"].map(normalize_contract)
    resultados["OPERADOR"] = resultados["ACORDO POR"].map(normalize_operator)
    resultados = resultados[~resultados["OPERADOR"].map(is_excluded_operator)].copy()
    resultados["DATA_ACORDO"] = pd.to_datetime(resultados["EMISSÃO"], errors="coerce")
    resultados["DATA_PAGAMENTO"] = pd.to_datetime(resultados["DATA DO PAGAMENTO"], errors="coerce")
    resultados["DATA_VENCIMENTO"] = pd.to_datetime(resultados["DATA DE VENCIMENTO"], errors="coerce")
    resultados["VALOR_NEGOCIADO"] = pd.to_numeric(resultados["VALOR DO BANCO - META"], errors="coerce").fillna(0)
    resultados = resultados.merge(
        contratos_lookup[["CONTRATO_KEY", "CPF_CNPJ_KEY"]],
        on="CONTRATO_KEY",
        how="left",
    )
    resultados["CPC_CLIENT_KEY"] = resultados["CPF_CNPJ_KEY"].fillna(resultados["CONTRATO_KEY"])
    resultados["PARCELA_NUM"] = resultados["PARCELA"].map(parse_installment_number)
    honorarios_col = next((col for col in resultados.columns if "HONOR" in normalize_status(col)), None)
    if honorarios_col:
        resultados["HONORARIOS_ESCRITORIO_BASE"] = pd.to_numeric(resultados[honorarios_col], errors="coerce").fillna(0)
    else:
        resultados["HONORARIOS_ESCRITORIO_BASE"] = 0.0
    resultados["DPD"] = pd.to_numeric(resultados["DPD"], errors="coerce")
    resultados["FAIXA_ATRASO"] = resultados["DPD"].map(atraso_faixa)
    resultados["SEGMENTO_DPD"] = resultados["DPD FORMULA"].map(segmento_dpd)
    resultados["REGIÃO"] = resultados["REGIÃO"].fillna(resultados.get("UF", "Sem região")).map(normalize_text)
    resultados["UF"] = resultados["UF"].map(normalize_text)
    resultados["CAMPANHA"] = resultados["CAMPANHA"].map(normalize_text)
    resultados["STATUS"] = resultados["STATUS"].map(normalize_text).str.upper()
    resultados["STATUS_KEY"] = resultados["STATUS"].map(normalize_status)
    resultados["IS_ACORDO"] = resultados["CONTRATO_KEY"].notna() & resultados["OPERADOR"].notna()
    resultados["IS_EM_ABERTO"] = resultados["STATUS_KEY"].eq("EM ABERTO")
    resultados["IS_NAO_PAGOU"] = resultados["STATUS_KEY"].eq("NAO PAGOU")
    resultados["IS_PAGO"] = resultados["STATUS_KEY"].eq("PAGOU") | (resultados["DATA_PAGAMENTO"].notna())
    resultados["VALOR_PAGO"] = np.where(resultados["IS_PAGO"], resultados["VALOR_NEGOCIADO"], 0)
    resultados["HONORARIOS_ESCRITORIO"] = np.where(resultados["IS_PAGO"], resultados["HONORARIOS_ESCRITORIO_BASE"], 0)
    resultados["VALOR_EM_ABERTO"] = np.where(resultados["IS_EM_ABERTO"], resultados["VALOR_NEGOCIADO"], 0)
    resultados["VALOR_NAO_PAGOU"] = np.where(resultados["IS_NAO_PAGOU"], resultados["VALOR_NEGOCIADO"], 0)
    resultados["MES_RESULTADO"] = resultados["MÊS"].map(normalize_text).str.upper()
    resultados["MES_NUM"] = pd.to_numeric(resultados["Nº CORRESPONDENTE AO MÊS"], errors="coerce")
    resultados["MES"] = resultados["DATA_ACORDO"].dt.to_period("M").astype(str)
    resultados = resultados[resultados["IS_ACORDO"]].copy()

    return eventos, contratos, resultados


def apply_filters(eventos, resultados):
    st.sidebar.title("Filtros")

    operadores = sorted(set(eventos["OPERADOR"].dropna()) | set(resultados["OPERADOR"].dropna()))
    regioes = sorted(resultados["REGIÃO"].replace("", np.nan).dropna().unique())
    faixas = ["000-030", "031-060", "061-090", "091-120", "121-180", "181-360", "361+", "Sem atraso"]
    segmentos_dpd = ["POTLOSS", "SALVAGE", "SALVAGE +"]
    campanhas = sorted(resultados["CAMPANHA"].replace("", np.nan).dropna().unique())
    produtos = sorted(eventos.get("PRODUTO", pd.Series(dtype=str)).replace("", np.nan).dropna().unique())
    meses_df = result_months_frame(resultados)
    meses = meses_df["MES_RESULTADO"].tolist()
    mes_padrao = []
    if not meses_df.empty:
        mes_atual = MONTH_NAMES_PT.get(pd.Timestamp.today().month)
        if DEFAULT_RESULT_MONTH in meses:
            mes_padrao = [DEFAULT_RESULT_MONTH]
        elif mes_atual in meses:
            mes_padrao = [mes_atual]
        else:
            meses_validos = meses_df[meses_df["MES_NUM"] <= pd.Timestamp.today().month]
            base_mes = meses_validos if not meses_validos.empty else meses_df
            mes_padrao = [base_mes.sort_values(["MES_NUM", "MES_RESULTADO"]).iloc[-1]["MES_RESULTADO"]]

    operador_sel = st.sidebar.multiselect("Operador", operadores)
    mes_sel = st.sidebar.multiselect("Mês do resultado", meses, default=mes_padrao)
    regiao_sel = st.sidebar.multiselect("Região", regioes)
    faixa_sel = st.sidebar.multiselect("Faixa de atraso", faixas)
    segmento_dpd_sel = st.sidebar.multiselect("Segmento DPD", segmentos_dpd)
    campanha_sel = st.sidebar.multiselect("Campanha", campanhas)
    produto_sel = st.sidebar.multiselect("Produto", produtos)
    incluir_auto = st.sidebar.toggle("Incluir AUTO/importação nos acionamentos", value=False)

    min_data = eventos["DATA"].min()
    max_data = eventos["DATA"].max()
    min_event_date = min_data.date()
    max_event_date = max_data.date()
    st.sidebar.caption(f"Eventos disponiveis: {min_event_date:%d/%m/%Y} a {max_event_date:%d/%m/%Y}")
    if "event_start_date" not in st.session_state:
        st.session_state["event_start_date"] = min_event_date
    if "event_end_date" not in st.session_state:
        st.session_state["event_end_date"] = max_event_date
    if st.session_state["event_start_date"] < min_event_date or st.session_state["event_start_date"] > max_event_date:
        st.session_state["event_start_date"] = min_event_date
    if st.session_state["event_end_date"] < min_event_date or st.session_state["event_end_date"] > max_event_date:
        st.session_state["event_end_date"] = max_event_date

    data_inicio = st.sidebar.date_input(
        "Inicio dos eventos",
        min_value=min_event_date,
        max_value=max_event_date,
        key="event_start_date",
        format="DD/MM/YYYY",
    )
    data_fim = st.sidebar.date_input(
        "Fim dos eventos",
        min_value=min_event_date,
        max_value=max_event_date,
        key="event_end_date",
        format="DD/MM/YYYY",
    )
    if data_inicio > data_fim:
        st.sidebar.warning("A data inicial nao pode ser maior que a final.")
    else:
        inicio, fim = pd.Timestamp(data_inicio), pd.Timestamp(data_fim) + pd.Timedelta(days=1)
        eventos = eventos[(eventos["DATA"].isna()) | ((eventos["DATA"] >= inicio) & (eventos["DATA"] < fim))]

    if operador_sel:
        eventos = eventos[eventos["OPERADOR"].isin(operador_sel)]
        resultados = resultados[resultados["OPERADOR"].isin(operador_sel)]
    if mes_sel:
        resultados = resultados[resultados["MES_RESULTADO"].isin(mes_sel)]
    if regiao_sel:
        resultados = resultados[resultados["REGIÃO"].isin(regiao_sel)]
    if faixa_sel:
        eventos = eventos[eventos["FAIXA_ATRASO"].isin(faixa_sel)]
        resultados = resultados[resultados["FAIXA_ATRASO"].isin(faixa_sel)]
    if segmento_dpd_sel:
        eventos = eventos[eventos["SEGMENTO_DPD"].isin(segmento_dpd_sel)]
        resultados = resultados[resultados["SEGMENTO_DPD"].isin(segmento_dpd_sel)]
    if campanha_sel:
        resultados = resultados[resultados["CAMPANHA"].isin(campanha_sel)]
        contratos_campanha = set(resultados["CONTRATO_KEY"].dropna())
        eventos = eventos[eventos["CONTRATO_KEY"].isin(contratos_campanha)]
    if produto_sel and "PRODUTO" in eventos.columns:
        eventos = eventos[eventos["PRODUTO"].isin(produto_sel)]
    if not incluir_auto:
        eventos = eventos[eventos["IS_ACIONAMENTO"]]

    return eventos, resultados, operador_sel


def ensure_cpc_client_key(eventos, resultados):
    eventos = eventos.copy()
    resultados = resultados.copy()
    if "CPC_CLIENT_KEY" not in eventos.columns:
        if "CPF_CNPJ_KEY" in eventos.columns:
            eventos["CPC_CLIENT_KEY"] = eventos["CPF_CNPJ_KEY"].fillna(eventos["CONTRATO_KEY"])
        else:
            eventos["CPC_CLIENT_KEY"] = eventos["CONTRATO_KEY"]
    if "CPC_CLIENT_KEY" not in resultados.columns:
        if "CPF_CNPJ_KEY" in resultados.columns:
            resultados["CPC_CLIENT_KEY"] = resultados["CPF_CNPJ_KEY"].fillna(resultados["CONTRATO_KEY"])
        else:
            resultados["CPC_CLIENT_KEY"] = resultados["CONTRATO_KEY"]
    return eventos, resultados


def aggregate_operator(eventos, resultados):
    ev = eventos.groupby("OPERADOR", dropna=True).agg(
        acionamentos=("EVENTO_TXT", "size"),
        clientes_trabalhados=("CONTRATO_KEY", "nunique"),
        contatos_efetivos=("IS_CONTATO_EFETIVO", "sum"),
        contatos_cliente=("IS_CONTATO_CLIENTE", "sum"),
        cpcs=("IS_CPC", "sum"),
        clientes_cpc=("CPC_CLIENT_KEY", lambda s: s[eventos.loc[s.index, "IS_CPC"]].nunique()),
    )
    rs = resultados.groupby("OPERADOR", dropna=True).agg(
        acordos=("CONTRATO_KEY", "count"),
        pagamentos=("IS_PAGO", "sum"),
        acordos_sem_pagamento=("IS_PAGO", lambda s: (~s).sum()),
        acordos_em_aberto=("IS_EM_ABERTO", "sum"),
        acordos_nao_pagou=("IS_NAO_PAGOU", "sum"),
        valor_negociado=("VALOR_NEGOCIADO", "sum"),
        valor_pago=("VALOR_PAGO", "sum"),
        valor_em_aberto=("VALOR_EM_ABERTO", "sum"),
        valor_nao_pagou=("VALOR_NAO_PAGOU", "sum"),
        ticket_medio=("VALOR_NEGOCIADO", "mean"),
    )
    df = ev.join(rs, how="outer").fillna(0).reset_index()
    df["tx_contato"] = safe_div(df["contatos_efetivos"], df["acionamentos"])
    df["tx_acordo"] = safe_div(df["acordos"], df["cpcs"])
    df["tx_acordo_cliente_cpc"] = safe_div(df["acordos"], df["clientes_cpc"])
    df["tx_pagamento"] = safe_div(df["pagamentos"], df["acordos"])
    df["efetividade_pagamento"] = safe_div(df["pagamentos"], df["pagamentos"] + df["acordos_nao_pagou"])
    df["tx_pagamento_cpc"] = safe_div(df["pagamentos"], df["cpcs"])
    df["tx_sem_pagamento"] = safe_div(df["acordos_sem_pagamento"], df["acordos"])
    df["valor_quebra"] = df["valor_nao_pagou"]
    df["meta_individual"] = operator_goal_series(df["OPERADOR"], selected_months_count(resultados))
    df["pct_quebra"] = safe_div(df["valor_quebra"], df["meta_individual"])
    df["recuperacao"] = safe_div(df["valor_pago"], df["valor_negociado"])
    df["score"] = (
        df["tx_contato"].rank(pct=True) * 0.15
        + df["tx_acordo"].rank(pct=True) * 0.25
        + df["tx_pagamento"].rank(pct=True) * 0.25
        + df["valor_pago"].rank(pct=True) * 0.25
        + df["clientes_trabalhados"].rank(pct=True) * 0.10
    )
    return df.sort_values(["score", "valor_pago"], ascending=False)


def aggregate_cpc_operator(eventos, resultados):
    cpc_eventos = eventos[eventos["IS_CPC"]].copy()
    ev = cpc_eventos.groupby("OPERADOR", dropna=True).agg(
        cpcs=("EVENTO_TXT", "size"),
        clientes_cpc=("CPC_CLIENT_KEY", "nunique"),
        contratos_cpc=("CONTRATO_KEY", "nunique"),
    )
    ev["cpcs_unicos"] = ev["clientes_cpc"]
    cpc_clientes = cpc_eventos[["OPERADOR", "CPC_CLIENT_KEY"]].dropna().drop_duplicates()
    resultado_cliente = resultados.groupby(["OPERADOR", "CPC_CLIENT_KEY"], dropna=True).agg(
        qtd_acordos=("CONTRATO_KEY", "count"),
        teve_pagamento=("IS_PAGO", "max"),
        teve_em_aberto=("IS_EM_ABERTO", "max"),
        teve_nao_pagou=("IS_NAO_PAGOU", "max"),
        valor_negociado=("VALOR_NEGOCIADO", "sum"),
        valor_pago=("VALOR_PAGO", "sum"),
        valor_em_aberto=("VALOR_EM_ABERTO", "sum"),
        valor_nao_pagou=("VALOR_NAO_PAGOU", "sum"),
    ).reset_index()
    cpc_resultado = cpc_clientes.merge(resultado_cliente, on=["OPERADOR", "CPC_CLIENT_KEY"], how="left")
    cpc_resultado["qtd_acordos"] = cpc_resultado["qtd_acordos"].fillna(0)
    cpc_resultado["teve_pagamento"] = cpc_resultado["teve_pagamento"].fillna(False).astype(bool)
    cpc_resultado["teve_em_aberto"] = cpc_resultado["teve_em_aberto"].fillna(False).astype(bool)
    cpc_resultado["teve_nao_pagou"] = cpc_resultado["teve_nao_pagou"].fillna(False).astype(bool)
    cpc_resultado["valor_negociado"] = cpc_resultado["valor_negociado"].fillna(0)
    cpc_resultado["valor_pago"] = cpc_resultado["valor_pago"].fillna(0)
    cpc_resultado["valor_em_aberto"] = cpc_resultado["valor_em_aberto"].fillna(0)
    cpc_resultado["valor_nao_pagou"] = cpc_resultado["valor_nao_pagou"].fillna(0)
    cpc_resultado["teve_acordo"] = cpc_resultado["qtd_acordos"] > 0
    cpc_resultado["acordo_sem_pagamento"] = cpc_resultado["teve_acordo"] & ~cpc_resultado["teve_pagamento"]

    rs = cpc_resultado.groupby("OPERADOR", dropna=True).agg(
        acordos=("teve_acordo", "sum"),
        pagamentos=("teve_pagamento", "sum"),
        acordos_sem_pagamento=("acordo_sem_pagamento", "sum"),
        acordos_em_aberto=("teve_em_aberto", "sum"),
        acordos_nao_pagou=("teve_nao_pagou", "sum"),
        valor_negociado=("valor_negociado", "sum"),
        valor_pago=("valor_pago", "sum"),
        valor_em_aberto=("valor_em_aberto", "sum"),
        valor_nao_pagou=("valor_nao_pagou", "sum"),
    )
    df = ev.join(rs, how="left").fillna(0).reset_index()
    df["ticket_medio"] = safe_div(df["valor_negociado"], df["acordos"])
    df["tx_cpc_acordo"] = safe_div(df["acordos"], df["contratos_cpc"])
    df["tx_cpc_pagamento"] = safe_div(df["pagamentos"], df["contratos_cpc"])
    df["tx_cpc_unico_acordo"] = safe_div(df["acordos"], df["cpcs_unicos"])
    df["tx_cpc_unico_pagamento"] = safe_div(df["pagamentos"], df["cpcs_unicos"])
    df["tx_acordo_pagamento"] = safe_div(df["pagamentos"], df["acordos"])
    df["efetividade_pagamento"] = safe_div(df["pagamentos"], df["pagamentos"] + df["acordos_nao_pagou"])
    df["tx_acordo_sem_pagamento"] = safe_div(df["acordos_sem_pagamento"], df["acordos"])
    df["valor_quebra"] = df["valor_nao_pagou"]
    df["meta_individual"] = operator_goal_series(df["OPERADOR"], selected_months_count(resultados))
    df["pct_quebra"] = safe_div(df["valor_quebra"], df["meta_individual"])
    df["recuperacao"] = safe_div(df["valor_pago"], df["valor_negociado"])
    df["quartil_cpc_acordo"] = quartile_label(df["tx_cpc_unico_acordo"], min_series=df["cpcs_unicos"], min_value=3)
    df["quartil_cpc_pagamento"] = quartile_label(df["tx_cpc_unico_pagamento"], min_series=df["cpcs_unicos"], min_value=3)
    df["quartil_cpc_volume"] = quartile_label(df["cpcs_unicos"], min_series=df["cpcs_unicos"], min_value=1)
    df["diagnostico_cpc"] = np.where(
        df["cpcs_unicos"] < 3,
        "Sem base",
        np.select(
            [
                df["tx_cpc_unico_pagamento"] >= 0.20,
                df["tx_cpc_unico_pagamento"] >= 0.10,
                df["tx_cpc_unico_pagamento"] >= 0.05,
            ],
            ["Alta conversão", "Boa conversão", "Atenção"],
            default="Crítico",
        ),
    )
    return df.sort_values(["pagamentos", "valor_pago", "tx_cpc_pagamento"], ascending=False)


FAIXA_ATRASO_ORDER = ["000-030", "031-060", "061-090", "091-120", "121-180", "181-360", "361+"]


def aggregate_operator_faixa(eventos, resultados):
    ev = eventos.groupby(["OPERADOR", "FAIXA_ATRASO"], dropna=True).agg(
        acionamentos=("EVENTO_TXT", "size"),
        cpcs=("IS_CPC", "sum"),
    ).reset_index()
    rs = resultados.groupby(["OPERADOR", "FAIXA_ATRASO"], dropna=True).agg(
        acordos=("CONTRATO_KEY", "count"),
        pagamentos=("IS_PAGO", "sum"),
        valor_pago=("VALOR_PAGO", "sum"),
    ).reset_index()
    df = ev.merge(rs, on=["OPERADOR", "FAIXA_ATRASO"], how="outer").fillna(0)
    df = df[df["FAIXA_ATRASO"].isin(FAIXA_ATRASO_ORDER)]
    df["tx_contato"] = safe_div(df["cpcs"], df["acionamentos"])
    df["tx_acordo"] = safe_div(df["acordos"], df["cpcs"])
    df["tx_pagamento_cpc"] = safe_div(df["pagamentos"], df["cpcs"])
    return df


def melhor_faixa_por_operador(faixa_operador_df, min_cpcs=5):
    elegivel = faixa_operador_df[faixa_operador_df["cpcs"] >= min_cpcs].copy()
    if elegivel.empty:
        return pd.DataFrame()

    melhor_cpc = (
        elegivel.sort_values(["OPERADOR", "tx_contato"], ascending=[True, False])
        .groupby("OPERADOR")
        .head(1)[["OPERADOR", "FAIXA_ATRASO", "tx_contato", "acionamentos", "cpcs"]]
        .rename(columns={"FAIXA_ATRASO": "Melhor faixa (contato → CPC)", "tx_contato": "_tx_contato", "acionamentos": "_acion_cpc", "cpcs": "_cpcs_cpc"})
    )
    melhor_conversao = (
        elegivel.sort_values(["OPERADOR", "tx_pagamento_cpc"], ascending=[True, False])
        .groupby("OPERADOR")
        .head(1)[["OPERADOR", "FAIXA_ATRASO", "tx_pagamento_cpc", "cpcs", "pagamentos", "valor_pago"]]
        .rename(columns={"FAIXA_ATRASO": "Melhor faixa (CPC → pagamento)", "tx_pagamento_cpc": "_tx_pgto", "cpcs": "_cpcs_pgto", "pagamentos": "_pagtos_pgto", "valor_pago": "_valor_pgto"})
    )
    out = melhor_cpc.merge(melhor_conversao, on="OPERADOR", how="outer")
    out["Taxa contato → CPC"] = out["_tx_contato"].map(pct_fmt)
    out["Acionamentos na faixa"] = out["_acion_cpc"].map(num_fmt)
    out["CPCs na faixa (contato)"] = out["_cpcs_cpc"].map(num_fmt)
    out["Taxa CPC → pagamento"] = out["_tx_pgto"].map(pct_fmt)
    out["CPCs na faixa (conversão)"] = out["_cpcs_pgto"].map(num_fmt)
    out["Pagamentos na faixa"] = out["_pagtos_pgto"].map(num_fmt)
    out["Valor recebido na faixa"] = out["_valor_pgto"].map(money_fmt)
    out["_ordenacao"] = out["_valor_pgto"]
    out = out.sort_values("_ordenacao", ascending=False)
    return out[
        [
            "OPERADOR",
            "Melhor faixa (contato → CPC)",
            "Taxa contato → CPC",
            "Acionamentos na faixa",
            "CPCs na faixa (contato)",
            "Melhor faixa (CPC → pagamento)",
            "Taxa CPC → pagamento",
            "CPCs na faixa (conversão)",
            "Pagamentos na faixa",
            "Valor recebido na faixa",
        ]
    ].rename(columns={"OPERADOR": "Operador"})


def aggregate_resultados(resultados, dimension):
    df = resultados.groupby(dimension, dropna=False).agg(
        clientes=("CONTRATO_KEY", "nunique"),
        acordos=("CONTRATO_KEY", "count"),
        pagamentos=("IS_PAGO", "sum"),
        acordos_em_aberto=("IS_EM_ABERTO", "sum"),
        acordos_nao_pagou=("IS_NAO_PAGOU", "sum"),
        valor_negociado=("VALOR_NEGOCIADO", "sum"),
        valor_pago=("VALOR_PAGO", "sum"),
        valor_em_aberto=("VALOR_EM_ABERTO", "sum"),
        valor_nao_pagou=("VALOR_NAO_PAGOU", "sum"),
        ticket_medio=("VALOR_NEGOCIADO", "mean"),
    ).reset_index()
    df["tx_pagamento"] = safe_div(df["pagamentos"], df["acordos"])
    df["efetividade_pagamento"] = safe_div(df["pagamentos"], df["pagamentos"] + df["acordos_nao_pagou"])
    df["valor_quebra"] = df["valor_nao_pagou"]
    df["pct_quebra"] = safe_div(df["valor_quebra"], office_goal_for_resultados(resultados))
    df["recuperacao"] = safe_div(df["valor_pago"], df["valor_negociado"])
    return df


def group_label(df, columns):
    cols = [columns] if isinstance(columns, str) else list(columns)
    cols = [col for col in cols if col in df.columns]
    if not cols:
        return pd.Series("Sem grupo", index=df.index, dtype="object")
    labels = df[cols].apply(lambda row: " | ".join(normalize_text(value) or "Sem grupo" for value in row), axis=1)
    return labels.replace("", "Sem grupo")


def valid_group_mask(df, columns):
    cols = [columns] if isinstance(columns, str) else list(columns)
    cols = [col for col in cols if col in df.columns]
    if not cols:
        return pd.Series(False, index=df.index)
    mask = pd.Series(True, index=df.index)
    for col in cols:
        mask = mask & df[col].map(normalize_text).ne("")
    return mask


def add_projection_group(df, columns):
    grouped = df[valid_group_mask(df, columns)].copy()
    if grouped.empty:
        return grouped
    grouped["grupo"] = group_label(grouped, columns)
    return grouped


def value_band(value):
    amount = pd.to_numeric(value, errors="coerce")
    if pd.isna(amount):
        return "Sem valor"
    if amount < 1000:
        return "Ate 1k"
    if amount < 5000:
        return "1k-5k"
    if amount < 15000:
        return "5k-15k"
    if amount < 50000:
        return "15k-50k"
    return "50k+"


def no_contact_band(days):
    days = pd.to_numeric(days, errors="coerce")
    if pd.isna(days):
        return "Sem contato"
    if days < 7:
        return "0-6 dias"
    if days < 15:
        return "7-14 dias"
    if days < 30:
        return "15-29 dias"
    if days < 60:
        return "30-59 dias"
    return "60+ dias"


def profile_key(df, columns):
    cols = [col for col in columns if col in df.columns]
    if not cols:
        return pd.Series("GERAL", index=df.index, dtype="object")
    return df[cols].apply(lambda row: "||".join(normalize_text(value).upper() or "SEM_INFO" for value in row), axis=1)


def build_recovery_profile_rates(eventos_hist, resultados_hist):
    base_cols = ["CONTRATO_KEY", "SEGMENTO_DPD", "FAIXA_ATRASO", "REGIAO", "UF", "TOTAL ABERTO"]
    if "PRODUTO" in eventos_hist.columns:
        base_cols.append("PRODUTO")
    base_cols = [col for col in base_cols if col in eventos_hist.columns]
    if not base_cols or eventos_hist.empty:
        return {}, {}

    hist = eventos_hist[base_cols + ["IS_CPC", "DATA"]].dropna(subset=["CONTRATO_KEY"]).copy()
    if hist.empty:
        return {}, {}

    profile_aggs = {col: (col, lambda s: normalize_text(s.dropna().iloc[-1]) if not s.dropna().empty else "") for col in base_cols if col != "CONTRATO_KEY"}
    profile_aggs["cpcs_hist_base"] = ("IS_CPC", "sum")
    profile_aggs["ultimo_acionamento_base"] = ("DATA", "max")
    hist = hist.groupby("CONTRATO_KEY", dropna=True).agg(**profile_aggs).reset_index()

    resultados_early = resultados_hist.copy()
    if "PARCELA_NUM" in resultados_early.columns:
        resultados_early["PARCELA_NAO_PAGA_NUM"] = resultados_early["PARCELA_NUM"].where(resultados_early["IS_NAO_PAGOU"])
    else:
        resultados_early["PARCELA_NAO_PAGA_NUM"] = np.nan
    resultados_contrato = resultados_early.groupby("CONTRATO_KEY", dropna=True).agg(
        acordos_hist_base=("CONTRATO_KEY", "count"),
        pagamentos_hist_base=("IS_PAGO", "sum"),
        acordos_nao_pagou_hist_base=("IS_NAO_PAGOU", "sum"),
        valor_negociado_hist_base=("VALOR_NEGOCIADO", "sum"),
        valor_pago_hist_base=("VALOR_PAGO", "sum"),
        primeira_parcela_nao_paga_hist=("PARCELA_NAO_PAGA_NUM", "min"),
    ).reset_index()
    hist = hist.merge(resultados_contrato, on="CONTRATO_KEY", how="left")
    for col in [
        "acordos_hist_base",
        "pagamentos_hist_base",
        "acordos_nao_pagou_hist_base",
        "valor_negociado_hist_base",
        "valor_pago_hist_base",
        "primeira_parcela_nao_paga_hist",
    ]:
        hist[col] = hist[col].fillna(0)

    hist["inadimplencia_precoce_tipo"] = hist["primeira_parcela_nao_paga_hist"].map(early_default_type)
    hist["flag_cpc_perfil"] = np.where(hist["cpcs_hist_base"] > 0, "SIM", "NAO")
    valor_historico = pd.to_numeric(hist["valor_negociado_hist_base"], errors="coerce").fillna(0)
    if "TOTAL ABERTO" in hist.columns:
        total_aberto_hist = pd.to_numeric(hist["TOTAL ABERTO"], errors="coerce").fillna(0)
        valor_historico = valor_historico.where(valor_historico > 0, total_aberto_hist)
    hist["valor_faixa"] = valor_historico.map(value_band)
    hoje = pd.Timestamp.today().normalize()
    hist["dias_sem_contato_perfil"] = (hoje - pd.to_datetime(hist["ultimo_acionamento_base"], errors="coerce")).dt.days
    hist["dias_sem_contato_faixa"] = hist["dias_sem_contato_perfil"].map(no_contact_band)

    if "REGIAO" in hist.columns and "REGIÃO" not in hist.columns:
        hist["REGIÃO"] = hist["REGIAO"]
    if "UF" in hist.columns and "uf" not in hist.columns:
        hist["uf"] = hist["UF"]

    profile_sets = [
        ["inadimplencia_precoce_tipo", "SEGMENTO_DPD", "FAIXA_ATRASO", "REGIÃO", "uf", "PRODUTO", "flag_cpc_perfil", "valor_faixa", "dias_sem_contato_faixa"],
        ["inadimplencia_precoce_tipo", "SEGMENTO_DPD", "FAIXA_ATRASO", "REGIÃO", "uf", "flag_cpc_perfil", "valor_faixa"],
        ["inadimplencia_precoce_tipo", "SEGMENTO_DPD", "FAIXA_ATRASO", "flag_cpc_perfil"],
        ["SEGMENTO_DPD", "FAIXA_ATRASO", "REGIÃO", "uf", "PRODUTO", "flag_cpc_perfil", "valor_faixa", "dias_sem_contato_faixa"],
        ["SEGMENTO_DPD", "FAIXA_ATRASO", "REGIÃO", "uf", "flag_cpc_perfil", "valor_faixa"],
        ["SEGMENTO_DPD", "FAIXA_ATRASO", "REGIÃO", "flag_cpc_perfil"],
        ["SEGMENTO_DPD", "FAIXA_ATRASO", "flag_cpc_perfil"],
        ["SEGMENTO_DPD", "flag_cpc_perfil"],
        ["SEGMENTO_DPD"],
    ]

    rates = {}
    for idx, cols in enumerate(profile_sets):
        cols = [col for col in cols if col in hist.columns]
        if not cols:
            continue
        frame = hist.copy()
        frame["_profile_key"] = profile_key(frame, cols)
        grouped = frame.groupby("_profile_key", dropna=False).agg(
            contratos_hist=("CONTRATO_KEY", "nunique"),
            acordos_hist_perfil=("acordos_hist_base", lambda s: (s > 0).sum()),
            pagamentos_hist_perfil=("pagamentos_hist_base", lambda s: (s > 0).sum()),
            quebras_hist_perfil=("acordos_nao_pagou_hist_base", lambda s: (s > 0).sum()),
            valor_negociado_hist_perfil=("valor_negociado_hist_base", "sum"),
            valor_pago_hist_perfil=("valor_pago_hist_base", "sum"),
        ).reset_index()
        grouped["prob_acordo_perfil"] = safe_div(grouped["acordos_hist_perfil"], grouped["contratos_hist"])
        grouped["prob_pagamento_perfil"] = safe_div(grouped["pagamentos_hist_perfil"], grouped["contratos_hist"])
        grouped["recuperacao_media_perfil"] = safe_div(grouped["valor_pago_hist_perfil"], grouped["valor_negociado_hist_perfil"])
        grouped["risco_quebra_perfil"] = safe_div(grouped["quebras_hist_perfil"], grouped["acordos_hist_perfil"])
        rates[idx] = {"columns": cols, "data": grouped}

    totals = {
        "contratos_hist": hist["CONTRATO_KEY"].nunique(),
        "acordos_hist_perfil": int((hist["acordos_hist_base"] > 0).sum()),
        "pagamentos_hist_perfil": int((hist["pagamentos_hist_base"] > 0).sum()),
        "quebras_hist_perfil": int((hist["acordos_nao_pagou_hist_base"] > 0).sum()),
        "valor_negociado_hist_perfil": float(hist["valor_negociado_hist_base"].sum()),
        "valor_pago_hist_perfil": float(hist["valor_pago_hist_base"].sum()),
    }
    totals["prob_acordo_perfil"] = scalar_safe_div(totals["acordos_hist_perfil"], totals["contratos_hist"])
    totals["prob_pagamento_perfil"] = scalar_safe_div(totals["pagamentos_hist_perfil"], totals["contratos_hist"])
    totals["recuperacao_media_perfil"] = scalar_safe_div(totals["valor_pago_hist_perfil"], totals["valor_negociado_hist_perfil"])
    totals["risco_quebra_perfil"] = scalar_safe_div(totals["quebras_hist_perfil"], totals["acordos_hist_perfil"])
    for profile in rates.values():
        grouped = profile["data"]
        weight = (grouped["contratos_hist"] / (grouped["contratos_hist"] + 20)).fillna(0)
        for col in ["prob_acordo_perfil", "prob_pagamento_perfil", "recuperacao_media_perfil", "risco_quebra_perfil"]:
            grouped[col] = (grouped[col] * weight) + totals[col] * (1 - weight)
            grouped[col] = grouped[col].clip(lower=0, upper=1)
        profile["data"] = grouped
    return rates, totals


def apply_recovery_profile_rates(workplan_df, profile_rates, global_rates):
    df = workplan_df.copy()
    defaults = {
        "contratos_hist": global_rates.get("contratos_hist", 0),
        "prob_acordo_perfil": global_rates.get("prob_acordo_perfil", 0),
        "prob_pagamento_perfil": global_rates.get("prob_pagamento_perfil", 0),
        "recuperacao_media_perfil": global_rates.get("recuperacao_media_perfil", 0),
        "risco_quebra_perfil": global_rates.get("risco_quebra_perfil", 0),
        "base_probabilidade": "Média geral",
    }
    for col, value in defaults.items():
        df[col] = value

    if "uf" not in df.columns and "UF" in df.columns:
        df["uf"] = df["UF"]
    df["flag_cpc_perfil"] = np.where((df["cpcs_hist"] > 0) | df["flag_cpc"].eq("SIM"), "SIM", "NAO")
    df["valor_faixa"] = df["total_amount_due"].map(value_band)
    df["dias_sem_contato_faixa"] = df["dias_sem_contato"].map(no_contact_band)

    metric_cols = [
        "contratos_hist",
        "prob_acordo_perfil",
        "prob_pagamento_perfil",
        "recuperacao_media_perfil",
        "risco_quebra_perfil",
    ]
    for idx in sorted(profile_rates):
        cols = profile_rates[idx]["columns"]
        rates = profile_rates[idx]["data"]
        if not cols or rates.empty or not all(col in df.columns for col in cols):
            continue
        lookup = rates.set_index("_profile_key")[metric_cols].to_dict("index")
        keys = profile_key(df, cols)
        matched = keys.map(lookup)
        mask = matched.notna() & df["base_probabilidade"].eq("Média geral")
        if not mask.any():
            continue
        matched_df = pd.DataFrame(matched[mask].tolist(), index=df.index[mask])
        for col in metric_cols:
            df.loc[mask, col] = matched_df[col]
        df.loc[mask, "base_probabilidade"] = "Perfil: " + " + ".join(cols)

    return df


def abordagem_workplan(row):
    reasons = []
    if row.get("inadimplencia_precoce_tipo") == "FPD":
        reasons.append("FPD: baixa propensao, cobrar com menor intensidade")
    elif row.get("inadimplencia_precoce_tipo") == "EPD":
        reasons.append("EPD: validar historico antes de priorizar")
    if row.get("cpcs_hist", 0) > 0 and row.get("acordos_hist", 0) <= 0:
        reasons.append("CPC recente sem acordo")
    if row.get("total_amount_due", 0) >= row.get("valor_alto_limite", 0) and row.get("dias_sem_contato", 0) >= 30:
        reasons.append("Alto valor sem contato ha 30+ dias")
    if row.get("probabilidade_alta_limite", 0) > 0 and row.get("probabilidade_recuperacao", 0) >= row.get("probabilidade_alta_limite", 0):
        reasons.append("Perfil com alta conversao historica")
    if row.get("flag_cobravel") == "SIM" and row.get("acionamentos_hist", 0) <= 0:
        reasons.append("Contrato cobravel sem tentativa recente")
    if row.get("acordos_nao_pagou_hist", 0) > 0 or row.get("risco_quebra_perfil", 0) >= 0.40:
        reasons.append("Quebrou acordo antes, abordar com cautela")
    if row.get("base_probabilidade") == "Média geral":
        reasons.append("Sem perfil historico especifico")
    return "; ".join(reasons[:3]) if reasons else "Prospectar conforme perfil historico"


def expected_recovery_by_group(workplan_view, eventos_hist, resultados_hist, workplan_col, eventos_col, resultados_col):
    workplan_cols = [workplan_col] if isinstance(workplan_col, str) else list(workplan_col)
    if workplan_view.empty or not any(col in workplan_view.columns for col in workplan_cols):
        return pd.DataFrame()

    carteira = add_projection_group(workplan_view, workplan_col)
    if carteira.empty:
        return pd.DataFrame()
    carteira_df = (
        carteira.groupby("grupo", dropna=False)
        .agg(
            contratos_elegiveis=("CONTRATO_KEY", "nunique"),
            carteira_elegivel=("total_amount_due", "sum"),
        )
        .reset_index()
    )

    eventos_base = eventos_hist[eventos_hist["IS_ACIONAMENTO"]].copy()
    eventos_cols = [eventos_col] if isinstance(eventos_col, str) else list(eventos_col)
    if any(col in eventos_base.columns for col in eventos_cols):
        eventos_base = eventos_base[valid_group_mask(eventos_base, eventos_col)].copy()
        eventos_base["grupo"] = group_label(eventos_base, eventos_col)
        eventos_df = (
            eventos_base.groupby("grupo", dropna=False)
            .agg(
                acionamentos_hist=("EVENTO_TXT", "size"),
                contatos_cliente_hist=("IS_CONTATO_CLIENTE", "sum"),
                cpcs_hist=("IS_CPC", "sum"),
            )
            .reset_index()
        )
    else:
        eventos_df = pd.DataFrame(columns=["grupo", "acionamentos_hist", "contatos_cliente_hist", "cpcs_hist"])

    resultados_cols = [resultados_col] if isinstance(resultados_col, str) else list(resultados_col)
    if any(col in resultados_hist.columns for col in resultados_cols):
        resultados_base = resultados_hist[valid_group_mask(resultados_hist, resultados_col)].copy()
        resultados_base["grupo"] = group_label(resultados_base, resultados_col)
        resultados_df = (
            resultados_base.groupby("grupo", dropna=False)
            .agg(
                acordos_hist=("CONTRATO_KEY", "count"),
                pagamentos_hist=("IS_PAGO", "sum"),
                valor_negociado_hist=("VALOR_NEGOCIADO", "sum"),
                valor_pago_hist=("VALOR_PAGO", "sum"),
            )
            .reset_index()
        )
    else:
        resultados_df = pd.DataFrame(columns=["grupo", "acordos_hist", "pagamentos_hist", "valor_negociado_hist", "valor_pago_hist"])

    df = carteira_df.merge(eventos_df, on="grupo", how="left").merge(resultados_df, on="grupo", how="left")
    fill_cols = [
        "acionamentos_hist",
        "contatos_cliente_hist",
        "cpcs_hist",
        "acordos_hist",
        "pagamentos_hist",
        "valor_negociado_hist",
        "valor_pago_hist",
    ]
    for col in fill_cols:
        df[col] = df[col].fillna(0)

    global_rates = {
        "taxa_contato": scalar_safe_div(eventos_base["IS_CONTATO_CLIENTE"].sum(), len(eventos_base)),
        "taxa_cpc": scalar_safe_div(eventos_base["IS_CPC"].sum(), eventos_base["IS_CONTATO_CLIENTE"].sum()),
        "taxa_acordo": scalar_safe_div(len(resultados_hist), eventos_base["IS_CPC"].sum()),
        "taxa_pagamento": scalar_safe_div(resultados_hist["IS_PAGO"].sum(), len(resultados_hist)),
        "percentual_medio_recuperado": scalar_safe_div(resultados_hist["VALOR_PAGO"].sum(), resultados_hist["VALOR_NEGOCIADO"].sum()),
    }

    df["taxa_contato_grupo"] = safe_div(df["contatos_cliente_hist"], df["acionamentos_hist"])
    df["taxa_cpc_grupo"] = safe_div(df["cpcs_hist"], df["contatos_cliente_hist"])
    df["taxa_acordo_grupo"] = safe_div(df["acordos_hist"], df["cpcs_hist"])
    df["taxa_pagamento_grupo"] = safe_div(df["pagamentos_hist"], df["acordos_hist"])
    df["percentual_medio_recuperado_grupo"] = safe_div(df["valor_pago_hist"], df["valor_negociado_hist"])

    usar_media_contato = df["acionamentos_hist"] < 30
    usar_media_cpc = df["contatos_cliente_hist"] < 10
    usar_media_acordo = df["cpcs_hist"] < 10
    usar_media_pagamento = df["acordos_hist"] < 5
    usar_media_recuperado = df["valor_negociado_hist"] <= 0

    df["taxa_contato"] = np.where(usar_media_contato, global_rates["taxa_contato"], df["taxa_contato_grupo"])
    df["taxa_cpc"] = np.where(usar_media_cpc, global_rates["taxa_cpc"], df["taxa_cpc_grupo"])
    df["taxa_acordo"] = np.where(usar_media_acordo, global_rates["taxa_acordo"], df["taxa_acordo_grupo"])
    df["taxa_pagamento"] = np.where(usar_media_pagamento, global_rates["taxa_pagamento"], df["taxa_pagamento_grupo"])
    df["percentual_medio_recuperado"] = np.where(
        usar_media_recuperado,
        global_rates["percentual_medio_recuperado"],
        df["percentual_medio_recuperado_grupo"],
    )
    for col in ["taxa_contato", "taxa_cpc", "taxa_acordo", "taxa_pagamento", "percentual_medio_recuperado"]:
        df[col] = pd.Series(df[col]).fillna(0).clip(lower=0, upper=1)

    df["recuperacao_esperada"] = (
        df["carteira_elegivel"]
        * df["taxa_contato"]
        * df["taxa_cpc"]
        * df["taxa_acordo"]
        * df["taxa_pagamento"]
        * df["percentual_medio_recuperado"]
    )
    df["recuperacao_esperada_pct_carteira"] = safe_div(df["recuperacao_esperada"], df["carteira_elegivel"])
    usa_alguma_media = usar_media_contato | usar_media_cpc | usar_media_acordo | usar_media_pagamento | usar_media_recuperado
    usa_todas_medias = usar_media_contato & usar_media_cpc & usar_media_acordo & usar_media_pagamento & usar_media_recuperado
    df["base_taxas"] = np.select(
        [usa_todas_medias, usa_alguma_media],
        ["Média geral", "Grupo + média geral"],
        default="Grupo",
    )
    return df.sort_values("recuperacao_esperada", ascending=False)


def metric_card(label, value, help_text=None):
    value_text = str(value)
    title = help_text or f"{label}: {value_text}"
    compact_class = " metric-card--compact" if len(value_text) >= 14 else ""
    st.markdown(
        f"""
        <div class="metric-card{compact_class}" title="{escape(title)}">
            <div class="metric-card__label">{escape(str(label))}</div>
            <div class="metric-card__value">{escape(value_text)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def gauge_point(cx, cy, radius, ratio):
    angle = math.pi * (1 - ratio)
    return cx + radius * math.cos(angle), cy - radius * math.sin(angle)


def gauge_path(cx, cy, radius, start_ratio, end_ratio):
    start_x, start_y = gauge_point(cx, cy, radius, start_ratio)
    end_x, end_y = gauge_point(cx, cy, radius, end_ratio)
    return f"M {start_x:.2f} {start_y:.2f} A {radius} {radius} 0 0 1 {end_x:.2f} {end_y:.2f}"


def hex_to_rgb(color):
    color = color.lstrip("#")
    return tuple(int(color[i:i + 2], 16) for i in (0, 2, 4))


def blend_hex(start, end, ratio):
    ratio = min(max(ratio, 0), 1)
    start_rgb = hex_to_rgb(start)
    end_rgb = hex_to_rgb(end)
    blended = [round(a + (b - a) * ratio) for a, b in zip(start_rgb, end_rgb)]
    return "#{:02x}{:02x}{:02x}".format(*blended)


def meta_progress_color(pct):
    pct = 0 if pd.isna(pct) else min(max(float(pct), 0), 1)
    if pct <= 0.40:
        return blend_hex("#9f1d2f", "#de6a76", pct / 0.40)
    if pct <= 0.70:
        return blend_hex("#c95616", "#f2a23a", (pct - 0.40) / 0.30)
    if pct <= 0.80:
        return blend_hex("#d4a514", "#ffe071", (pct - 0.70) / 0.10)
    return blend_hex("#1f7a4d", "#72d391", (pct - 0.80) / 0.20)


def business_day_clock_details(month_labels, today=None):
    today = (today or pd.Timestamp.now(tz="America/Sao_Paulo").tz_localize(None)).normalize()
    month_lookup = {normalize_status(name): month for month, name in MONTH_NAMES_PT.items()}
    operational_holidays = {pd.Timestamp("2026-06-04")}
    selected_months = []
    for label in month_labels or []:
        month_num = month_lookup.get(normalize_status(label))
        if month_num:
            selected_months.append(month_num)
    if not selected_months:
        selected_months = [today.month]

    elapsed_days = 0
    total_days = 0
    remaining_days = 0
    for month_num in selected_months:
        month_start = pd.Timestamp(year=today.year, month=month_num, day=1)
        month_end = month_start + pd.offsets.MonthEnd(0)
        month_business_days = pd.bdate_range(month_start, month_end)
        previous_month_start = month_start - pd.offsets.MonthBegin(1)
        previous_month_end = month_start - pd.Timedelta(days=1)
        previous_month_business_days = pd.bdate_range(previous_month_start, previous_month_end)
        period_start = previous_month_business_days[-1] if len(previous_month_business_days) else previous_month_end
        period_end = month_business_days[-2] if len(month_business_days) >= 2 else month_end
        period_business_days = pd.bdate_range(period_start, period_end)
        period_business_days = period_business_days[~period_business_days.normalize().isin(operational_holidays)]
        month_total_days = OPERATIONAL_BUSINESS_DAY_TOTALS.get((today.year, month_num), len(period_business_days))
        elapsed_for_month = 0
        total_days += month_total_days
        if today >= period_end:
            elapsed_for_month = month_total_days
        elif today >= period_start:
            elapsed_for_month = min(len(period_business_days[period_business_days <= today]), month_total_days)
        elapsed_days += elapsed_for_month
        remaining_days += max(month_total_days - elapsed_for_month, 0)

    return {
        "elapsed_days": elapsed_days,
        "total_days": total_days,
        "remaining_days": remaining_days,
        "ratio": scalar_safe_div(elapsed_days, total_days),
    }


def business_day_clock_ratio(month_labels, today=None):
    return business_day_clock_details(month_labels, today).get("ratio", 0)


def meta_gauge(
    value,
    target,
    month_label,
    open_today_count,
    open_today_value,
    paid_today_count,
    paid_today_value,
    avg_paid_ticket=0,
    pos_paid_value=0,
    pos_paid_value_pct=0,
    pos_paid_count_pct=0,
    clock_target=0,
    clock_gap=0,
    clock_ratio=0,
    daily_target=0,
    remaining_business_days=0,
    open_today_rows=None,
    paid_today_rows=None,
    title="Recebimento Total",
):
    value = 0 if pd.isna(value) else float(value)
    target = 0 if pd.isna(target) else float(target)
    open_today_count = 0 if pd.isna(open_today_count) else int(open_today_count)
    open_today_value = 0 if pd.isna(open_today_value) else float(open_today_value)
    paid_today_count = 0 if pd.isna(paid_today_count) else int(paid_today_count)
    paid_today_value = 0 if pd.isna(paid_today_value) else float(paid_today_value)
    avg_paid_ticket = 0 if pd.isna(avg_paid_ticket) else float(avg_paid_ticket)
    pos_paid_value = 0 if pd.isna(pos_paid_value) else float(pos_paid_value)
    pos_paid_value_pct = 0 if pd.isna(pos_paid_value_pct) else float(pos_paid_value_pct)
    pos_paid_count_pct = 0 if pd.isna(pos_paid_count_pct) else float(pos_paid_count_pct)
    clock_target = 0 if pd.isna(clock_target) else float(clock_target)
    clock_gap = 0 if pd.isna(clock_gap) else float(clock_gap)
    clock_ratio = 0 if pd.isna(clock_ratio) else float(clock_ratio)
    daily_target = 0 if pd.isna(daily_target) else float(daily_target)
    remaining_business_days = 0 if pd.isna(remaining_business_days) else int(remaining_business_days)
    if target <= 0:
        st.info("Sem meta geral cadastrada para montar o indicador.")
        return

    gap = max(target - value, 0)
    progress = min(max(value / target, 0), 1)
    pct_target = value / target if target else 0
    clock_status = "Dentro" if value >= clock_target else "Abaixo"
    clock_status_color = "#9ee7ef" if value >= clock_target else "#fbbf24"
    color = meta_progress_color(pct_target)
    open_today_rows = open_today_rows if open_today_rows is not None else pd.DataFrame()
    if open_today_rows.empty:
        open_today_html = '<div class="meta-tooltip__empty">Sem boletos em aberto para hoje.</div>'
    else:
        items = []
        for _, row in open_today_rows.sort_values(["OPERADOR", "NOME DO CLIENTE"]).iterrows():
            cliente = normalize_text(row.get("NOME DO CLIENTE", "Cliente sem nome"))
            operador = normalize_text(row.get("OPERADOR", ""))
            valor = money_fmt(row.get("VALOR_EM_ABERTO", 0))
            items.append(
                f'<div class="meta-tooltip__row"><span>{escape(cliente)}</span><span>{escape(operador)}</span><strong>{escape(valor)}</strong></div>'
            )
        open_today_html = "".join(items)

    paid_today_rows = paid_today_rows if paid_today_rows is not None else pd.DataFrame()
    if paid_today_rows.empty:
        paid_today_html = '<div class="meta-tooltip__empty">Sem recebimentos registrados hoje.</div>'
    else:
        items = []
        for _, row in paid_today_rows.sort_values(["OPERADOR", "NOME DO CLIENTE"]).iterrows():
            cliente = normalize_text(row.get("NOME DO CLIENTE", "Cliente sem nome"))
            operador = normalize_text(row.get("OPERADOR", ""))
            valor = money_fmt(row.get("VALOR_PAGO", 0))
            items.append(
                f'<div class="meta-tooltip__row"><span>{escape(cliente)}</span><span>{escape(operador)}</span><strong>{escape(valor)}</strong></div>'
            )
        paid_today_html = "".join(items)

    cx, cy, radius = 320, 230, 185
    bg_path = gauge_path(cx, cy, radius, 0, 1)
    value_path = gauge_path(cx, cy, radius, 0, progress)

    st.markdown(
        f"""
        <style>
        .meta-panel__side {{
            width: 390px;
            display: grid;
            grid-template-columns: repeat(2, minmax(0, 1fr));
            gap: 10px;
        }}
        .meta-panel__card {{
            border: 1px solid rgba(125, 211, 252, .24);
            border-radius: 7px;
            background: rgba(2, 47, 63, .72);
            box-shadow: inset 0 1px 0 rgba(255,255,255,.08);
            color: #f8fafc;
            overflow: visible;
        }}
        .meta-panel__card--wide {{
            grid-column: 1 / -1;
        }}
        .meta-panel__card-title {{
            padding: 7px 10px;
            border-bottom: 1px solid rgba(125, 211, 252, .18);
            color: rgba(255,255,255,.86);
            font-size: .82rem;
            font-weight: 800;
            text-align: center;
            text-transform: uppercase;
        }}
        .meta-panel__value {{
            padding: 12px 10px;
            text-align: center;
            font-size: 1.08rem;
            font-weight: 750;
        }}
        .meta-panel__daily-row {{
            position: relative;
            display: grid;
            grid-template-columns: 1fr auto;
            gap: 8px;
            padding: 8px 10px;
            border-bottom: 1px solid rgba(125, 211, 252, .18);
            cursor: default;
            font-size: .86rem;
        }}
        .meta-panel__daily-row span {{
            color: rgba(255,255,255,.82);
        }}
        .meta-panel__today-value {{
            padding: 9px 10px;
            text-align: center;
            color: #9ee7ef;
            font-weight: 760;
        }}
        .meta-panel__subvalue {{
            padding: 0 10px 10px;
            text-align: center;
            color: rgba(255,255,255,.78);
            font-size: .84rem;
            font-weight: 700;
        }}
        .meta-panel__expected {{
            padding: 0 10px 6px;
            color: #9ee7ef;
            font-size: 1rem;
            font-weight: 820;
        }}
        .meta-tooltip {{
            display: none;
            position: absolute;
            top: 34px;
            right: 0;
            z-index: 20;
            width: 460px;
            max-height: 260px;
            overflow: auto;
            padding: 10px;
            border: 1px solid rgba(125, 211, 252, .34);
            border-radius: 7px;
            background: #07111f;
            box-shadow: 0 18px 34px rgba(0,0,0,.35);
        }}
        .meta-panel__daily-row:hover .meta-tooltip {{
            display: block;
        }}
        .meta-tooltip__title {{
            margin-bottom: 8px;
            color: #f8fafc;
            font-size: .82rem;
            font-weight: 800;
        }}
        .meta-tooltip__row {{
            display: grid;
            grid-template-columns: minmax(150px, 1fr) 120px auto;
            gap: 8px;
            align-items: center;
            padding: 6px 0;
            border-top: 1px solid rgba(148,163,184,.18);
            color: rgba(255,255,255,.84);
            font-size: .78rem;
        }}
        .meta-tooltip__row span {{
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
        }}
        .meta-tooltip__row strong {{
            color: #9ee7ef;
        }}
        .meta-tooltip__empty {{
            color: rgba(255,255,255,.78);
            font-size: .8rem;
        }}
        </style>
        <div style="border:1px solid rgba(47,111,115,.55);border-radius:8px;padding:12px 14px;background:rgba(15,23,42,.10);max-width:1120px;margin:12px auto 18px;">
            <div style="display:grid;grid-template-columns:minmax(0, 1fr) auto;gap:18px;align-items:start;">
                <div>
                    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:2px;">
                        <div style="color:#ffffff;font-size:1.65rem;font-weight:650;line-height:1.1;">{escape(title)}</div>
                        <div style="color:#ffffff;font-size:1.25rem;font-weight:650;line-height:1.1;">{escape(month_label)}</div>
                    </div>
                    <svg viewBox="0 0 640 290" width="100%" height="260" style="display:block;max-height:260px;" role="img" aria-label="{escape(title)}">
                        <path d="{bg_path}" fill="none" stroke="rgba(255,255,255,.14)" stroke-width="70" stroke-linecap="butt"/>
                        <path d="{value_path}" fill="none" stroke="{color}" stroke-width="70" stroke-linecap="butt"/>
                        <rect x="260" y="145" width="120" height="54" rx="10" fill="rgba(255,255,255,.72)"/>
                        <text x="320" y="181" text-anchor="middle" fill="#2f6f73" font-size="28" font-weight="700">{escape(pct_fmt(pct_target))}</text>
                        <text x="320" y="246" text-anchor="middle" fill="#ffffff" font-size="34" font-weight="500">{escape(money_fmt(value))}</text>
                        <text x="118" y="274" fill="rgba(255,255,255,.72)" font-size="15">{escape(money_fmt(0))}</text>
                        <text x="466" y="274" fill="rgba(255,255,255,.72)" font-size="15">{escape(money_fmt(target))}</text>
                    </svg>
                </div>
                <div class="meta-panel__side">
                    <div class="meta-panel__card">
                        <div class="meta-panel__card-title">GAP</div>
                        <div class="meta-panel__value">{escape(money_fmt(gap))}</div>
                    </div>
                    <div class="meta-panel__card">
                        <div class="meta-panel__card-title">Meta relogio</div>
                        <div class="meta-panel__value">{escape(money_fmt(clock_target))}</div>
                        <div class="meta-panel__subvalue meta-panel__expected">Esperado {escape(pct_fmt(clock_ratio))}</div>
                        <div class="meta-panel__subvalue" style="color:{clock_status_color};">{escape(clock_status)} | falta {escape(money_fmt(clock_gap))}</div>
                    </div>
                    <div class="meta-panel__card">
                        <div class="meta-panel__card-title">Meta diaria</div>
                        <div class="meta-panel__value">{escape(money_fmt(daily_target))}</div>
                        <div class="meta-panel__subvalue">{escape(num_fmt(remaining_business_days))} dias uteis restantes</div>
                    </div>
                    <div class="meta-panel__card">
                        <div class="meta-panel__card-title">Ticket medio</div>
                        <div class="meta-panel__value">{escape(money_fmt(avg_paid_ticket))}</div>
                        <div class="meta-panel__subvalue">Pos retomada: {escape(pct_fmt(pos_paid_value_pct))} do recebido</div>
                        <div class="meta-panel__subvalue">{escape(money_fmt(pos_paid_value))} | {escape(pct_fmt(pos_paid_count_pct))} dos acordos</div>
                    </div>
                    <div class="meta-panel__card meta-panel__card--wide">
                        <div class="meta-panel__card-title">Hoje</div>
                        <div class="meta-panel__daily-row">
                            <span>Recebido</span><strong>{escape(num_fmt(paid_today_count))}</strong>
                            <div class="meta-tooltip">
                                <div class="meta-tooltip__title">Recebimentos de hoje</div>
                                {paid_today_html}
                            </div>
                        </div>
                        <div class="meta-panel__today-value">{escape(money_fmt(paid_today_value))}</div>
                        <div class="meta-panel__daily-row">
                            <span>Em aberto</span><strong>{escape(num_fmt(open_today_count))}</strong>
                            <div class="meta-tooltip">
                                <div class="meta-tooltip__title">Boletos em aberto hoje</div>
                                {open_today_html}
                            </div>
                        </div>
                        <div class="meta-panel__today-value">{escape(money_fmt(open_today_value))}</div>
                    </div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def bar_chart(df, x, y, color=None, tooltip=None, title=None, sort="-x", height=320):
    if df.empty:
        st.info("Sem dados para os filtros selecionados.")
        return
    x_encoding = x if not isinstance(x, str) else alt.X(x, title=None)
    y_encoding = y if not isinstance(y, str) else alt.Y(y, title=None, sort=sort)
    chart = (
        alt.Chart(df)
        .mark_bar(cornerRadiusTopRight=3, cornerRadiusBottomRight=3)
        .encode(
            x=x_encoding,
            y=y_encoding,
            color=alt.Color(color, scale=alt.Scale(range=CORP_PALETTE), legend=None) if color else alt.value("#2f6f73"),
            tooltip=tooltip or list(df.columns),
        )
        .properties(height=height, title=title)
    )
    stretch_altair_chart(chart)


def line_chart(df, x, y, color, title):
    if df.empty:
        st.info("Sem dados para os filtros selecionados.")
        return
    chart = (
        alt.Chart(df)
        .mark_line(point=True)
        .encode(
            x=alt.X(x, title=None),
            y=alt.Y(y, title=None),
            color=alt.Color(color, scale=alt.Scale(range=CORP_PALETTE)),
            tooltip=list(df.columns),
        )
        .properties(height=280, title=title)
    )
    stretch_altair_chart(chart)


def heatmap(df, x, y, metric, title):
    if df.empty:
        st.info("Sem dados para os filtros selecionados.")
        return
    tooltip_cols = [
        col for col in df.columns
        if not (col in {"valor_negociado", "valor_pago", "ticket_medio"} and f"{col}_br" in df.columns)
    ]
    chart = (
        alt.Chart(df)
        .mark_rect()
        .encode(
            x=alt.X(x, title=None),
            y=alt.Y(y, title=None),
            color=alt.Color(metric, scale=alt.Scale(scheme="tealblues"), title=None),
            tooltip=tooltip_cols,
        )
        .properties(height=420, title=title)
    )
    stretch_altair_chart(chart)


def payment_profile_analysis(resultados):
    dims = [
        ("OPERADOR", "Operador"),
        ("FAIXA_ATRASO", "Faixa de atraso"),
        ("SEGMENTO_DPD", "Segmento DPD"),
        ("REGIÃO", "Região"),
        ("UF", "UF"),
        ("CAMPANHA", "Campanha"),
    ]
    rows = []
    for col, label in dims:
        if col not in resultados.columns:
            continue
        grouped = (
            resultados.assign(perfil=resultados[col].replace("", np.nan).fillna("Sem informacao"))
            .groupby("perfil", dropna=False)
            .agg(
                clientes=("CONTRATO_KEY", "nunique"),
                acordos=("CONTRATO_KEY", "count"),
                pagamentos=("IS_PAGO", "sum"),
                acordos_em_aberto=("IS_EM_ABERTO", "sum"),
                acordos_nao_pagou=("IS_NAO_PAGOU", "sum"),
                valor_negociado=("VALOR_NEGOCIADO", "sum"),
                valor_pago=("VALOR_PAGO", "sum"),
                valor_em_aberto=("VALOR_EM_ABERTO", "sum"),
                valor_nao_pagou=("VALOR_NAO_PAGOU", "sum"),
            )
            .reset_index()
        )
        grouped["dimensao"] = label
        rows.append(grouped)
    if not rows:
        return pd.DataFrame()
    df = pd.concat(rows, ignore_index=True)
    df["tx_pagamento"] = safe_div(df["pagamentos"], df["acordos"])
    df["efetividade_pagamento"] = safe_div(df["pagamentos"], df["pagamentos"] + df["acordos_nao_pagou"])
    df["recuperacao"] = safe_div(df["valor_pago"], df["valor_negociado"])
    df["pagamentos_esperados_abertos"] = df["acordos_em_aberto"] * df["efetividade_pagamento"].fillna(0)
    df["pagamentos_previstos_total"] = df["pagamentos"] + df["pagamentos_esperados_abertos"]
    df["valor_esperado_aberto"] = df["valor_em_aberto"] * df["efetividade_pagamento"].fillna(0)
    df["score_eficiencia_pagamento"] = (
        df["tx_pagamento"].fillna(0) * 0.45
        + df["efetividade_pagamento"].fillna(0) * 0.35
        + df["recuperacao"].fillna(0) * 0.20
    )
    df["base_confiavel"] = np.select(
        [df["acordos"] >= 20, df["acordos"] >= 8],
        ["Alta", "Media"],
        default="Baixa",
    )
    return df.sort_values(["score_eficiencia_pagamento", "valor_pago", "acordos"], ascending=False)


def workplan_analytics_section(workplan_view):
    if workplan_view.empty:
        st.info("Sem contratos nos filtros atuais para montar a visão analítica.")
        return

    st.subheader("Captação de clientes - visão analítica")

    region_col = next(
        (
            col
            for col in ["REGIÃO", "REGIÃƒO", "REGIÃƒÆ’O", "REGIÃƒÆ’Ã†â€™O", "regiao", "uf", "state"]
            if col in workplan_view.columns
        ),
        None,
    )
    if region_col is None:
        region_col = "_grupo_geografico"
        workplan_view = workplan_view.copy()
        workplan_view[region_col] = "Sem regiao/UF"

    top20_valor = workplan_view.sort_values("valor_esperado_recuperacao", ascending=False).head(20)["valor_esperado_recuperacao"].sum()
    total_valor_esperado = workplan_view["valor_esperado_recuperacao"].sum()
    alta_prioridade = workplan_view[workplan_view["prioridade_workplan"].eq("Alta")]
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        metric_card("Valor esperado", money_fmt(total_valor_esperado))
    with c2:
        metric_card("Top 20 contratos", money_fmt(top20_valor))
    with c3:
        metric_card("Prioridade alta", num_fmt(len(alta_prioridade)))
    with c4:
        metric_card("Chance media", pct_fmt(workplan_view["probabilidade_recuperacao"].mean()))

    prioridade_resumo = (
        workplan_view.groupby("prioridade_workplan", dropna=False)
        .agg(
            clientes=("CONTRATO_KEY", "nunique"),
            valor_potencial=("valor_potencial", "sum"),
            valor_esperado_recuperacao=("valor_esperado_recuperacao", "sum"),
            probabilidade_recuperacao=("probabilidade_recuperacao", "mean"),
        )
        .reset_index()
    )
    prioridade_resumo["ordem_prioridade"] = prioridade_resumo["prioridade_workplan"].map({"Alta": 0, "MÃ©dia": 1, "Baixa": 2}).fillna(3)
    prioridade_resumo = prioridade_resumo.sort_values("ordem_prioridade")
    prioridade_display = display_fields(prioridade_resumo)
    prioridade_chart = (
        alt.Chart(prioridade_display)
        .mark_bar(cornerRadiusTopRight=3, cornerRadiusBottomRight=3)
        .encode(
            x=alt.X("valor_esperado_recuperacao:Q", title="Valor esperado (R$)"),
            y=alt.Y("prioridade_workplan:N", sort=["Alta", "MÃ©dia", "Baixa"], title=None),
            color=alt.Color("prioridade_workplan:N", title="Prioridade", scale=alt.Scale(range=CORP_PALETTE)),
            tooltip=[
                alt.Tooltip("prioridade_workplan:N", title="Prioridade"),
                alt.Tooltip("clientes_br:N", title="Clientes"),
                alt.Tooltip("valor_potencial_br:N", title="Carteira"),
                alt.Tooltip("valor_esperado_recuperacao_br:N", title="Valor esperado"),
                alt.Tooltip("probabilidade_recuperacao_br:N", title="Chance media"),
            ],
        )
        .properties(height=260, title="Valor esperado por prioridade")
    )

    motivo_df = (
        workplan_view.groupby("motivo_abordagem", dropna=False)
        .agg(
            clientes=("CONTRATO_KEY", "nunique"),
            valor_potencial=("valor_potencial", "sum"),
            valor_esperado_recuperacao=("valor_esperado_recuperacao", "sum"),
            probabilidade_recuperacao=("probabilidade_recuperacao", "mean"),
        )
        .reset_index()
        .sort_values("valor_esperado_recuperacao", ascending=False)
        .head(10)
    )
    motivo_display = display_fields(motivo_df)
    motivo_chart = (
        alt.Chart(motivo_display)
        .mark_bar(cornerRadiusTopRight=3, cornerRadiusBottomRight=3)
        .encode(
            x=alt.X("valor_esperado_recuperacao:Q", title="Valor esperado (R$)"),
            y=alt.Y("motivo_abordagem:N", sort="-x", title=None),
            color=alt.value("#2f6f73"),
            tooltip=[
                alt.Tooltip("motivo_abordagem:N", title="Motivo"),
                alt.Tooltip("clientes_br:N", title="Clientes"),
                alt.Tooltip("valor_potencial_br:N", title="Valor potencial"),
                alt.Tooltip("valor_esperado_recuperacao_br:N", title="Valor esperado"),
                alt.Tooltip("probabilidade_recuperacao_br:N", title="Chance média"),
            ],
        )
        .properties(height=330, title="Oportunidade por motivo de abordagem")
    )

    c1, c2 = st.columns([1.2, 1])
    with c1:
        stretch_altair_chart(prioridade_chart)
    with c2:
        stretch_altair_chart(motivo_chart)

    if "inadimplencia_precoce_tipo" in workplan_view.columns:
        fpd_epd_df = (
            workplan_view.groupby("inadimplencia_precoce_tipo", dropna=False)
            .agg(
                clientes=("CONTRATO_KEY", "nunique"),
                valor_potencial=("valor_potencial", "sum"),
                valor_esperado_recuperacao=("valor_esperado_recuperacao", "sum"),
                probabilidade_recuperacao=("probabilidade_recuperacao", "mean"),
            )
            .reset_index()
        )
        fpd_epd_df["ordem_fpd_epd"] = fpd_epd_df["inadimplencia_precoce_tipo"].map({"Demais": 0, "EPD": 1, "FPD": 2}).fillna(3)
        fpd_epd_display = display_fields(fpd_epd_df.sort_values("ordem_fpd_epd"))
        fpd_epd_chart = (
            alt.Chart(fpd_epd_display)
            .mark_bar(cornerRadiusTopRight=3, cornerRadiusBottomRight=3)
            .encode(
                x=alt.X("valor_esperado_recuperacao:Q", title="Valor esperado (R$)"),
                y=alt.Y("inadimplencia_precoce_tipo:N", sort=["Demais", "EPD", "FPD"], title=None),
                color=alt.Color("inadimplencia_precoce_tipo:N", title="FPD/EPD", scale=alt.Scale(range=CORP_PALETTE)),
                tooltip=[
                    alt.Tooltip("inadimplencia_precoce_tipo:N", title="Tipo"),
                    alt.Tooltip("clientes_br:N", title="Clientes"),
                    alt.Tooltip("valor_potencial_br:N", title="Carteira"),
                    alt.Tooltip("valor_esperado_recuperacao_br:N", title="Valor esperado"),
                    alt.Tooltip("probabilidade_recuperacao_br:N", title="Chance media"),
                ],
            )
            .properties(height=190, title="Probabilidade e valor esperado por FPD/EPD")
        )
        stretch_altair_chart(fpd_epd_chart)

    st.markdown("#### Exportar clientes por prioridade e perfil")
    st.caption("Filtre pelo perfil do cliente (segmento DPD / faixa de atraso) e baixe a lista para a operação prospectar e cobrar.")
    exp1, exp2 = st.columns(2)
    with exp1:
        export_prioridades = st.multiselect(
            "Prioridade para exportar",
            ["Alta", "MÃƒÂ©dia", "Baixa"],
            default=["Alta"],
            key="workplan_export_prioridades",
        )
        motivos_opcoes = motivo_df["motivo_abordagem"].dropna().tolist()
        export_motivos = st.multiselect(
            "Motivo de abordagem para exportar",
            motivos_opcoes,
            default=[],
            key="workplan_export_motivos",
        )
    with exp2:
        segmentos_export_opcoes = sorted(workplan_view["SEGMENTO_DPD"].dropna().unique().tolist()) if "SEGMENTO_DPD" in workplan_view.columns else []
        export_segmentos = st.multiselect(
            "Segmento DPD (tipo de cliente) para exportar",
            segmentos_export_opcoes,
            default=[],
            key="workplan_export_segmentos",
        )
        faixas_export_opcoes = sorted(workplan_view["FAIXA_ATRASO"].dropna().unique().tolist()) if "FAIXA_ATRASO" in workplan_view.columns else []
        export_faixas = st.multiselect(
            "Faixa de atraso para exportar",
            faixas_export_opcoes,
            default=[],
            key="workplan_export_faixas",
        )
    export_clientes = workplan_view.copy()
    if export_prioridades:
        export_clientes = export_clientes[export_clientes["prioridade_workplan"].isin(export_prioridades)]
    if export_motivos:
        export_clientes = export_clientes[export_clientes["motivo_abordagem"].isin(export_motivos)]
    if export_segmentos:
        export_clientes = export_clientes[export_clientes["SEGMENTO_DPD"].isin(export_segmentos)]
    if export_faixas:
        export_clientes = export_clientes[export_clientes["FAIXA_ATRASO"].isin(export_faixas)]
    export_clientes = export_clientes.sort_values(["valor_esperado_recuperacao", "score_recuperacao"], ascending=False)
    export_cols = [
        "agreement_no",
        "cust_name",
        "cpf_cnpj",
        "prioridade_workplan",
        "motivo_abordagem",
        "inadimplencia_precoce_tipo",
        "no_first_ins_unpaid",
        "probabilidade_recuperacao",
        "valor_potencial",
        "valor_esperado_recuperacao",
        "valor_esperado_aberto",
        "score_recuperacao",
        "SEGMENTO_DPD",
        "FAIXA_ATRASO",
        region_col,
        "uf",
        "dpd",
        "dias_sem_contato",
        "acionamentos_hist",
        "cpcs_hist",
        "acordos_hist",
        "pagamentos_hist",
        "status_base",
        "status_cpc",
        "city",
        "state",
    ]
    export_cols = [col for col in export_cols if col in export_clientes.columns]
    cexp1, cexp2 = st.columns([1, 2])
    with cexp1:
        metric_card("Clientes no export", num_fmt(len(export_clientes)))
    with cexp2:
        st.download_button(
            "Baixar clientes filtrados (.xlsx)",
            data=dataframe_to_excel_bytes(export_clientes[export_cols], sheet_name="Clientes"),
            file_name="clientes_prioridade_workplan.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            disabled=export_clientes.empty,
        )

    c1, c2 = st.columns(2)
    with c1:
        matriz_df = (
            workplan_view.groupby(["SEGMENTO_DPD", "FAIXA_ATRASO"], dropna=False)
            .agg(
                clientes=("CONTRATO_KEY", "nunique"),
                valor_esperado_recuperacao=("valor_esperado_recuperacao", "sum"),
                probabilidade_recuperacao=("probabilidade_recuperacao", "mean"),
            )
            .reset_index()
        )
        matriz_display = display_fields(matriz_df)
        matriz_chart = (
            alt.Chart(matriz_display)
            .mark_rect()
            .encode(
                x=alt.X("FAIXA_ATRASO:N", title="Faixa de atraso"),
                y=alt.Y("SEGMENTO_DPD:N", title="Segmento DPD"),
                color=alt.Color("valor_esperado_recuperacao:Q", scale=alt.Scale(scheme="tealblues"), title="Valor esperado"),
                tooltip=[
                    alt.Tooltip("SEGMENTO_DPD:N", title="Segmento"),
                    alt.Tooltip("FAIXA_ATRASO:N", title="Faixa"),
                    alt.Tooltip("clientes_br:N", title="Clientes"),
                    alt.Tooltip("valor_esperado_recuperacao_br:N", title="Valor esperado"),
                    alt.Tooltip("probabilidade_recuperacao_br:N", title="Chance média"),
                ],
            )
            .properties(height=320, title="Mapa de oportunidade por segmento e faixa")
        )
        stretch_altair_chart(matriz_chart)
    with c2:
        regiao_df = (
            workplan_view.groupby(region_col, dropna=False)
            .agg(
                clientes=("CONTRATO_KEY", "nunique"),
                valor_potencial=("valor_potencial", "sum"),
                valor_esperado_recuperacao=("valor_esperado_recuperacao", "sum"),
                probabilidade_recuperacao=("probabilidade_recuperacao", "mean"),
            )
            .reset_index()
            .sort_values("valor_esperado_recuperacao", ascending=False)
            .head(12)
        )
        regiao_display = display_fields(regiao_df)
        bar_chart(
            regiao_display,
            x="valor_esperado_recuperacao:Q",
            y=f"{region_col}:N",
            tooltip=[
                alt.Tooltip(f"{region_col}:N", title="Regiao/UF"),
                alt.Tooltip("clientes_br:N", title="Clientes"),
                alt.Tooltip("valor_potencial_br:N", title="Valor potencial"),
                alt.Tooltip("valor_esperado_recuperacao_br:N", title="Valor esperado"),
                alt.Tooltip("probabilidade_recuperacao_br:N", title="Chance média"),
            ],
            title="Top regiões por valor esperado",
            height=320,
        )

    chance_df = workplan_view.copy()
    chance_df["faixa_chance"] = pd.cut(
        chance_df["probabilidade_recuperacao"],
        bins=[-0.001, 0.10, 0.20, 0.35, 0.50, 0.70, 1.00],
        labels=["0-10%", "10-20%", "20-35%", "35-50%", "50-70%", "70%+"],
    )
    chance_df = (
        chance_df.groupby("faixa_chance", dropna=False)
        .agg(
            clientes=("CONTRATO_KEY", "nunique"),
            valor_potencial=("valor_potencial", "sum"),
            valor_esperado_recuperacao=("valor_esperado_recuperacao", "sum"),
        )
        .reset_index()
    )
    chance_display = display_fields(chance_df)
    chance_chart = (
        alt.Chart(chance_display)
        .mark_bar(cornerRadiusTopLeft=3, cornerRadiusTopRight=3)
        .encode(
            x=alt.X("faixa_chance:N", title="Faixa de chance"),
            y=alt.Y("clientes:Q", title="Clientes"),
            color=alt.value("#b7791f"),
            tooltip=[
                alt.Tooltip("faixa_chance:N", title="Faixa de chance"),
                alt.Tooltip("clientes_br:N", title="Clientes"),
                alt.Tooltip("valor_potencial_br:N", title="Valor potencial"),
                alt.Tooltip("valor_esperado_recuperacao_br:N", title="Valor esperado"),
            ],
        )
        .properties(height=260, title="Distribuição de clientes por chance estimada")
    )
    stretch_altair_chart(chance_chart)


def display_fields(df):
    out = df.copy()
    money_cols = [
        "valor_negociado",
        "valor_pago",
        "valor_em_aberto",
        "valor_nao_pagou",
        "valor_quebra",
        "ticket_medio",
        "meta_individual",
        "saldo_meta_individual",
        "meta_geral_escritorio",
        "total_amount_due",
        "valor_negociado_hist",
        "valor_pago_hist",
        "carteira_elegivel",
        "recuperacao_esperada",
        "valor_potencial",
        "valor_esperado_recuperacao",
        "valor_esperado_aberto",
    ]
    pct_cols = [
        "tx_contato",
        "tx_acordo",
        "tx_acordo_cliente_cpc",
        "tx_pagamento",
        "efetividade_pagamento",
        "pct_quebra",
        "tx_pagamento_cpc",
        "tx_sem_pagamento",
        "tx_cpc_acordo",
        "tx_cpc_pagamento",
        "tx_cpc_unico_acordo",
        "tx_cpc_unico_pagamento",
        "tx_acordo_pagamento",
        "tx_acordo_sem_pagamento",
        "recuperacao",
        "score",
        "atingimento_meta_individual",
        "pct_aberto_meta_individual",
        "participacao_meta_geral",
        "score_recuperacao",
        "taxa_contato",
        "taxa_cpc",
        "taxa_acordo",
        "taxa_pagamento",
        "percentual_medio_recuperado",
        "recuperacao_esperada_pct_carteira",
        "probabilidade_recuperacao",
        "prob_acordo_perfil",
        "prob_pagamento_perfil",
        "recuperacao_media_perfil",
        "risco_quebra_perfil",
        "score_eficiencia_pagamento",
    ]
    num_cols = [
        "acionamentos",
        "clientes_trabalhados",
        "contatos_efetivos",
        "contatos_cliente",
        "cpcs",
        "cpcs_unicos",
        "clientes_cpc",
        "contratos_cpc",
        "acordos",
        "pagamentos",
        "acordos_sem_pagamento",
        "acordos_em_aberto",
        "acordos_nao_pagou",
        "clientes",
        "contratos_elegiveis",
        "dpd",
        "dias_sem_contato",
        "acionamentos_hist",
        "contatos_cliente_hist",
        "cpcs_hist",
        "acordos_hist",
        "pagamentos_hist",
        "contratos_hist",
    ]
    for col in money_cols:
        if col in out:
            out[f"{col}_br"] = out[col].map(money_fmt)
    for col in pct_cols:
        if col in out:
            out[f"{col}_br"] = out[col].map(pct_fmt)
    for col in num_cols:
        if col in out:
            out[f"{col}_br"] = out[col].map(num_fmt)
    return out


def formatted_table(df):
    out = df.copy()
    for col in [
        "valor_negociado",
        "valor_pago",
        "valor_em_aberto",
        "valor_nao_pagou",
        "valor_quebra",
        "ticket_medio",
        "meta_individual",
        "saldo_meta_individual",
        "meta_geral_escritorio",
        "total_amount_due",
        "valor_negociado_hist",
        "valor_pago_hist",
        "carteira_elegivel",
        "recuperacao_esperada",
        "valor_potencial",
        "valor_esperado_recuperacao",
    ]:
        if col in out:
            out[col] = out[col].map(money_fmt)
    for col in [
        "tx_contato",
        "tx_acordo",
        "tx_acordo_cliente_cpc",
        "tx_pagamento",
        "efetividade_pagamento",
        "pct_quebra",
        "tx_pagamento_cpc",
        "tx_sem_pagamento",
        "tx_cpc_acordo",
        "tx_cpc_pagamento",
        "tx_cpc_unico_acordo",
        "tx_cpc_unico_pagamento",
        "tx_acordo_pagamento",
        "tx_acordo_sem_pagamento",
        "recuperacao",
        "score",
        "atingimento_meta_individual",
        "pct_aberto_meta_individual",
        "participacao_meta_geral",
        "score_recuperacao",
        "taxa_contato",
        "taxa_cpc",
        "taxa_acordo",
        "taxa_pagamento",
        "percentual_medio_recuperado",
        "recuperacao_esperada_pct_carteira",
        "probabilidade_recuperacao",
        "prob_acordo_perfil",
        "prob_pagamento_perfil",
        "recuperacao_media_perfil",
        "risco_quebra_perfil",
    ]:
        if col in out:
            out[col] = out[col].map(pct_fmt)
    for col in [
        "acionamentos",
        "clientes_trabalhados",
        "contatos_efetivos",
        "contatos_cliente",
        "cpcs",
        "cpcs_unicos",
        "clientes_cpc",
        "contratos_cpc",
        "acordos",
        "pagamentos",
        "acordos_sem_pagamento",
        "acordos_em_aberto",
        "acordos_nao_pagou",
        "clientes",
        "contratos_elegiveis",
        "dpd",
        "dias_sem_contato",
        "acionamentos_hist",
        "contatos_cliente_hist",
        "cpcs_hist",
        "acordos_hist",
        "pagamentos_hist",
        "contratos_hist",
    ]:
        if col in out:
            out[col] = out[col].map(num_fmt)
    return out


def help_config(df):
    config = {}
    for col in df.columns:
        if col in FIELD_HELP:
            label, help_text = FIELD_HELP[col]
            config[col] = st.column_config.TextColumn(label=label, help=help_text)
    return config


def data_table(df, **kwargs):
    formatted = formatted_table(df)
    stretch_dataframe(
        formatted,
        column_config=help_config(formatted),
        hide_index=True,
        **kwargs,
    )


def dataframe_to_excel_bytes(df, sheet_name="Workplan", extra_sheets=None):
    money_cols = {
        "valor_negociado",
        "valor_pago",
        "valor_em_aberto",
        "valor_nao_pagou",
        "valor_quebra",
        "ticket_medio",
        "meta_individual",
        "saldo_meta_individual",
        "meta_geral_escritorio",
        "total_amount_due",
        "valor_negociado_hist",
        "valor_pago_hist",
        "carteira_elegivel",
        "recuperacao_esperada",
        "valor_potencial",
        "valor_esperado_recuperacao",
        "honorarios_escritorio",
        "HONORARIOS_ESCRITORIO",
    }
    pct_cols = {
        "tx_contato",
        "tx_acordo",
        "tx_acordo_cliente_cpc",
        "tx_pagamento",
        "efetividade_pagamento",
        "pct_quebra",
        "tx_pagamento_cpc",
        "tx_sem_pagamento",
        "tx_cpc_acordo",
        "tx_cpc_pagamento",
        "tx_cpc_unico_acordo",
        "tx_cpc_unico_pagamento",
        "tx_acordo_pagamento",
        "tx_acordo_sem_pagamento",
        "recuperacao",
        "score",
        "atingimento_meta_individual",
        "pct_aberto_meta_individual",
        "participacao_meta_geral",
        "score_recuperacao",
        "taxa_contato",
        "taxa_cpc",
        "taxa_acordo",
        "taxa_pagamento",
        "percentual_medio_recuperado",
        "recuperacao_esperada_pct_carteira",
        "probabilidade_recuperacao",
        "prob_acordo_perfil",
        "prob_pagamento_perfil",
        "recuperacao_media_perfil",
        "risco_quebra_perfil",
    }
    int_cols = {
        "acionamentos",
        "clientes_trabalhados",
        "contatos_efetivos",
        "contatos_cliente",
        "cpcs",
        "cpcs_unicos",
        "clientes_cpc",
        "contratos_cpc",
        "acordos",
        "pagamentos",
        "acordos_sem_pagamento",
        "acordos_em_aberto",
        "acordos_nao_pagou",
        "clientes",
        "contratos_elegiveis",
        "dpd",
        "dias_sem_contato",
        "acionamentos_hist",
        "contatos_cliente_hist",
        "cpcs_hist",
        "acordos_hist",
        "pagamentos_hist",
        "contratos_hist",
    }

    def format_sheet(ws, source_df):
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        if source_df.empty:
            return

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        header_fill = PatternFill("solid", fgColor="213547")
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill

        for idx, col in enumerate(source_df.columns, start=1):
            letter = get_column_letter(idx)
            sample = [str(col)] + ["" if pd.isna(value) else str(value) for value in source_df[col].head(100)]
            width = min(max(max(len(value) for value in sample) + 2, 10), 42)
            ws.column_dimensions[letter].width = width

            if col in money_cols:
                number_format = 'R$ #,##0.00'
            elif col in pct_cols:
                number_format = '0.0%'
            elif col in int_cols:
                number_format = '#,##0'
            else:
                number_format = None

            if number_format:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=idx).number_format = number_format

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        for extra_sheet_name, extra_df in (extra_sheets or {}).items():
            extra_df.to_excel(writer, index=False, sheet_name=extra_sheet_name)
        format_sheet(writer.book[sheet_name], df)
        for extra_sheet_name, extra_df in (extra_sheets or {}).items():
            format_sheet(writer.book[extra_sheet_name], extra_df)
    return output.getvalue()


def glossary():
    with st.expander("Glossário dos indicadores"):
        items = [
            ("CPC", "Eventos iniciados por 02, 03, 04 ou 05."),
            ("Valor negociado", "Soma da coluna VALOR DO BANCO - META."),
            ("Valor recebido", "Valor negociado apenas dos acordos pagos; não pagos entram como R$ 0,00."),
            ("Taxa CPC -> acordo", "Contratos com CPC que geraram acordo / contratos com CPC."),
            ("Taxa CPC -> pagamento", "Contratos com CPC que geraram pagamento / contratos com CPC."),
            ("Acordos sem pagamento", "Acordos que ainda não possuem status PAGOU nem data de pagamento."),
            ("Em aberto", "Acordos ainda pendentes, com status EM ABERTO."),
            ("Não pagou", "Acordos vencidos/sem efetivação, com status NÃO PAGOU."),
            ("Recuperação", "Valor recebido / valor negociado."),
        ]
        for name, desc in items:
            st.markdown(f"**{name}:** {desc}")


eventos_raw, contratos_raw, resultados_raw = load_data(data_file_versions())
workplan_raw, workplan_error = load_workplan()
eventos, resultados, operadores_filtrados = apply_filters(eventos_raw, resultados_raw)
eventos, resultados = ensure_cpc_client_key(eventos, resultados)
eventos_raw, resultados_raw = ensure_cpc_client_key(eventos_raw, resultados_raw)
operador_df = aggregate_operator(eventos, resultados)
cpc_df = aggregate_cpc_operator(eventos, resultados)
faixa_operador_df = aggregate_operator_faixa(eventos, resultados)
workplan_df = build_workplan_analysis(workplan_raw, eventos_raw[eventos_raw["IS_ACIONAMENTO"]], resultados_raw)

st.title("Performance Operacional por Operador")
st.caption("Análise executiva de acionamentos, contatos, acordos, pagamentos, recuperação e eficiência por segmento.")
glossary()

total_clientes = eventos["CONTRATO_KEY"].nunique()
total_acionamentos = len(eventos)
total_contatos = int(eventos["IS_CONTATO_EFETIVO"].sum())
total_acordos = len(resultados)
total_pagamentos = int(resultados["IS_PAGO"].sum())
total_em_aberto = int(resultados["IS_EM_ABERTO"].sum())
total_nao_pagou = int(resultados["IS_NAO_PAGOU"].sum())
base_quebras = total_pagamentos + total_nao_pagou
valor_negociado = resultados["VALOR_NEGOCIADO"].sum()
valor_pago = resultados["VALOR_PAGO"].sum()
valor_em_aberto = resultados["VALOR_EM_ABERTO"].sum()
valor_nao_pagou = resultados["VALOR_NAO_PAGOU"].sum()
meta_geral_atual = office_goal_for_resultados(resultados)
tx_pagamento_geral = total_pagamentos / total_acordos if total_acordos else 0
efetividade_pagamento_geral = total_pagamentos / (total_pagamentos + total_nao_pagou) if (total_pagamentos + total_nao_pagou) else 0
pct_quebra_geral = valor_nao_pagou / meta_geral_atual if meta_geral_atual else 0

kpi_row1 = st.columns(5)
kpi_row2 = st.columns(6)
kpi_row3 = st.columns(3)
with kpi_row1[0]:
    metric_card("Clientes", num_fmt(total_clientes))
with kpi_row1[1]:
    metric_card("Acionamentos", num_fmt(total_acionamentos))
with kpi_row1[2]:
    metric_card("Contatos efetivos", num_fmt(total_contatos))
with kpi_row1[3]:
    metric_card("Acordos", num_fmt(total_acordos))
with kpi_row1[4]:
    metric_card("Pagamentos", num_fmt(total_pagamentos))
with kpi_row2[0]:
    metric_card("Em aberto", num_fmt(total_em_aberto))
with kpi_row2[1]:
    metric_card("NÃ£o pagou", num_fmt(total_nao_pagou))
with kpi_row2[2]:
    metric_card("Negociado", money_fmt(valor_negociado))
with kpi_row2[5]:
    metric_card("Taxa pgto", pct_fmt(tx_pagamento_geral))

with kpi_row2[3]:
    metric_card("Recebido", money_fmt(valor_pago))
with kpi_row2[4]:
    metric_card("Recuperação", pct_fmt(valor_pago / valor_negociado if valor_negociado else 0))

tabs = st.tabs(["Visão Geral", "Operadores", "CPC", "Faixa de Atraso", "DPD", "Região", "Matriz", "Metas", "Workplan", "Pagamentos", "Insights"])

with kpi_row3[0]:
    metric_card("% quebras", pct_fmt(pct_quebra_geral))
with kpi_row3[1]:
    metric_card("Valor quebras", money_fmt(valor_nao_pagou))
with kpi_row3[2]:
    metric_card("Base quebras", f"{num_fmt(total_nao_pagou)} de {num_fmt(base_quebras)}")

with tabs[0]:
    c1, c2 = st.columns([1.2, 1])
    with c1:
        top = display_fields(operador_df.head(12))
        bar_chart(
            top,
            x=alt.X("score:Q", axis=alt.Axis(format="%")),
            y="OPERADOR:N",
            tooltip=[
                "OPERADOR",
                "score_br",
                "acionamentos_br",
                "cpcs_br",
                "acordos_br",
                "pagamentos_br",
                "valor_pago_br",
            ],
            title="Ranking geral de performance",
        )
    with c2:
        funil = pd.DataFrame(
            {
                "Etapa": ["Acionamentos", "Contatos efetivos", "Acordos", "Pagamentos", "Em aberto", "Não pagou"],
                "Volume": [total_acionamentos, total_contatos, total_acordos, total_pagamentos, total_em_aberto, total_nao_pagou],
            }
        )
        bar_chart(funil, x="Volume:Q", y="Etapa:N", title="Funil operacional", sort=None, height=320)

    status_df = pd.DataFrame(
        {
            "Status": ["Pagou", "Em aberto", "Não pagou", "Outros sem pagamento"],
            "Contratos": [
                total_pagamentos,
                total_em_aberto,
                total_nao_pagou,
                max(total_acordos - total_pagamentos - total_em_aberto - total_nao_pagou, 0),
            ],
            "Valor": [
                valor_pago,
                valor_em_aberto,
                valor_nao_pagou,
                max(valor_negociado - valor_pago - valor_em_aberto - valor_nao_pagou, 0),
            ],
        }
    )
    status_df = display_fields(status_df.rename(columns={"Contratos": "clientes", "Valor": "valor_negociado"}))
    bar_chart(
        status_df,
        x="valor_negociado:Q",
        y="Status:N",
        color="Status:N",
        tooltip=["Status", "clientes_br", "valor_negociado_br"],
        title="Distribuição financeira por status",
        sort=None,
        height=260,
    )

    quebra_chart = display_fields(
        operador_df[operador_df["acordos_nao_pagou"] > 0]
        .sort_values("valor_quebra", ascending=False)
        .head(10)
    )
    bar_chart(
        quebra_chart,
        x="valor_quebra:Q",
        y="OPERADOR:N",
        color="OPERADOR:N",
        tooltip=["OPERADOR", "acordos_nao_pagou_br", "pct_quebra_br", "valor_quebra_br"],
        title="Quebras por operador",
        height=320,
    )

    by_mes_eventos = eventos.groupby("MES").agg(acionamentos=("EVENTO_TXT", "size"), contatos=("IS_CONTATO_EFETIVO", "sum")).reset_index()
    by_mes_result = resultados.groupby("MES").agg(acordos=("CONTRATO_KEY", "count"), pagamentos=("IS_PAGO", "sum")).reset_index()
    by_mes = by_mes_eventos.merge(by_mes_result, on="MES", how="outer").fillna(0)
    trend = by_mes.melt("MES", value_vars=["acionamentos", "contatos", "acordos", "pagamentos"], var_name="Indicador", value_name="Volume")
    line_chart(trend, "MES:N", "Volume:Q", "Indicador:N", "Evolução mensal")

with tabs[1]:
    c1, c2 = st.columns(2)
    with c1:
        chart_operador = display_fields(operador_df.head(15))
        chart_operador["OPERADOR_LABEL"] = chart_operador["OPERADOR"].map(format_operator_label)
        bar_chart(
            chart_operador,
            x="valor_pago:Q",
            y=alt.Y("OPERADOR_LABEL:N", title=None, sort="-x", axis=alt.Axis(labelLimit=200)),
            tooltip=[
                alt.Tooltip("OPERADOR_LABEL:N", title="Operador"),
                "valor_pago_br",
                "pagamentos_br",
                "tx_pagamento_br",
                "score_br",
            ],
            title="Valor recuperado por operador",
        )
    with c2:
        volume_eficiencia = display_fields(operador_df[operador_df["acionamentos"] > 0])
        volume_eficiencia["OPERADOR_LABEL"] = volume_eficiencia["OPERADOR"].map(format_operator_label)
        scatter = (
            alt.Chart(volume_eficiencia)
            .mark_circle(size=120, opacity=0.78)
            .encode(
                x=alt.X("acionamentos:Q", title="Acionamentos"),
                y=alt.Y("tx_acordo:Q", title="Conversão contato/acordo", axis=alt.Axis(format="%")),
                size=alt.Size("valor_pago:Q", title="Valor recebido"),
                color=alt.Color("tx_pagamento:Q", scale=alt.Scale(scheme="tealblues"), title="Acordo/pagamento"),
                tooltip=[
                    alt.Tooltip("OPERADOR_LABEL:N", title="Operador"),
                    "acionamentos_br",
                    "cpcs_br",
                    "acordos_br",
                    "pagamentos_br",
                    "valor_pago_br",
                    "tx_acordo_br",
                    "tx_pagamento_br",
                ],
            )
            .properties(height=320, title="Volume versus eficiência")
        )
        stretch_altair_chart(scatter)

    st.subheader("Ranking detalhado")
    cols = [
        "OPERADOR",
        "clientes_trabalhados",
        "acionamentos",
        "cpcs",
        "clientes_cpc",
        "acordos",
        "pagamentos",
        "acordos_sem_pagamento",
        "acordos_em_aberto",
        "acordos_nao_pagou",
        "pct_quebra",
        "tx_contato",
        "tx_acordo",
        "tx_pagamento_cpc",
        "tx_pagamento",
        "efetividade_pagamento",
        "tx_sem_pagamento",
        "valor_negociado",
        "valor_pago",
        "valor_em_aberto",
        "valor_nao_pagou",
        "valor_quebra",
        "ticket_medio",
        "recuperacao",
        "score",
    ]
    data_table(operador_df[cols])

    st.markdown("---")
    st.subheader("Melhor faixa de atraso por operador")
    st.caption(
        "Para cada operador, a faixa de atraso (DPD) onde ele tem a melhor taxa de contato → CPC "
        "e a melhor conversão CPC → pagamento. Mínimo de 5 CPCs na faixa para entrar no ranking."
    )
    melhor_faixa_df = melhor_faixa_por_operador(faixa_operador_df, min_cpcs=5)
    if melhor_faixa_df.empty:
        st.info("Nenhum operador atingiu o mínimo de CPCs por faixa para esse recorte de filtros.")
    else:
        stretch_dataframe(melhor_faixa_df, hide_index=True)

        top_n_faixa = 9
        destaque_operadores = operador_df.head(top_n_faixa)["OPERADOR"].tolist()
        destaque_labels = [format_operator_label(op) for op in destaque_operadores]

        pequenos_multiplos = faixa_operador_df[
            faixa_operador_df["OPERADOR"].isin(destaque_operadores) & (faixa_operador_df["cpcs"] >= 5)
        ].copy()
        if pequenos_multiplos.empty:
            st.info("Sem faixas com volume mínimo de CPCs para os operadores em destaque.")
        else:
            pequenos_multiplos["OPERADOR_LABEL"] = pequenos_multiplos["OPERADOR"].map(format_operator_label)
            melhor_idx = (
                pequenos_multiplos.sort_values("tx_pagamento_cpc", ascending=False).groupby("OPERADOR").head(1).index
            )
            pequenos_multiplos["destaque"] = "Demais faixas"
            pequenos_multiplos.loc[melhor_idx, "destaque"] = "Melhor faixa"
            pequenos_multiplos["rotulo_melhor"] = np.where(
                pequenos_multiplos["destaque"].eq("Melhor faixa"),
                pequenos_multiplos["tx_pagamento_cpc"].map(pct_fmt),
                "",
            )

            barras = (
                alt.Chart(pequenos_multiplos)
                .mark_bar(cornerRadiusTopLeft=3, cornerRadiusTopRight=3, size=18)
                .encode(
                    x=alt.X("FAIXA_ATRASO:N", title=None, sort=FAIXA_ATRASO_ORDER, axis=alt.Axis(labelAngle=0, labelFontSize=9)),
                    y=alt.Y("tx_pagamento_cpc:Q", title=None, axis=alt.Axis(format="%", labelFontSize=9)),
                    color=alt.Color(
                        "destaque:N",
                        scale=alt.Scale(domain=["Melhor faixa", "Demais faixas"], range=[CORP_PALETTE[1], CORP_PALETTE[3]]),
                        legend=alt.Legend(title=None, orient="top"),
                    ),
                    tooltip=[
                        alt.Tooltip("OPERADOR_LABEL:N", title="Operador"),
                        alt.Tooltip("FAIXA_ATRASO:N", title="Faixa"),
                        alt.Tooltip("cpcs:Q", title="CPCs", format=","),
                        alt.Tooltip("pagamentos:Q", title="Pagamentos", format=","),
                        alt.Tooltip("tx_pagamento_cpc:Q", title="CPC → pagamento", format=".1%"),
                    ],
                )
            )
            rotulos = (
                alt.Chart(pequenos_multiplos)
                .mark_text(dy=-6, fontSize=9, fontWeight="bold", color=CORP_PALETTE[1])
                .encode(
                    x=alt.X("FAIXA_ATRASO:N", sort=FAIXA_ATRASO_ORDER),
                    y=alt.Y("tx_pagamento_cpc:Q"),
                    text="rotulo_melhor:N",
                )
            )
            pequenos_multiplos_chart = (
                (barras + rotulos)
                .properties(width=190, height=130)
                .facet(
                    facet=alt.Facet("OPERADOR_LABEL:N", title=None, sort=destaque_labels),
                    columns=3,
                )
                .resolve_scale(y="independent")
                .properties(title=f"Conversão CPC → pagamento por faixa de atraso (top {top_n_faixa} operadores)")
            )
            st.caption("Barra destacada = faixa com melhor conversão CPC → pagamento do operador. Faixas sem barra: menos de 5 CPCs, sem base.")
            stretch_altair_chart(pequenos_multiplos_chart)

with tabs[2]:
    st.subheader("Conversão CPC para acordos e pagamentos")
    st.caption("CPC considerado pelos eventos iniciados por 02, 03, 04 e 05. O KPI geral remove duplicidade por CPF/CNPJ.")

    cpc_eventos_geral = (
        eventos[eventos["IS_CPC"]]
        .dropna(subset=["CPC_CLIENT_KEY"])
        [["CPC_CLIENT_KEY"]]
        .drop_duplicates()
    )
    resultado_cliente_geral = (
        resultados.dropna(subset=["CPC_CLIENT_KEY"])
        .groupby("CPC_CLIENT_KEY", dropna=True)
        .agg(
            qtd_acordos=("CONTRATO_KEY", "count"),
            teve_pagamento=("IS_PAGO", "max"),
            teve_em_aberto=("IS_EM_ABERTO", "max"),
            teve_nao_pagou=("IS_NAO_PAGOU", "max"),
        )
        .reset_index()
    )
    cpc_resultado_geral = cpc_eventos_geral.merge(resultado_cliente_geral, on="CPC_CLIENT_KEY", how="left")
    cpc_resultado_geral["qtd_acordos"] = cpc_resultado_geral["qtd_acordos"].fillna(0)
    cpc_resultado_geral["teve_pagamento"] = cpc_resultado_geral["teve_pagamento"].fillna(False).astype(bool)
    cpc_resultado_geral["teve_em_aberto"] = cpc_resultado_geral["teve_em_aberto"].fillna(False).astype(bool)
    cpc_resultado_geral["teve_nao_pagou"] = cpc_resultado_geral["teve_nao_pagou"].fillna(False).astype(bool)

    _unicos_total = float(len(cpc_eventos_geral))
    _acordos_geral = int((cpc_resultado_geral["qtd_acordos"] > 0).sum())
    _pagamentos_geral = int(cpc_resultado_geral["teve_pagamento"].sum())
    _em_aberto_geral = int(cpc_resultado_geral["teve_em_aberto"].sum())
    _nao_pagou_geral = int(cpc_resultado_geral["teve_nao_pagou"].sum())
    _tx_acordo_geral = _acordos_geral / _unicos_total if _unicos_total else 0
    _tx_pag_geral = _pagamentos_geral / _unicos_total if _unicos_total else 0

    c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
    with c1:
        metric_card("CPCs únicos", num_fmt(_unicos_total))
    with c2:
        metric_card("Tx CPC único → acordo", pct_fmt(_tx_acordo_geral))
    with c3:
        metric_card("Tx CPC único → pagto", pct_fmt(_tx_pag_geral))
    with c4:
        metric_card("Acordos após CPC", num_fmt(_acordos_geral))
    with c5:
        metric_card("Pagamentos", num_fmt(_pagamentos_geral))
    with c6:
        metric_card("Em aberto", num_fmt(_em_aberto_geral))
    with c7:
        metric_card("Não pagou", num_fmt(_nao_pagou_geral))

    c1, c2 = st.columns(2)
    with c1:
        cpc_chart = display_fields(cpc_df[cpc_df["cpcs_unicos"] > 0].sort_values("tx_cpc_unico_acordo", ascending=False).head(15))
        bar_chart(
            cpc_chart,
            x=alt.X("tx_cpc_unico_acordo:Q", axis=alt.Axis(format="%")),
            y="OPERADOR:N",
            tooltip=["OPERADOR", "cpcs_unicos_br", "acordos_br", "tx_cpc_unico_acordo_br", "valor_negociado_br"],
            title="Conversão CPC único → acordo por negociador",
        )
    with c2:
        cpc_pag_chart = display_fields(cpc_df[cpc_df["cpcs_unicos"] > 0].sort_values("tx_cpc_unico_pagamento", ascending=False).head(15))
        bar_chart(
            cpc_pag_chart,
            x=alt.X("tx_cpc_unico_pagamento:Q", axis=alt.Axis(format="%")),
            y="OPERADOR:N",
            tooltip=["OPERADOR", "cpcs_unicos_br", "pagamentos_br", "tx_cpc_unico_pagamento_br", "valor_pago_br"],
            title="Conversão CPC único → pagamento por negociador",
        )

    c1, c2 = st.columns(2)
    with c1:
        sem_pg = display_fields(cpc_df[cpc_df["acordos_sem_pagamento"] > 0].sort_values("acordos_sem_pagamento", ascending=False).head(15))
        bar_chart(
            sem_pg,
            x="acordos_sem_pagamento:Q",
            y="OPERADOR:N",
            tooltip=[
                "OPERADOR",
                "acordos_br",
                "acordos_sem_pagamento_br",
                "tx_acordo_sem_pagamento_br",
                "valor_negociado_br",
            ],
            title="Acordos convertidos sem pagamento",
        )
    with c2:
        cpc_scatter = display_fields(cpc_df[cpc_df["cpcs_unicos"] > 0])
        scatter = (
            alt.Chart(cpc_scatter)
            .mark_circle(size=130, opacity=0.78)
            .encode(
                x=alt.X("tx_cpc_unico_acordo:Q", title="CPC único → acordo", axis=alt.Axis(format="%")),
                y=alt.Y("tx_cpc_unico_pagamento:Q", title="CPC único → pagamento", axis=alt.Axis(format="%")),
                size=alt.Size("valor_pago:Q", title="Valor recebido"),
                color=alt.Color("quartil_cpc_acordo:N", title="Quartil"),
                tooltip=[
                    "OPERADOR",
                    "cpcs_unicos_br",
                    "acordos_br",
                    "pagamentos_br",
                    "acordos_sem_pagamento_br",
                    "tx_cpc_unico_acordo_br",
                    "tx_cpc_unico_pagamento_br",
                    "efetividade_pagamento_br",
                    "valor_pago_br",
                    "quartil_cpc_acordo",
                ],
            )
            .properties(height=320, title="Qualidade da conversão (CPC único → acordo × CPC único → pagamento)")
        )
        stretch_altair_chart(scatter)

    st.subheader("Quartil de conversão CPC")
    st.caption("Classificação pelo percentual de CPCs únicos que resultaram em pagamento. Mínimo de 3 CPCs únicos para entrar no ranking.")

    diagnostico_cpc_resumo = cpc_df["diagnostico_cpc"].value_counts().reset_index()
    diagnostico_cpc_resumo.columns = ["Diagnóstico", "Operadores"]
    c1, c2 = st.columns([1, 1.3])
    with c1:
        st.subheader("Resumo por diagnóstico")
        stretch_dataframe(diagnostico_cpc_resumo, hide_index=True)
    with c2:
        chart_cpc_quartil = display_fields(
            cpc_df[cpc_df["cpcs_unicos"] >= 3]
            .sort_values("tx_cpc_unico_pagamento", ascending=False)
            .head(15)
        )
        bar_chart(
            chart_cpc_quartil,
            x=alt.X("tx_cpc_unico_pagamento:Q", axis=alt.Axis(format="%")),
            y="OPERADOR:N",
            color="diagnostico_cpc:N",
            tooltip=[
                "OPERADOR",
                "cpcs_unicos_br",
                "tx_cpc_unico_acordo_br",
                "tx_cpc_unico_pagamento_br",
                "efetividade_pagamento_br",
                "valor_pago_br",
                "diagnostico_cpc",
            ],
            title="Conversão CPC único → pagamento por diagnóstico",
            height=360,
        )

    st.subheader("Operadores em cada quartil")
    grupos_cpc = cpc_operator_groups(cpc_df)
    stretch_dataframe(
        grupos_cpc,
        hide_index=True,
        column_config={
            "Métrica": st.column_config.TextColumn("Métrica", help="Indicador usado para montar o quartil."),
            "Grupo": st.column_config.TextColumn("Grupo", help="Grupo do quartil. Q4 é destaque; Q1 é crítico."),
            "Qtd. operadores": st.column_config.NumberColumn("Qtd. operadores"),
            "Operadores": st.column_config.TextColumn("Operadores"),
        },
    )

    st.subheader("Análise descritiva por quartil")
    st.caption("Médias calculadas sobre os operadores com pelo menos 3 CPCs únicos dentro do quartil. Totais são a soma consolidada de todos os operadores do grupo.")
    desc_df = cpc_quartil_descritivo(cpc_df)
    if not desc_df.empty:
        stretch_dataframe(desc_df, hide_index=True)

        _descricoes_quartil = {
            "Q4 - destaque": "Representam o padrão de excelência do grupo. Vale mapear as boas práticas desses operadores para disseminar à equipe.",
            "Q3 - bom": "Acima da mediana; com foco na qualidade da negociação e acompanhamento do funil pós-CPC têm potencial para alcançar o grupo destaque.",
            "Q2 - atenção": "Abaixo da mediana. Recomenda-se analisar as perdas entre CPC e fechamento de acordo: script de abordagem, timing de retorno e perfil da carteira trabalhada.",
            "Q1 - crítico": "Conversão crítica. Necessita acompanhamento gerencial individualizado, capacitação direcionada e revisão da estratégia de negociação.",
            "Sem base": "Menos de 3 CPCs únicos no período filtrado; sem base estatística para classificação no ranking.",
        }
        for _, row in desc_df.iterrows():
            grupo = row["Quartil"]
            ops = int(row["Qtd. operadores"])
            pag_media = row["CPC → pgto (média)"]
            descricao = _descricoes_quartil.get(grupo, "")
            if "Sem base" in grupo:
                st.markdown(f"**{grupo}** — {ops} operador(es). {descricao}")
            else:
                st.markdown(f"**{grupo}** ({ops} operador(es)) — média CPC → pagamento: **{pag_media}**. {descricao}")

    st.subheader("Quartil de volume CPC")
    st.caption("Classificação pela quantidade de CPCs únicos gerados no período. Mínimo de 1 CPC único para entrar no ranking.")

    _vq4 = cpc_df[cpc_df["quartil_cpc_volume"].eq("Q4 - destaque")]
    _vq3 = cpc_df[cpc_df["quartil_cpc_volume"].eq("Q3 - bom")]
    _vq2 = cpc_df[cpc_df["quartil_cpc_volume"].eq("Q2 - atenção")]
    _vq1 = cpc_df[cpc_df["quartil_cpc_volume"].eq("Q1 - crítico")]
    _vsb = cpc_df[cpc_df["quartil_cpc_volume"].eq("Sem base")]
    _vbase = cpc_df[cpc_df["cpcs_unicos"] >= 1]

    kv1, kv2, kv3, kv4, kv5, kv6 = st.columns(6)
    with kv1:
        metric_card("Ranqueados", num_fmt(len(cpc_df) - len(_vsb)))
    with kv2:
        metric_card("Q4 — Destaque", num_fmt(len(_vq4)))
    with kv3:
        metric_card("Q3 — Bom", num_fmt(len(_vq3)))
    with kv4:
        metric_card("Q2 — Atenção", num_fmt(len(_vq2)))
    with kv5:
        metric_card("Q1 — Crítico", num_fmt(len(_vq1)))
    with kv6:
        metric_card("Sem base", num_fmt(len(_vsb)))

    kv7, kv8, kv9, kv10 = st.columns(4)
    with kv7:
        metric_card("CPCs únicos (total)", num_fmt(cpc_df["cpcs_unicos"].sum()))
    with kv8:
        _media_geral = _vbase["cpcs_unicos"].mean() if not _vbase.empty else float("nan")
        metric_card("Média geral CPCs únicos", num_fmt(_media_geral) if not pd.isna(_media_geral) else "—")
    with kv9:
        _media_q4 = _vq4["cpcs_unicos"].mean() if not _vq4.empty else float("nan")
        metric_card("Média CPCs únicos (Q4)", num_fmt(_media_q4) if not pd.isna(_media_q4) else "—")
    with kv10:
        metric_card("Valor recebido (Q4)", money_fmt(_vq4["valor_pago"].sum()))

    c1, c2 = st.columns(2)
    with c1:
        chart_cpc_vol = display_fields(
            cpc_df[cpc_df["cpcs_unicos"] >= 1]
            .sort_values("cpcs_unicos", ascending=False)
            .head(15)
        )
        bar_chart(
            chart_cpc_vol,
            x=alt.X("cpcs_unicos:Q"),
            y="OPERADOR:N",
            color="quartil_cpc_volume:N",
            tooltip=[
                "OPERADOR",
                "cpcs_unicos_br",
                "acordos_br",
                "pagamentos_br",
                "tx_cpc_unico_acordo_br",
                "tx_cpc_unico_pagamento_br",
                "valor_pago_br",
                "quartil_cpc_volume",
            ],
            title="Volume de CPCs únicos por operador (top 15)",
            height=360,
        )
    with c2:
        _scatter_vol = display_fields(cpc_df[cpc_df["cpcs_unicos"] >= 1])
        if not _scatter_vol.empty:
            scatter_vol_chart = (
                alt.Chart(_scatter_vol)
                .mark_circle(size=90, opacity=0.85)
                .encode(
                    x=alt.X("cpcs_unicos:Q", title="CPCs únicos"),
                    y=alt.Y("valor_pago:Q", title="Valor recebido (R$)"),
                    color=alt.Color("quartil_cpc_volume:N", title="Quartil"),
                    tooltip=[
                        "OPERADOR",
                        "cpcs_unicos_br",
                        "valor_pago_br",
                        "acordos_br",
                        "pagamentos_br",
                        "tx_cpc_unico_pagamento_br",
                        "quartil_cpc_volume",
                    ],
                )
                .properties(height=360, title="CPCs únicos × Valor recebido por quartil")
            )
            stretch_altair_chart(scatter_vol_chart)
        else:
            st.info("Sem dados para o gráfico de dispersão.")

    vol_resumo = (
        cpc_df["quartil_cpc_volume"]
        .value_counts()
        .reindex(["Q4 - destaque", "Q3 - bom", "Q2 - atenção", "Q1 - crítico", "Sem base"])
        .dropna()
        .reset_index()
    )
    vol_resumo.columns = ["Quartil", "Operadores"]
    stretch_dataframe(vol_resumo, hide_index=True)

    st.subheader("Análise descritiva por volume de CPC")
    st.caption("Médias e totais calculados sobre os operadores de cada quartil com ao menos 1 CPC único.")
    desc_vol_df = cpc_volume_quartil_descritivo(cpc_df)
    if not desc_vol_df.empty:
        stretch_dataframe(desc_vol_df, hide_index=True)

        _descricoes_volume_quartil = {
            "Q4 - destaque": "Alta produtividade de contato produtivo; esses operadores lideram em volume de CPCs únicos. Vale mapear a abordagem e cadência de acionamento desse grupo para disseminar à equipe.",
            "Q3 - bom": "Volume acima da mediana; bom ritmo de contatos produtivos. Com incremento na frequência de follow-up há espaço para alcançar o grupo destaque.",
            "Q2 - atenção": "Volume abaixo da mediana. Recomenda-se revisar a gestão da carteira ativa e a frequência de acionamentos para ampliar o número de contatos produtivos.",
            "Q1 - crítico": "Volume crítico de CPCs únicos. Necessita acompanhamento gerencial para identificar bloqueios operacionais: carteira insuficiente, ausências ou baixa eficiência nos acionamentos.",
            "Sem base": "Sem CPCs únicos no período filtrado; operadores sem contatos produtivos registrados.",
        }
        for _, row in desc_vol_df.iterrows():
            grupo = row["Quartil"]
            ops = int(row["Qtd. operadores"])
            media_cpc = row["CPCs únicos (média)"]
            descricao = _descricoes_volume_quartil.get(grupo, "")
            if "Sem base" in grupo:
                st.markdown(f"**{grupo}** — {ops} operador(es). {descricao}")
            else:
                st.markdown(f"**{grupo}** ({ops} operador(es)) — média CPCs únicos: **{media_cpc}**. {descricao}")

    st.subheader("Tabela analítica CPC")
    cpc_cols = [
        "OPERADOR",
        "cpcs_unicos",
        "acordos",
        "pagamentos",
        "acordos_sem_pagamento",
        "acordos_em_aberto",
        "acordos_nao_pagou",
        "pct_quebra",
        "tx_cpc_unico_acordo",
        "tx_cpc_unico_pagamento",
        "tx_acordo_pagamento",
        "efetividade_pagamento",
        "tx_acordo_sem_pagamento",
        "valor_negociado",
        "valor_pago",
        "valor_em_aberto",
        "valor_nao_pagou",
        "valor_quebra",
        "ticket_medio",
        "recuperacao",
        "quartil_cpc_volume",
        "quartil_cpc_acordo",
        "quartil_cpc_pagamento",
        "diagnostico_cpc",
    ]
    data_table(cpc_df[cpc_cols])

with tabs[3]:
    pos_retomado_contratos = set(resultados.loc[resultados["CAMPANHA"].map(is_pos_retomado), "CONTRATO_KEY"])
    resultados_faixa = resultados[~resultados["CAMPANHA"].map(is_pos_retomado)].copy()
    eventos_faixa = eventos[~eventos["CONTRATO_KEY"].isin(pos_retomado_contratos)].copy()
    st.caption("Contratos da campanha Pós Retomado são excluídos desta análise por faixa de atraso.")

    faixa_df = aggregate_resultados(resultados_faixa, "FAIXA_ATRASO")
    ev_faixa = eventos_faixa.groupby("FAIXA_ATRASO").agg(acionamentos=("EVENTO_TXT", "size"), contatos_efetivos=("IS_CONTATO_EFETIVO", "sum")).reset_index()
    faixa_df = faixa_df.merge(ev_faixa, on="FAIXA_ATRASO", how="outer").fillna(0)
    faixa_df["tx_contato"] = safe_div(faixa_df["contatos_efetivos"], faixa_df["acionamentos"])
    faixa_chart = display_fields(faixa_df)

    # ── KPIs ──────────────────────────────────────────────────────────────────
    _f_faixas_ativas = int(faixa_df[faixa_df["acordos"] > 0]["FAIXA_ATRASO"].nunique())
    _f_acordos_total = int(faixa_df["acordos"].sum())
    _f_pago_total = faixa_df["valor_pago"].sum()
    _f_pag_sum = float(faixa_df["pagamentos"].sum())
    _f_npag_sum = float(faixa_df["acordos_nao_pagou"].sum())
    _f_efetividade = _f_pag_sum / (_f_pag_sum + _f_npag_sum) if (_f_pag_sum + _f_npag_sum) > 0 else 0
    _f_melhor = str(faixa_df.sort_values("valor_pago", ascending=False).iloc[0]["FAIXA_ATRASO"]) if not faixa_df.empty else "-"
    _f_maior_volume = str(faixa_df.sort_values("acordos", ascending=False).iloc[0]["FAIXA_ATRASO"]) if not faixa_df.empty else "-"

    kf1, kf2, kf3, kf4, kf5 = st.columns(5)
    with kf1:
        metric_card("Faixas com acordos", num_fmt(_f_faixas_ativas))
    with kf2:
        metric_card("Total de acordos", num_fmt(_f_acordos_total))
    with kf3:
        metric_card("Valor total recebido", money_fmt(_f_pago_total))
    with kf4:
        metric_card("Efetividade de pagamento", pct_fmt(_f_efetividade))
    with kf5:
        metric_card("Faixa líder (valor)", _f_melhor)

    st.markdown("---")

    c1, c2 = st.columns(2)
    with c1:
        bar_chart(
            faixa_chart,
            x="valor_pago:Q",
            y="FAIXA_ATRASO:N",
            tooltip=["FAIXA_ATRASO", "valor_pago_br", "valor_negociado_br", "acordos_br", "pagamentos_br"],
            title="Valor recuperado por faixa",
            sort=None,
        )
    with c2:
        bar_chart(
            faixa_chart,
            x=alt.X("tx_pagamento:Q", axis=alt.Axis(format="%")),
            y="FAIXA_ATRASO:N",
            tooltip=["FAIXA_ATRASO", "tx_pagamento_br", "acordos_br", "pagamentos_br", "recuperacao_br"],
            title="Conversão acordo → pagamento por faixa",
            sort=None,
        )

    c3, c4 = st.columns(2)
    with c3:
        bar_chart(
            faixa_chart,
            x="ticket_medio:Q",
            y="FAIXA_ATRASO:N",
            tooltip=["FAIXA_ATRASO", "ticket_medio_br", "acordos_br", "valor_negociado_br"],
            title="Ticket médio por faixa",
            sort=None,
        )
    with c4:
        bar_chart(
            faixa_chart,
            x=alt.X("efetividade_pagamento:Q", axis=alt.Axis(format="%")),
            y="FAIXA_ATRASO:N",
            tooltip=["FAIXA_ATRASO", "efetividade_pagamento_br", "pagamentos_br", "acordos_nao_pagou_br", "valor_quebra_br"],
            title="Efetividade de pagamento por faixa",
            sort=None,
        )

    best_faixa = (
        resultados_faixa.groupby(["FAIXA_ATRASO", "OPERADOR"])
        .agg(
            acordos=("CONTRATO_KEY", "count"),
            pagamentos=("IS_PAGO", "sum"),
            acordos_em_aberto=("IS_EM_ABERTO", "sum"),
            acordos_nao_pagou=("IS_NAO_PAGOU", "sum"),
            valor_pago=("VALOR_PAGO", "sum"),
            valor_em_aberto=("VALOR_EM_ABERTO", "sum"),
            valor_nao_pagou=("VALOR_NAO_PAGOU", "sum"),
        )
        .reset_index()
    )
    best_faixa["tx_pagamento"] = safe_div(best_faixa["pagamentos"], best_faixa["acordos"])
    best_faixa["efetividade_pagamento"] = safe_div(best_faixa["pagamentos"], best_faixa["pagamentos"] + best_faixa["acordos_nao_pagou"])
    best_faixa["valor_quebra"] = best_faixa["valor_nao_pagou"]
    best_faixa["meta_individual"] = operator_goal_series(best_faixa["OPERADOR"], selected_months_count(resultados_faixa))
    best_faixa["pct_quebra"] = safe_div(best_faixa["valor_quebra"], best_faixa["meta_individual"])
    best_faixa = best_faixa.sort_values(["FAIXA_ATRASO", "valor_pago", "tx_pagamento"], ascending=[True, False, False]).groupby("FAIXA_ATRASO").head(1)
    st.subheader("Melhor operador por faixa")
    data_table(best_faixa)

with tabs[4]:
    resultados_segmento = resultados[resultados["SEGMENTO_DPD"].ne("Sem DPD")].copy()
    eventos_segmento = eventos[eventos["SEGMENTO_DPD"].ne("Sem DPD")].copy()
    segmento_df = aggregate_resultados(resultados_segmento, "SEGMENTO_DPD")
    ev_segmento = eventos_segmento.groupby("SEGMENTO_DPD").agg(
        acionamentos=("EVENTO_TXT", "size"),
        contatos_efetivos=("IS_CONTATO_EFETIVO", "sum"),
    ).reset_index()
    segmento_df = segmento_df.merge(ev_segmento, on="SEGMENTO_DPD", how="outer").fillna(0)
    segmento_df["tx_contato"] = safe_div(segmento_df["contatos_efetivos"], segmento_df["acionamentos"])
    segmento_order = ["POTLOSS", "SALVAGE", "SALVAGE +"]
    segmento_df["SEGMENTO_DPD"] = pd.Categorical(segmento_df["SEGMENTO_DPD"], categories=segmento_order, ordered=True)
    segmento_df = segmento_df.sort_values("SEGMENTO_DPD")
    segmento_chart = display_fields(segmento_df)

    # ── KPIs por segmento ────────────────────────────────────────────────────
    def _seg(col, seg):
        row = segmento_df[segmento_df["SEGMENTO_DPD"].astype(str).eq(seg)]
        return row[col].iloc[0] if not row.empty else 0

    ks1, ks2, ks3, ks4, ks5 = st.columns(5)
    with ks1:
        metric_card("POTLOSS — Recebido", money_fmt(_seg("valor_pago", "POTLOSS")))
    with ks2:
        metric_card("SALVAGE — Recebido", money_fmt(_seg("valor_pago", "SALVAGE")))
    with ks3:
        metric_card("SALVAGE+ — Recebido", money_fmt(_seg("valor_pago", "SALVAGE +")))
    with ks4:
        _dpd_pag = float(segmento_df["pagamentos"].sum())
        _dpd_npag = float(segmento_df["acordos_nao_pagou"].sum())
        metric_card("Efetividade geral DPD", pct_fmt(_dpd_pag / (_dpd_pag + _dpd_npag) if (_dpd_pag + _dpd_npag) else 0))
    with ks5:
        _dpd_melhor = segmento_df.sort_values("recuperacao", ascending=False).iloc[0]["SEGMENTO_DPD"] if not segmento_df.empty else "-"
        metric_card("Melhor % recuperação", str(_dpd_melhor))

    st.markdown("---")

    c1, c2 = st.columns(2)
    with c1:
        bar_chart(
            segmento_chart,
            x="valor_pago:Q",
            y="SEGMENTO_DPD:N",
            tooltip=["SEGMENTO_DPD", "valor_pago_br", "valor_negociado_br", "acordos_br", "pagamentos_br", "recuperacao_br"],
            title="Valor recebido por segmento DPD",
            sort=segmento_order,
        )
    with c2:
        bar_chart(
            segmento_chart,
            x=alt.X("tx_pagamento:Q", axis=alt.Axis(format="%")),
            y="SEGMENTO_DPD:N",
            tooltip=["SEGMENTO_DPD", "tx_pagamento_br", "acordos_br", "pagamentos_br", "tx_contato_br"],
            title="Conversão acordo → pagamento por segmento DPD",
            sort=segmento_order,
        )

    c3, c4 = st.columns(2)
    with c3:
        bar_chart(
            segmento_chart,
            x="ticket_medio:Q",
            y="SEGMENTO_DPD:N",
            tooltip=["SEGMENTO_DPD", "ticket_medio_br", "acordos_br", "valor_negociado_br"],
            title="Ticket médio por segmento DPD",
            sort=segmento_order,
        )
    with c4:
        bar_chart(
            segmento_chart,
            x=alt.X("efetividade_pagamento:Q", axis=alt.Axis(format="%")),
            y="SEGMENTO_DPD:N",
            tooltip=["SEGMENTO_DPD", "efetividade_pagamento_br", "pagamentos_br", "acordos_nao_pagou_br"],
            title="Efetividade de pagamento por segmento DPD",
            sort=segmento_order,
        )

    best_segmento = (
        resultados_segmento.groupby(["SEGMENTO_DPD", "OPERADOR"])
        .agg(
            acordos=("CONTRATO_KEY", "count"),
            pagamentos=("IS_PAGO", "sum"),
            acordos_em_aberto=("IS_EM_ABERTO", "sum"),
            acordos_nao_pagou=("IS_NAO_PAGOU", "sum"),
            valor_pago=("VALOR_PAGO", "sum"),
            valor_em_aberto=("VALOR_EM_ABERTO", "sum"),
            valor_nao_pagou=("VALOR_NAO_PAGOU", "sum"),
            valor_negociado=("VALOR_NEGOCIADO", "sum"),
        )
        .reset_index()
    )
    best_segmento["tx_pagamento"] = safe_div(best_segmento["pagamentos"], best_segmento["acordos"])
    best_segmento["efetividade_pagamento"] = safe_div(best_segmento["pagamentos"], best_segmento["pagamentos"] + best_segmento["acordos_nao_pagou"])
    best_segmento["valor_quebra"] = best_segmento["valor_nao_pagou"]
    best_segmento["meta_individual"] = operator_goal_series(best_segmento["OPERADOR"], selected_months_count(resultados_segmento))
    best_segmento["pct_quebra"] = safe_div(best_segmento["valor_quebra"], best_segmento["meta_individual"])
    best_segmento["recuperacao"] = safe_div(best_segmento["valor_pago"], best_segmento["valor_negociado"])
    best_segmento = best_segmento.sort_values(["SEGMENTO_DPD", "valor_pago", "tx_pagamento"], ascending=[True, False, False]).groupby("SEGMENTO_DPD").head(3)
    st.subheader("Top operadores por segmento DPD")
    data_table(best_segmento)

    st.subheader("Resumo por segmento DPD")
    data_table(segmento_df[["SEGMENTO_DPD", "clientes", "acionamentos", "contatos_efetivos", "tx_contato", "acordos", "pagamentos", "acordos_em_aberto", "acordos_nao_pagou", "pct_quebra", "efetividade_pagamento", "valor_negociado", "valor_pago", "valor_em_aberto", "valor_nao_pagou", "valor_quebra", "recuperacao"]])

with tabs[5]:
    regiao_df = aggregate_resultados(resultados, "REGIÃO").sort_values("valor_pago", ascending=False)
    regiao_chart = display_fields(regiao_df)

    # ── KPIs ──────────────────────────────────────────────────────────────────
    _r_total_regioes = int(regiao_df[regiao_df["acordos"] > 0]["REGIÃO"].nunique())
    _r_pago_total = regiao_df["valor_pago"].sum()
    _r_top_reg = str(regiao_df.iloc[0]["REGIÃO"]) if not regiao_df.empty else "-"
    _r_top_pago = regiao_df.iloc[0]["valor_pago"] if not regiao_df.empty else 0
    _r_top3_pct = regiao_df.head(3)["valor_pago"].sum() / _r_pago_total if _r_pago_total else 0
    _r_rec_media = float(regiao_df["recuperacao"].mean()) if not regiao_df.empty else 0

    kr1, kr2, kr3, kr4, kr5 = st.columns(5)
    with kr1:
        metric_card("Regiões com acordos", num_fmt(_r_total_regioes))
    with kr2:
        metric_card("Região líder", _r_top_reg)
    with kr3:
        metric_card("Valor recebido (líder)", money_fmt(_r_top_pago))
    with kr4:
        metric_card("Concentração top 3", pct_fmt(_r_top3_pct))
    with kr5:
        metric_card("% recuperação média", pct_fmt(_r_rec_media))

    st.markdown("---")

    c1, c2 = st.columns(2)
    with c1:
        bar_chart(
            regiao_chart,
            x="valor_pago:Q",
            y="REGIÃO:N",
            tooltip=["REGIÃO", "valor_pago_br", "valor_negociado_br", "acordos_br", "pagamentos_br"],
            title="Valor recebido por região",
        )
    with c2:
        bar_chart(
            regiao_chart,
            x=alt.X("recuperacao:Q", axis=alt.Axis(format="%")),
            y="REGIÃO:N",
            tooltip=["REGIÃO", "recuperacao_br", "valor_pago_br", "valor_negociado_br"],
            title="% Recuperação por região",
        )

    c3, c4 = st.columns(2)
    with c3:
        bar_chart(
            regiao_chart,
            x="acordos:Q",
            y="REGIÃO:N",
            tooltip=["REGIÃO", "acordos_br", "pagamentos_br", "clientes_br"],
            title="Acordos por região",
        )
    with c4:
        bar_chart(
            regiao_chart,
            x="ticket_medio:Q",
            y="REGIÃO:N",
            tooltip=["REGIÃO", "ticket_medio_br", "acordos_br", "valor_negociado_br"],
            title="Ticket médio por região",
        )

    best_regiao = (
        resultados.groupby(["REGIÃO", "OPERADOR"])
        .agg(
            acordos=("CONTRATO_KEY", "count"),
            pagamentos=("IS_PAGO", "sum"),
            acordos_em_aberto=("IS_EM_ABERTO", "sum"),
            acordos_nao_pagou=("IS_NAO_PAGOU", "sum"),
            valor_pago=("VALOR_PAGO", "sum"),
            valor_em_aberto=("VALOR_EM_ABERTO", "sum"),
            valor_nao_pagou=("VALOR_NAO_PAGOU", "sum"),
            valor_negociado=("VALOR_NEGOCIADO", "sum"),
        )
        .reset_index()
    )
    best_regiao["efetividade_pagamento"] = safe_div(best_regiao["pagamentos"], best_regiao["pagamentos"] + best_regiao["acordos_nao_pagou"])
    best_regiao["valor_quebra"] = best_regiao["valor_nao_pagou"]
    best_regiao["meta_individual"] = operator_goal_series(best_regiao["OPERADOR"], selected_months_count(resultados))
    best_regiao["pct_quebra"] = safe_div(best_regiao["valor_quebra"], best_regiao["meta_individual"])
    best_regiao["recuperacao"] = safe_div(best_regiao["valor_pago"], best_regiao["valor_negociado"])
    best_regiao = best_regiao.sort_values(["REGIÃO", "valor_pago", "recuperacao"], ascending=[True, False, False]).groupby("REGIÃO").head(3)
    st.subheader("Top operadores por região")
    data_table(best_regiao)

with tabs[6]:
    matrix = (
        resultados.groupby(["OPERADOR", "FAIXA_ATRASO", "REGIÃO"])
        .agg(
            acordos=("CONTRATO_KEY", "count"),
            pagamentos=("IS_PAGO", "sum"),
            acordos_em_aberto=("IS_EM_ABERTO", "sum"),
            acordos_nao_pagou=("IS_NAO_PAGOU", "sum"),
            valor_pago=("VALOR_PAGO", "sum"),
            valor_em_aberto=("VALOR_EM_ABERTO", "sum"),
            valor_nao_pagou=("VALOR_NAO_PAGOU", "sum"),
            valor_negociado=("VALOR_NEGOCIADO", "sum"),
        )
        .reset_index()
    )
    matrix["tx_pagamento"] = safe_div(matrix["pagamentos"], matrix["acordos"])
    matrix["efetividade_pagamento"] = safe_div(matrix["pagamentos"], matrix["pagamentos"] + matrix["acordos_nao_pagou"])
    matrix["valor_quebra"] = matrix["valor_nao_pagou"]
    matrix["meta_individual"] = operator_goal_series(matrix["OPERADOR"], selected_months_count(resultados))
    matrix["pct_quebra"] = safe_div(matrix["valor_quebra"], matrix["meta_individual"])
    matrix["recuperacao"] = safe_div(matrix["valor_pago"], matrix["valor_negociado"])

    # ── KPIs ──────────────────────────────────────────────────────────────────
    _mx_combos = int(matrix[matrix["acordos"] > 0][["OPERADOR", "FAIXA_ATRASO"]].drop_duplicates().shape[0])
    _mx_pago_total = matrix["valor_pago"].sum()
    _mx_top5_pct = matrix.sort_values("valor_pago", ascending=False).head(5)["valor_pago"].sum() / _mx_pago_total if _mx_pago_total else 0
    _mx_top_row = matrix.sort_values("valor_pago", ascending=False).iloc[0] if not matrix.empty else None
    _mx_top_label = f"{_mx_top_row['OPERADOR']} / {_mx_top_row['FAIXA_ATRASO']}" if _mx_top_row is not None else "-"
    _mx_top_valor = float(_mx_top_row["valor_pago"]) if _mx_top_row is not None else 0

    km1, km2, km3, km4 = st.columns(4)
    with km1:
        metric_card("Combinações ativas", num_fmt(_mx_combos))
    with km2:
        metric_card("Maior combinação", str(_mx_top_label))
    with km3:
        metric_card("Valor (maior combinação)", money_fmt(_mx_top_valor))
    with km4:
        metric_card("Concentração top 5", pct_fmt(_mx_top5_pct))

    st.markdown("---")

    metric_choice = st.selectbox(
        "Métrica da matriz",
        ["valor_pago", "valor_em_aberto", "valor_nao_pagou", "valor_quebra", "tx_pagamento", "efetividade_pagamento", "pct_quebra", "recuperacao", "acordos", "pagamentos", "acordos_em_aberto", "acordos_nao_pagou"],
        index=0,
    )
    heat_data = matrix.groupby(["OPERADOR", "FAIXA_ATRASO"]).agg(
        {metric_choice: "sum" if metric_choice in ["valor_pago", "valor_em_aberto", "valor_nao_pagou", "valor_quebra", "acordos", "pagamentos", "acordos_em_aberto", "acordos_nao_pagou"] else "mean"}
    ).reset_index()
    heatmap(display_fields(heat_data), "FAIXA_ATRASO:N", "OPERADOR:N", f"{metric_choice}:Q", "Operador x faixa de atraso")

    st.subheader("Top 10 combinações Operador × Faixa")
    top_combos = matrix.sort_values("valor_pago", ascending=False).head(10).copy()
    top_combos["combo"] = top_combos["OPERADOR"].astype(str) + " / " + top_combos["FAIXA_ATRASO"].astype(str)
    top_combos_chart = display_fields(top_combos)
    bar_chart(
        top_combos_chart,
        x="valor_pago:Q",
        y=alt.Y("combo:N", sort="-x", title=None),
        tooltip=["combo", "valor_pago_br", "acordos_br", "pagamentos_br", "tx_pagamento_br", "recuperacao_br"],
        title="Top 10 combinações por valor recebido",
        height=320,
    )

    st.subheader("Matriz analítica por operador, faixa e região")
    data_table(matrix.sort_values(["valor_pago", "pagamentos"], ascending=False))

with tabs[7]:
    st.subheader("Metas e quartis de atingimento")
    st.caption("Meta geral lida da aba METAS da planilha de resultados. Meta mensal: R$ 200.000 por negociador; Victor Lima usa a meta individual de R$ 900.000 referente aos casos de pós retomado.")

    metas_df, meses_meta, meta_geral = build_meta_analysis(operador_df, resultados, operadores_filtrados)
    recebido_meta_geral = resultados["VALOR_PAGO"].sum()
    honorarios_escritorio = resultados["HONORARIOS_ESCRITORIO"].sum()
    valor_aberto_meta_geral = resultados["VALOR_EM_ABERTO"].sum()
    hoje = pd.Timestamp.now(tz="America/Sao_Paulo").tz_localize(None).normalize()
    abertos_hoje = resultados[resultados["IS_EM_ABERTO"] & resultados["DATA_VENCIMENTO"].dt.normalize().eq(hoje)]
    boletos_abertos_hoje = len(abertos_hoje)
    valor_aberto_hoje = abertos_hoje["VALOR_EM_ABERTO"].sum()
    recebidos_hoje = resultados[resultados["IS_PAGO"] & resultados["DATA_PAGAMENTO"].dt.normalize().eq(hoje)]
    boletos_recebidos_hoje = len(recebidos_hoje)
    valor_recebido_hoje = recebidos_hoje["VALOR_PAGO"].sum()
    pagamentos_total = resultados["IS_PAGO"].sum()
    ticket_medio_recebimento = scalar_safe_div(recebido_meta_geral, pagamentos_total)
    pagos_pos_retomada = resultados[resultados["IS_PAGO"] & resultados["CAMPANHA"].map(is_pos_retomado)]
    valor_pos_retomada = pagos_pos_retomada["VALOR_PAGO"].sum()
    pct_acordos_pos_retomada = scalar_safe_div(len(pagos_pos_retomada), pagamentos_total)
    pct_recebido_pos_retomada = scalar_safe_div(valor_pos_retomada, recebido_meta_geral)
    meses_texto = ", ".join(meses_meta) if meses_meta else "Sem mês filtrado"
    meta_relogio_info = business_day_clock_details(meses_meta, hoje)
    meta_relogio_pct = meta_relogio_info["ratio"]
    meta_relogio_valor = meta_geral * meta_relogio_pct
    gap_meta_relogio = max(meta_relogio_valor - recebido_meta_geral, 0)
    dias_uteis_restantes = int(meta_relogio_info["remaining_days"])
    gap_meta_total = max(meta_geral - recebido_meta_geral, 0)
    meta_diaria_necessaria = (
        scalar_safe_div(gap_meta_total, dias_uteis_restantes)
        if dias_uteis_restantes
        else gap_meta_total
    )

    c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
    with c1:
        metric_card("Mês analisado", meses_texto)
    with c2:
        metric_card("Meta geral escritório", money_fmt(meta_geral))
    with c3:
        metric_card("Honorarios escritorio", money_fmt(honorarios_escritorio))
    with c4:
        metric_card("Recebido", money_fmt(recebido_meta_geral))
    with c5:
        metric_card("% meta geral", pct_fmt(recebido_meta_geral / meta_geral if meta_geral else 0))
    with c6:
        metric_card("Em aberto", money_fmt(valor_aberto_meta_geral))
    with c7:
        metric_card("% aberto/meta", pct_fmt(valor_aberto_meta_geral / meta_geral if meta_geral else 0))

    meta_gauge(
        recebido_meta_geral,
        meta_geral,
        meses_texto,
        boletos_abertos_hoje,
        valor_aberto_hoje,
        boletos_recebidos_hoje,
        valor_recebido_hoje,
        ticket_medio_recebimento,
        valor_pos_retomada,
        pct_recebido_pos_retomada,
        pct_acordos_pos_retomada,
        meta_relogio_valor,
        gap_meta_relogio,
        meta_relogio_pct,
        meta_diaria_necessaria,
        dias_uteis_restantes,
        abertos_hoje,
        recebidos_hoje,
    )

    region_meta_df = build_region_goal_map(meses_meta, resultados)
    region_meta_map(region_meta_df, f"% da meta por região — {meses_texto}")

    meta_resumo = metas_df["diagnostico_meta"].value_counts().reset_index()
    meta_resumo.columns = ["Diagnóstico", "Operadores"]
    c1, c2 = st.columns([1, 1.3])
    with c1:
        st.subheader("Resumo por diagnóstico")
        stretch_dataframe(meta_resumo, hide_index=True)
    with c2:
        chart_meta = display_fields(metas_df.sort_values("atingimento_meta_individual", ascending=False).head(15))
        bar_chart(
            chart_meta,
            x=alt.X("atingimento_meta_individual:Q", axis=alt.Axis(format="%")),
            y="OPERADOR:N",
            color="diagnostico_meta:N",
            tooltip=[
                "OPERADOR",
                "valor_pago_br",
                "valor_em_aberto_br",
                "meta_individual_br",
                "atingimento_meta_individual_br",
                "pct_aberto_meta_individual_br",
                "saldo_meta_individual_br",
                "participacao_meta_geral_br",
                "diagnostico_meta",
            ],
            title="Atingimento da meta individual",
            height=360,
        )

    st.subheader("Operadores em cada quartil")
    grupos_meta = meta_operator_groups(metas_df)
    stretch_dataframe(
        grupos_meta,
        hide_index=True,
        column_config={
            "Métrica": st.column_config.TextColumn("Métrica", help="Indicador usado para montar o quartil."),
            "Grupo": st.column_config.TextColumn("Grupo", help="Grupo do quartil. Q4 é destaque; Q1 é crítico."),
            "Qtd. operadores": st.column_config.NumberColumn("Qtd. operadores", help="Quantidade de operadores no grupo."),
            "Operadores": st.column_config.TextColumn("Operadores", help="Operadores classificados nesse grupo."),
        },
    )

    st.subheader("Tabela de metas por operador")
    meta_cols = [
        "OPERADOR",
        "nome_colaborador",
        "base_colaborador",
        "cargo_colaborador",
        "negociador_cadastrado",
        "valor_pago",
        "valor_em_aberto",
        "meta_individual",
        "atingimento_meta_individual",
        "pct_aberto_meta_individual",
        "efetividade_pagamento",
        "pct_quebra",
        "valor_quebra",
        "saldo_meta_individual",
        "meta_geral_escritorio",
        "participacao_meta_geral",
        "quartil_meta_individual",
        "quartil_meta_geral",
        "diagnostico_meta",
        "pagamentos",
        "acordos_nao_pagou",
        "acordos",
    ]
    data_table(metas_df[meta_cols])

with tabs[8]:
    st.subheader("Workplan e priorização futura")
    if workplan_error:
        st.warning(workplan_error)
    elif workplan_df.empty:
        st.info("Sem dados do Workplan para exibir.")
    else:
        st.caption(f"Workplan carregado do banco: {num_fmt(len(workplan_raw))} contratos brutos e {num_fmt(len(workplan_df))} contratos com chave valida.")
        elegivel_df = workplan_df[
            workplan_df["pagamentos_hist"].eq(0)
            & workplan_df["acordos_em_aberto_hist"].eq(0)
            & ~workplan_df["status_base"].map(is_pos_retomado)
            & workplan_df["flag_cobravel"].eq("SIM")
        ].copy()
        base_workplan = elegivel_df

        if base_workplan.empty:
            st.info("Sem contratos cobraveis, sem pagamento historico, sem acordo em aberto e fora de POS RETOMADO para recomendar.")

        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            metric_card("Contratos elegíveis", num_fmt(len(base_workplan)))
        with c2:
            metric_card("Valor em aberto", money_fmt(base_workplan["total_amount_due"].sum()))
        with c3:
            metric_card("Prioridade alta", num_fmt(base_workplan["prioridade_workplan"].eq("Alta").sum()))
        with c4:
            metric_card("Com CPC histórico", num_fmt((base_workplan["cpcs_hist"] > 0).sum()))
        with c5:
            excluidos = (
                workplan_df["pagamentos_hist"].gt(0)
                | workplan_df["acordos_em_aberto_hist"].gt(0)
                | workplan_df["status_base"].map(is_pos_retomado)
                | ~workplan_df["flag_cobravel"].eq("SIM")
            )
            metric_card("Fora da prospeccao", num_fmt(excluidos.sum()))

        prioridade_sel = st.multiselect("Prioridade Workplan", ["Alta", "Média", "Baixa"], default=["Alta", "Média"])
        segmento_sel = st.multiselect("Segmento Workplan", ["POTLOSS", "SALVAGE", "SALVAGE +"])
        fpd_epd_sel = st.multiselect("FPD/EPD", ["Demais", "EPD", "FPD"], default=["Demais", "EPD"])
        workplan_view = base_workplan.copy()
        if prioridade_sel:
            workplan_view = workplan_view[workplan_view["prioridade_workplan"].isin(prioridade_sel)]
        if segmento_sel:
            workplan_view = workplan_view[workplan_view["SEGMENTO_DPD"].isin(segmento_sel)]
        if fpd_epd_sel:
            workplan_view = workplan_view[workplan_view["inadimplencia_precoce_tipo"].isin(fpd_epd_sel)]

        c1, c2 = st.columns(2)
        with c1:
            prioridade_df = (
                workplan_view.groupby("prioridade_workplan")
                .agg(
                    contratos=("CONTRATO_KEY", "nunique"),
                    total_amount_due=("total_amount_due", "sum"),
                    valor_esperado_recuperacao=("valor_esperado_recuperacao", "sum"),
                    probabilidade_recuperacao=("probabilidade_recuperacao", "mean"),
                    score_recuperacao=("score_recuperacao", "mean"),
                )
                .reset_index()
            )
            prioridade_df = display_fields(prioridade_df.rename(columns={"contratos": "clientes"}))
            bar_chart(
                prioridade_df,
                x="total_amount_due:Q",
                y=alt.Y("prioridade_workplan:N", sort=["Alta", "Média", "Baixa"], title=None),
                color="prioridade_workplan:N",
                tooltip=[
                    "prioridade_workplan",
                    "clientes_br",
                    "total_amount_due_br",
                    "valor_esperado_recuperacao_br",
                    "probabilidade_recuperacao_br",
                    "score_recuperacao_br",
                ],
                title="Valor em aberto por prioridade",
                sort=None,
            )
        with c2:
            segmento_workplan = (
                workplan_view.groupby("SEGMENTO_DPD")
                .agg(
                    clientes=("CONTRATO_KEY", "nunique"),
                    total_amount_due=("total_amount_due", "sum"),
                    valor_esperado_recuperacao=("valor_esperado_recuperacao", "sum"),
                    probabilidade_recuperacao=("probabilidade_recuperacao", "mean"),
                    score_recuperacao=("score_recuperacao", "mean"),
                )
                .reset_index()
            )
            segmento_workplan = display_fields(segmento_workplan)
            bar_chart(
                segmento_workplan,
                x="total_amount_due:Q",
                y=alt.Y("SEGMENTO_DPD:N", sort=["POTLOSS", "SALVAGE", "SALVAGE +", "Sem DPD"], title=None),
                tooltip=[
                    "SEGMENTO_DPD",
                    "clientes_br",
                    "total_amount_due_br",
                    "valor_esperado_recuperacao_br",
                    "probabilidade_recuperacao_br",
                    "score_recuperacao_br",
                ],
                title="Valor em aberto por segmento",
                sort=None,
            )

        workplan_analytics_section(workplan_view)

        st.subheader("Recuperação esperada ponderada")
        st.caption("Carteira elegível x taxa de contato x taxa de CPC x taxa de acordo x taxa de pagamento x percentual médio recuperado.")
        projection_dims = {
            "Segmento DPD": ("SEGMENTO_DPD", "SEGMENTO_DPD", "SEGMENTO_DPD"),
            "Faixa de atraso": ("FAIXA_ATRASO", "FAIXA_ATRASO", "FAIXA_ATRASO"),
            "Região": ("REGIÃO", "REGIAO", "REGIÃO"),
            "UF/Estado": ("uf", "UF", "UF"),
            "Segmento + faixa": (
                ["SEGMENTO_DPD", "FAIXA_ATRASO"],
                ["SEGMENTO_DPD", "FAIXA_ATRASO"],
                ["SEGMENTO_DPD", "FAIXA_ATRASO"],
            ),
            "Segmento + região": (
                ["SEGMENTO_DPD", "REGIÃO"],
                ["SEGMENTO_DPD", "REGIAO"],
                ["SEGMENTO_DPD", "REGIÃO"],
            ),
            "Faixa + região": (
                ["FAIXA_ATRASO", "REGIÃO"],
                ["FAIXA_ATRASO", "REGIAO"],
                ["FAIXA_ATRASO", "REGIÃO"],
            ),
            "Segmento + faixa + região": (
                ["SEGMENTO_DPD", "FAIXA_ATRASO", "REGIÃO"],
                ["SEGMENTO_DPD", "FAIXA_ATRASO", "REGIAO"],
                ["SEGMENTO_DPD", "FAIXA_ATRASO", "REGIÃO"],
            ),
        }
        if "PRODUTO" in workplan_view.columns:
            projection_dims["Produto"] = ("PRODUTO", "PRODUTO", "PRODUTO")
            projection_dims["Produto + segmento"] = (
                ["PRODUTO", "SEGMENTO_DPD"],
                ["PRODUTO", "SEGMENTO_DPD"],
                ["PRODUTO", "SEGMENTO_DPD"],
            )
        projection_group = st.selectbox("Agrupamento da projeção", list(projection_dims.keys()))
        projection_df = expected_recovery_by_group(
            workplan_view,
            eventos_raw,
            resultados_raw,
            *projection_dims[projection_group],
        )
        if projection_df.empty:
            st.info("Sem base suficiente para calcular a recuperação esperada nos filtros atuais.")
        else:
            total_carteira_proj = projection_df["carteira_elegivel"].sum()
            total_recuperacao_proj = projection_df["recuperacao_esperada"].sum()
            c1, c2, c3 = st.columns(3)
            with c1:
                metric_card("Carteira elegível projetada", money_fmt(total_carteira_proj))
            with c2:
                metric_card("Recuperação esperada", money_fmt(total_recuperacao_proj))
            with c3:
                metric_card("% esperado da carteira", pct_fmt(total_recuperacao_proj / total_carteira_proj if total_carteira_proj else 0))

            projection_display = display_fields(projection_df)
            bar_chart(
                projection_display.head(12),
                x="recuperacao_esperada:Q",
                y="grupo:N",
                tooltip=[
                    "grupo",
                    "contratos_elegiveis_br",
                    "carteira_elegivel_br",
                    "recuperacao_esperada_br",
                    "recuperacao_esperada_pct_carteira_br",
                    "taxa_contato_br",
                    "taxa_cpc_br",
                    "taxa_acordo_br",
                    "taxa_pagamento_br",
                    "percentual_medio_recuperado_br",
                    "base_taxas",
                ],
                title=f"Recuperação esperada por {projection_group.lower()}",
            )
            projection_cols = [
                "grupo",
                "contratos_elegiveis",
                "carteira_elegivel",
                "recuperacao_esperada",
                "recuperacao_esperada_pct_carteira",
                "taxa_contato",
                "taxa_cpc",
                "taxa_acordo",
                "taxa_pagamento",
                "percentual_medio_recuperado",
                "acionamentos_hist",
                "contatos_cliente_hist",
                "cpcs_hist",
                "acordos_hist",
                "pagamentos_hist",
                "base_taxas",
            ]
            data_table(projection_df[projection_cols])
            projection_workplan_cols = projection_dims[projection_group][0]
            projection_clients = add_projection_group(workplan_view, projection_workplan_cols)
            projection_client_cols = [
                "grupo",
                "agreement_no",
                "cust_name",
                "cpf_cnpj",
                "PRODUTO",
                "prioridade_workplan",
                "score_recuperacao",
                "probabilidade_recuperacao",
                "valor_potencial",
                "valor_esperado_recuperacao",
                "base_probabilidade",
                "contratos_hist",
                "recuperacao_media_perfil",
                "risco_quebra_perfil",
                "motivo_abordagem",
                "inadimplencia_precoce_tipo",
                "no_first_ins_unpaid",
                "SEGMENTO_DPD",
                "FAIXA_ATRASO",
                "REGIÃO",
                "uf",
                "dpd",
                "total_amount_due",
                "dias_sem_contato",
                "acionamentos_hist",
                "cpcs_hist",
                "acordos_hist",
                "pagamentos_hist",
                "status_base",
                "status_cpc",
                "city",
                "state",
            ]
            projection_client_cols = [col for col in projection_client_cols if col in projection_clients.columns]
            st.download_button(
                "Exportar projeção Excel",
                data=dataframe_to_excel_bytes(
                    projection_df[projection_cols],
                    sheet_name="Resumo",
                    extra_sheets={"Clientes": projection_clients[projection_client_cols]},
                ),
                file_name="recuperacao_esperada_workplan.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        st.subheader("Contratos recomendados para novos acordos")
        workplan_cols = [
            "agreement_no",
            "cust_name",
            "cpf_cnpj",
            "PRODUTO",
            "prioridade_workplan",
            "score_recuperacao",
            "probabilidade_recuperacao",
            "valor_potencial",
            "valor_esperado_recuperacao",
            "base_probabilidade",
            "contratos_hist",
            "recuperacao_media_perfil",
            "risco_quebra_perfil",
            "motivo_abordagem",
            "inadimplencia_precoce_tipo",
            "no_first_ins_unpaid",
            "SEGMENTO_DPD",
            "FAIXA_ATRASO",
            "REGIÃO",
            "uf",
            "dpd",
            "total_amount_due",
            "dias_sem_contato",
            "acionamentos_hist",
            "cpcs_hist",
            "acordos_hist",
            "pagamentos_hist",
            "status_base",
            "status_cpc",
            "city",
            "state",
        ]
        workplan_cols = [col for col in workplan_cols if col in workplan_view.columns]
        selected_workplan_cols = st.multiselect(
            "Colunas para visualizar/exportar",
            workplan_cols,
            default=workplan_cols,
            format_func=lambda col: FIELD_HELP.get(col, (col, ""))[0],
        )
        if not selected_workplan_cols:
            st.info("Selecione pelo menos uma coluna para visualizar/exportar.")
        else:
            export_view = workplan_view[selected_workplan_cols].copy()
            st.caption(f"A visualização mostra os 100 primeiros contratos. A exportação usa todos os {num_fmt(len(export_view))} contratos filtrados.")
            data_table(export_view.head(100))
            st.download_button(
                "Exportar Excel",
                data=dataframe_to_excel_bytes(export_view),
                file_name="contratos_recomendados_workplan.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

with tabs[9]:
    st.subheader("Analitica de pagamentos por perfil")
    st.caption("A taxa de pagamento considera todos os acordos do filtro atual: pagamentos divididos por acordos. A efetividade vencida considera apenas pagos e nao pagos, deixando em aberto fora do denominador.")

    payment_profiles = payment_profile_analysis(resultados)
    pagamentos_previstos = total_pagamentos + (total_em_aberto * efetividade_pagamento_geral)
    valor_aberto_esperado = valor_em_aberto * efetividade_pagamento_geral

    p1, p2, p3, p4 = st.columns(4)
    with p1:
        metric_card("Taxa de pagamento", pct_fmt(tx_pagamento_geral), f"{num_fmt(total_pagamentos)} de {num_fmt(total_acordos)} acordos")
    with p2:
        metric_card("Efetividade vencida", pct_fmt(efetividade_pagamento_geral), f"{num_fmt(total_pagamentos)} de {num_fmt(base_quebras)} desfechos")
    with p3:
        metric_card("Pagamentos previstos", f"{pagamentos_previstos:,.1f}".replace(",", "X").replace(".", ",").replace("X", "."), "Pagos atuais + em aberto x efetividade")
    with p4:
        metric_card("Valor aberto esperado", money_fmt(valor_aberto_esperado), "Em aberto x efetividade vencida")

    if payment_profiles.empty:
        st.info("Sem dados de pagamento para os filtros atuais.")
    else:
        dimensoes = payment_profiles["dimensao"].dropna().unique().tolist()
        dimensoes_sel = st.multiselect("Perfilamento", dimensoes, default=dimensoes)
        min_acordos = st.slider("Base minima de acordos por perfil", 1, 50, 5)

        profile_view = payment_profiles[
            payment_profiles["dimensao"].isin(dimensoes_sel)
            & payment_profiles["acordos"].ge(min_acordos)
        ].copy()

        if profile_view.empty:
            st.info("Sem perfis com a base minima selecionada.")
        else:
            profile_view["perfil_completo"] = profile_view["dimensao"] + " | " + profile_view["perfil"].astype(str)
            profile_view["gap_taxa_pagamento"] = profile_view["tx_pagamento"] - tx_pagamento_geral
            profile_display = display_fields(profile_view)

            c1, c2 = st.columns([1.2, 1])
            with c1:
                top_profiles = profile_display.sort_values(["score_eficiencia_pagamento", "valor_pago"], ascending=False).head(15)
                bar_chart(
                    top_profiles,
                    x=alt.X("score_eficiencia_pagamento:Q", axis=alt.Axis(format="%"), title="Score de eficiencia"),
                    y=alt.Y("perfil_completo:N", title=None, sort="-x"),
                    color="dimensao:N",
                    tooltip=[
                        "dimensao",
                        "perfil",
                        "base_confiavel",
                        "acordos_br",
                        "pagamentos_br",
                        "tx_pagamento_br",
                        "efetividade_pagamento_br",
                        "recuperacao_br",
                        "valor_pago_br",
                    ],
                    title="Perfis com maior eficiencia de pagamento",
                    height=420,
                )
            with c2:
                forecast_profiles = profile_display[profile_display["acordos_em_aberto"] > 0].sort_values("valor_esperado_aberto", ascending=False).head(12)
                if forecast_profiles.empty:
                    st.info("Sem acordos em aberto nos perfis filtrados.")
                else:
                    bar_chart(
                        forecast_profiles,
                        x="valor_esperado_aberto:Q",
                        y=alt.Y("perfil_completo:N", title=None, sort="-x"),
                        color="dimensao:N",
                        tooltip=[
                            "dimensao",
                            "perfil",
                            "acordos_em_aberto_br",
                            "efetividade_pagamento_br",
                            "pagamentos_esperados_abertos",
                            "valor_em_aberto_br",
                            "valor_esperado_aberto_br",
                        ],
                        title="Potencial esperado nos acordos em aberto",
                        height=420,
                    )

            dimensao_resumo = (
                profile_view.groupby("dimensao", dropna=False)
                .agg(
                    perfis=("perfil", "count"),
                    acordos=("acordos", "sum"),
                    pagamentos=("pagamentos", "sum"),
                    acordos_em_aberto=("acordos_em_aberto", "sum"),
                    acordos_nao_pagou=("acordos_nao_pagou", "sum"),
                    valor_pago=("valor_pago", "sum"),
                    valor_em_aberto=("valor_em_aberto", "sum"),
                )
                .reset_index()
            )
            dimensao_resumo["tx_pagamento"] = safe_div(dimensao_resumo["pagamentos"], dimensao_resumo["acordos"])
            dimensao_resumo["efetividade_pagamento"] = safe_div(dimensao_resumo["pagamentos"], dimensao_resumo["pagamentos"] + dimensao_resumo["acordos_nao_pagou"])

            st.subheader("Resumo por tipo de perfil")
            data_table(display_fields(dimensao_resumo)[[
                "dimensao",
                "perfis",
                "acordos",
                "pagamentos",
                "acordos_em_aberto",
                "acordos_nao_pagou",
                "tx_pagamento",
                "efetividade_pagamento",
                "valor_pago",
                "valor_em_aberto",
            ]])

            st.subheader("Tabela detalhada dos perfis")
            detail_cols = [
                "dimensao",
                "perfil",
                "base_confiavel",
                "clientes",
                "acordos",
                "pagamentos",
                "acordos_em_aberto",
                "acordos_nao_pagou",
                "tx_pagamento",
                "efetividade_pagamento",
                "recuperacao",
                "score_eficiencia_pagamento",
                "valor_negociado",
                "valor_pago",
                "valor_em_aberto",
                "valor_esperado_aberto",
                "pagamentos_esperados_abertos",
                "pagamentos_previstos_total",
            ]
            data_table(profile_view[detail_cols].head(100))

with tabs[10]:
    avg_score = operador_df["score"].mean() if not operador_df.empty else 0
    oportunidades = operador_df[(operador_df["acionamentos"] >= operador_df["acionamentos"].median()) & (operador_df["score"] < avg_score)].sort_values("acionamentos", ascending=False)
    destaques = operador_df[operador_df["score"] >= operador_df["score"].quantile(0.75)].sort_values("score", ascending=False)
    faixas_oportunidade = aggregate_resultados(resultados, "FAIXA_ATRASO")
    faixas_oportunidade = faixas_oportunidade.sort_values(["valor_negociado", "recuperacao"], ascending=[False, True])
    segmentos_oportunidade = aggregate_resultados(resultados[resultados["SEGMENTO_DPD"].ne("Sem DPD")], "SEGMENTO_DPD")
    segmentos_oportunidade = segmentos_oportunidade.sort_values(["valor_negociado", "recuperacao"], ascending=[False, True])

    # ── KPIs de alerta ──────────────────────────────────────────────────────
    _i_destaques = len(destaques)
    _i_oportunidades = len(oportunidades)
    _i_valor_risco = resultados["VALOR_NAO_PAGOU"].sum()
    _i_valor_aberto = resultados["VALOR_EM_ABERTO"].sum()

    ki1, ki2, ki3, ki4 = st.columns(4)
    with ki1:
        metric_card("Operadores em destaque", num_fmt(_i_destaques), "Score ≥ percentil 75% da equipe")
    with ki2:
        metric_card("Oportunidade de melhoria", num_fmt(_i_oportunidades), "Alto volume de acionamento, score abaixo da média")
    with ki3:
        metric_card("Valor em risco (não pago)", money_fmt(_i_valor_risco), "Acordos com status Não Pagou — quebra de contrato")
    with ki4:
        metric_card("Potencial recuperável (aberto)", money_fmt(_i_valor_aberto), "Acordos em aberto — ainda podem ser convertidos")

    st.markdown("---")

    # ── Seção 1: Operadores em destaque ─────────────────────────────────────
    st.subheader("Operadores em destaque")
    st.caption("Score no percentil 75% ou acima. São as referências de produtividade e eficiência da equipe — candidatos a mentores e benchmarks internos.")
    if not destaques.empty:
        ci1, ci2 = st.columns([1.3, 1])
        with ci1:
            dest_chart = display_fields(destaques.head(10))
            bar_chart(
                dest_chart,
                x=alt.X("score:Q", axis=alt.Axis(format="%")),
                y=alt.Y("OPERADOR:N", sort="-x", title=None),
                tooltip=["OPERADOR", "score_br", "valor_pago_br", "tx_pagamento_br", "efetividade_pagamento_br", "recuperacao_br"],
                title="Score dos operadores em destaque",
                height=300,
            )
        with ci2:
            data_table(destaques[["OPERADOR", "score", "valor_pago", "tx_pagamento", "efetividade_pagamento", "recuperacao", "clientes_trabalhados"]].head(10))

    # ── Seção 2: Operadores com oportunidade ────────────────────────────────
    st.subheader("Alto volume com eficiência abaixo da média")
    st.caption("Operadores com volume de acionamentos ≥ mediana da equipe, mas score abaixo da média. Têm carteira ativa mas precisam melhorar conversão — foco prioritário de coaching.")
    if not oportunidades.empty:
        ci3, ci4 = st.columns([1.3, 1])
        with ci3:
            op_chart = display_fields(oportunidades.head(10))
            bar_chart(
                op_chart,
                x="acionamentos:Q",
                y=alt.Y("OPERADOR:N", sort="-x", title=None),
                tooltip=["OPERADOR", "acionamentos_br", "score_br", "tx_acordo_br", "tx_pagamento_br", "valor_pago_br"],
                title="Volume de acionamentos (oportunidade de conversão)",
                height=300,
            )
        with ci4:
            data_table(oportunidades[["OPERADOR", "score", "acionamentos", "tx_contato", "tx_acordo", "valor_pago"]].head(10))

    st.markdown("---")

    # ── Seção 3: Mapa de oportunidades por faixa ────────────────────────────
    st.subheader("Mapa de oportunidades por faixa de atraso")
    st.caption("Faixas ordenadas por maior carteira negociada com menor recuperação proporcional. Onde há mais valor em disputa e menor aproveitamento — ação imediata gera maior retorno.")
    ci5, ci6 = st.columns(2)
    with ci5:
        fo_chart = display_fields(faixas_oportunidade)
        bar_chart(
            fo_chart,
            x="valor_negociado:Q",
            y=alt.Y("FAIXA_ATRASO:N", sort="-x", title=None),
            tooltip=["FAIXA_ATRASO", "valor_negociado_br", "valor_pago_br", "recuperacao_br", "acordos_br"],
            title="Carteira negociada por faixa",
            sort=None,
        )
    with ci6:
        bar_chart(
            fo_chart,
            x=alt.X("recuperacao:Q", axis=alt.Axis(format="%")),
            y=alt.Y("FAIXA_ATRASO:N", sort="-x", title=None),
            tooltip=["FAIXA_ATRASO", "recuperacao_br", "valor_pago_br", "valor_negociado_br", "efetividade_pagamento_br"],
            title="% Recuperação por faixa",
            sort=None,
        )
    data_table(faixas_oportunidade[["FAIXA_ATRASO", "clientes", "acordos", "pagamentos", "acordos_em_aberto", "acordos_nao_pagou", "pct_quebra", "efetividade_pagamento", "valor_negociado", "valor_pago", "valor_em_aberto", "valor_nao_pagou", "valor_quebra", "recuperacao"]])

    # ── Seção 4: Oportunidade por segmento DPD ──────────────────────────────
    st.subheader("Oportunidade por segmento DPD")
    st.caption("POTLOSS: dívidas recentes, maior liquidez e probabilidade de acordo. SALVAGE e SALVAGE+: dívidas antigas, maior desconto necessário mas ticket potencialmente alto.")
    data_table(segmentos_oportunidade[["SEGMENTO_DPD", "clientes", "acordos", "pagamentos", "acordos_em_aberto", "acordos_nao_pagou", "pct_quebra", "efetividade_pagamento", "valor_negociado", "valor_pago", "valor_em_aberto", "valor_nao_pagou", "valor_quebra", "recuperacao"]])
